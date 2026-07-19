"""
parser.py — Sheet detection, header-row finding, and raw data extraction.

Handles any GP Excel format by scoring sheets and sniffing the header row.
"""

from __future__ import annotations

import re
import pandas as pd
import openpyxl
from datetime import datetime, date


DEAL_SHEET_KEYWORDS = [
    "portfolio", "workbook", "pwb", "deals", "investments",
    "track", "record", "transaction", "investment", "schedule",
]

# Keywords that strongly suggest a sheet is NOT deal-level (skip it)
_SKIP_KEYWORDS = ["cover", "contents", "toc", "instructions", "legend", "glossary",
                  "notes", "summary", "overview", "performance", "irr", "returns"]


def _sheet_score(name: str) -> int:
    """Score a sheet name by how likely it contains deal-level data."""
    lname = name.lower()
    if re.match(r"^sheet\d*$", lname):
        return 0
    score = sum(2 for kw in DEAL_SHEET_KEYWORDS if kw in lname)
    score -= sum(1 for kw in _SKIP_KEYWORDS if kw in lname)
    score += 1  # baseline for having a non-generic name
    return score


def detect_main_sheet(wb: openpyxl.Workbook) -> str:
    """Return the single sheet most likely to contain deal-level data."""
    scores = {name: _sheet_score(name) for name in wb.sheetnames}
    best = max(scores, key=scores.get)
    if scores[best] <= 0 and len(wb.sheetnames) > 1:
        return wb.sheetnames[-1]
    return best


def detect_deal_sheets(wb: openpyxl.Workbook) -> list[str]:
    """Return ALL sheets that may contain deal data, ordered by score descending."""
    scored = sorted(wb.sheetnames, key=_sheet_score, reverse=True)
    # Include sheets with score > 0; if none qualify, fall back to highest scorer
    positive = [s for s in scored if _sheet_score(s) > 0]
    return positive if positive else [scored[0]]


def detect_header_row(ws, max_search: int = 30) -> int:
    """Return the 1-based row index that has the most non-empty cells (the header row)."""
    best_row, best_count = 1, 0
    for i, row in enumerate(ws.iter_rows(max_row=max_search, values_only=True)):
        count = sum(1 for v in row if v is not None and str(v).strip() not in ("", "-"))
        if count > best_count:
            best_count = count
            best_row = i + 1
    return best_row


def _dedup_headers(headers: list[str]) -> list[str]:
    """
    Make column names unique:
    - Blank headers → _col<position>
    - Duplicate names → append .2, .3, … suffix
    """
    seen: dict[str, int] = {}
    result: list[str] = []
    for i, h in enumerate(headers):
        h = h.strip()
        if not h:
            h = f"_col{i + 1}"
        if h in seen:
            seen[h] += 1
            result.append(f"{h}.{seen[h]}")
        else:
            seen[h] = 0
            result.append(h)
    return result


def normalize_col_name(name: str) -> str:
    """Lowercase, strip footnote markers, collapse whitespace/underscores."""
    if not name:
        return ""
    s = str(name)
    # Remove footnote markers like (a), (b)(c), etc.
    s = re.sub(r"\([a-z]+\)", "", s)
    s = re.sub(r"\([0-9]+\)", "", s)
    # Replace newlines and multiple spaces
    s = re.sub(r"[\n\r]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s.lower()


def _as_date(v) -> date | None:
    """Coerce various date representations to a date object."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    s = str(v).strip()
    if not s or s in ("-", "n/a", "N/A", ""):
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            pass
    return None


def _safe_float(v) -> float | None:
    """Coerce a value to float, return None on failure or if dash/blank."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("-", "", "n/a", "N/A", "#N/A", "#VALUE!", "#REF!"):
        return None
    try:
        return float(s.replace(",", "").replace("%", ""))
    except (ValueError, TypeError):
        return None


def _parse_single_sheet(
    wb,
    chosen: str,
    header_rows: list[int] | None = None,
    first_data_row: int | None = None,
) -> tuple[pd.DataFrame, dict]:
    """
    Parse one worksheet and return (df, meta).

    Parameters
    ----------
    header_rows   : 1-based row indices of header row(s). When multiple rows are
                    given (multi-row header), their values are concatenated per column.
                    If None, auto-detected via detect_header_row.
    first_data_row: 1-based row where data begins. Defaults to max(header_rows)+1.
    """
    ws = wb[chosen]

    # ── Resolve header / data row positions ──────────────────────────
    if header_rows is None:
        detected = detect_header_row(ws, max_search=30)
        header_rows = [detected]
    if first_data_row is None:
        first_data_row = max(header_rows) + 1

    # ── Determine column count from the first header row ─────────────
    n_cols = 0
    for row_vals in ws.iter_rows(min_row=header_rows[0], max_row=header_rows[0], values_only=True):
        n_cols = len(row_vals)
        break

    if n_cols == 0:
        return pd.DataFrame(), {
            "gp_name": "", "report_date": None, "currency": "USD",
            "raw_sheet_name": chosen, "header_rows": header_rows,
            "first_data_row": first_data_row, "raw_headers": [],
        }

    # ── Build composite headers across all header rows ───────────────
    # For multi-row headers (e.g. group label in row 5, column names in row 6):
    #   col_parts[i] = ["Investment Details", "Company Name"]
    #   → header = "Investment Details | Company Name"
    # Consecutive duplicate parts (from merged cells propagating) are suppressed.
    col_parts: list[list[str]] = [[] for _ in range(n_cols)]
    for hrow in header_rows:
        for row_vals in ws.iter_rows(min_row=hrow, max_row=hrow, values_only=True):
            for i, v in enumerate(row_vals[:n_cols]):
                if v is not None:
                    sv = str(v).strip()
                    if sv and sv not in ("-",):
                        if not col_parts[i] or col_parts[i][-1] != sv:
                            col_parts[i].append(sv)
            break

    raw_headers = [" | ".join(parts) if parts else "" for parts in col_parts]
    raw_headers = _dedup_headers(raw_headers)

    # ── Read data rows ───────────────────────────────────────────────
    data_rows = []
    for row_vals in ws.iter_rows(min_row=first_data_row, values_only=True):
        vals = list(row_vals)
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        data_rows.append(vals[:len(raw_headers)])

    df = pd.DataFrame(data_rows, columns=raw_headers).dropna(how="all")

    # ── Metadata from rows above the header ──────────────────────────
    meta_max_row = min(header_rows[0] - 1, 10) if header_rows[0] > 1 else 0
    gp_name = ""
    report_date = None
    if meta_max_row >= 1:
        for row_vals in ws.iter_rows(max_row=meta_max_row, values_only=True):
            for v in row_vals:
                if v is None:
                    continue
                sv = str(v).strip()
                if report_date is None:
                    d = _as_date(v)
                    if d and d.year > 1990:
                        report_date = d
                if not gp_name and len(sv) > 3 and _as_date(v) is None:
                    name = re.split(r"\s+(equity|portfolio|workbook|track|record|data)", sv, flags=re.I)[0]
                    gp_name = name.strip()

    # ── Currency from a dedicated column ────────────────────────────
    currency = "USD"
    for col in raw_headers:
        if "currency" in col.lower() and not col.startswith("_col"):
            col_vals = df[col].dropna().unique()
            if len(col_vals) > 0:
                currency = str(col_vals[0]).strip()
                break

    meta = {
        "gp_name": gp_name,
        "report_date": report_date,
        "currency": currency,
        "raw_sheet_name": chosen,
        "header_rows": header_rows,
        "first_data_row": first_data_row,
        "raw_headers": raw_headers,
    }
    return df, meta


def parse_gp_file(
    file_bytes: bytes,
    sheet_name: str | list[str] | None = None,
    table_hint: dict | None = None,
) -> tuple[pd.DataFrame, dict]:
    """
    Parse a GP raw Excel file.

    Parameters
    ----------
    file_bytes  : bytes
        Raw Excel file bytes.
    sheet_name  : str | list[str] | None
        Explicit sheet(s) to parse.  When given, overrides table_hint.
        None → use table_hint or auto-detect.
    table_hint  : dict | None
        Profiler-derived hints for guided parsing.  Keys:
          "sheet_name"    : str         — sheet to open
          "header_rows"   : list[int]   — 1-based header row numbers
          "first_data_row": int         — 1-based first data row

        When sheet_name is None and table_hint is supplied, the hint's
        sheet is used and header_rows / first_data_row skip auto-detection.
    """
    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    all_sheet_names = list(wb.sheetnames)

    # ── Resolve which sheets to parse ────────────────────────────────
    # Priority: explicit sheet_name > table_hint.sheet_name > auto-detect
    hint_header_rows: list[int] | None = None
    hint_first_data_row: int | None    = None

    if sheet_name is not None:
        # Explicit override — ignore table_hint
        if isinstance(sheet_name, str):
            chosen_list = [sheet_name] if sheet_name in all_sheet_names else [detect_main_sheet(wb)]
        else:
            chosen_list = [s for s in sheet_name if s in all_sheet_names] or [detect_main_sheet(wb)]
    elif table_hint and table_hint.get("sheet_name") in all_sheet_names:
        chosen_list        = [table_hint["sheet_name"]]
        hint_header_rows   = table_hint.get("header_rows")
        hint_first_data_row = table_hint.get("first_data_row")
    else:
        chosen_list = [detect_main_sheet(wb)]

    # ── Parse each sheet ─────────────────────────────────────────────
    frames: list[pd.DataFrame] = []
    primary_meta: dict = {}

    for i, chosen in enumerate(chosen_list):
        # Only apply hints to the primary sheet (not when combining multiple)
        h_rows = hint_header_rows   if i == 0 else None
        h_first = hint_first_data_row if i == 0 else None

        df_s, meta_s = _parse_single_sheet(
            wb, chosen,
            header_rows=h_rows,
            first_data_row=h_first,
        )
        if len(chosen_list) > 1:
            df_s["_source_sheet"] = chosen
        frames.append(df_s)
        if i == 0:
            primary_meta = meta_s

    wb.close()

    combined = frames[0] if len(frames) == 1 else pd.concat(frames, ignore_index=True, sort=False)

    primary_meta["all_sheets"]    = all_sheet_names
    primary_meta["parsed_sheets"] = chosen_list
    primary_meta["raw_sheet_name"] = (
        chosen_list[0] if len(chosen_list) == 1 else ", ".join(chosen_list)
    )

    return combined, primary_meta


def detect_funds_in_df(df: pd.DataFrame, fund_col: str) -> list[str]:
    """Return unique fund names in the order they first appear."""
    if fund_col not in df.columns:
        return []
    seen, result = set(), []
    for v in df[fund_col].dropna():
        s = str(v).strip()
        if s and s not in seen:
            seen.add(s)
            result.append(s)
    return result
