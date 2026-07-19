"""
profiler.py — Stage 1: Workbook Profiling

Inspects a GP Excel workbook before any parsing or field mapping.
Returns a WorkbookProfile describing every sheet's structure:
  - All worksheets with visibility state
  - Used ranges and empty regions
  - Merged cell regions
  - Candidate tables with bounds, header rows, orientation
  - Multi-row header detection
  - Table orientation (row-per-record vs col-per-record)
  - Workbook-level metadata (GP name candidates, report date, currency)

No assumptions are made about the GP or template format.
"""

from __future__ import annotations

import io
import math
import re
from dataclasses import dataclass, asdict, field
from datetime import date, datetime
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════════
# Data Structures
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class CellRange:
    """Inclusive 1-based (row, col) bounds."""
    min_row: int
    max_row: int
    min_col: int
    max_col: int

    @property
    def row_count(self) -> int:
        return self.max_row - self.min_row + 1

    @property
    def col_count(self) -> int:
        return self.max_col - self.min_col + 1

    @property
    def area(self) -> int:
        return self.row_count * self.col_count

    def __str__(self) -> str:
        return (f"{get_column_letter(self.min_col)}{self.min_row}:"
                f"{get_column_letter(self.max_col)}{self.max_row}")


@dataclass
class MergedRegion:
    """A merged cell group with its display value."""
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    value: str        # value of the top-left cell

    @property
    def spans_rows(self) -> bool:
        return self.max_row > self.min_row

    @property
    def spans_cols(self) -> bool:
        return self.max_col > self.min_col

    def __str__(self) -> str:
        return (f"{get_column_letter(self.min_col)}{self.min_row}:"
                f"{get_column_letter(self.max_col)}{self.max_row}"
                f" = {self.value!r}")


@dataclass
class TableCandidate:
    """
    A rectangular region that looks like a structured data table.

    Attributes
    ----------
    sheet_name        : name of the containing sheet
    bounds            : full table bounds (includes header rows)
    header_rows       : 1-based absolute row numbers identified as header(s)
    first_data_row    : first non-header row (1-based)
    last_data_row     : last data row (1-based)
    orientation       : "row_per_record" (normal) | "col_per_record" (transposed)
    score             : 0–100 composite quality score
    sample_headers    : first ≤15 detected header strings
    row_count         : number of data rows (excluding headers)
    col_count         : number of columns
    density           : fraction of non-empty cells in the data region
    has_date_col      : at least one column contains dates
    has_numeric_col   : at least one column contains numeric values
    likely_deal_table : header keywords match PE deal terminology
    notes             : human-readable observations
    """
    sheet_name: str
    bounds: CellRange
    header_rows: list[int]
    first_data_row: int
    last_data_row: int
    orientation: str
    score: float
    sample_headers: list[str]
    row_count: int
    col_count: int
    density: float
    has_date_col: bool
    has_numeric_col: bool
    likely_deal_table: bool
    notes: list[str]

    def summary(self) -> str:
        hr = ",".join(str(r) for r in self.header_rows)
        return (f"[{self.sheet_name}] {self.bounds} | "
                f"hdr rows: {hr} | "
                f"{self.row_count} data rows × {self.col_count} cols | "
                f"score {self.score:.0f} | {self.orientation}")


@dataclass
class SheetProfile:
    name: str
    index: int              # 0-based position in workbook
    state: str              # "visible" | "hidden" | "veryHidden"
    used_range: Optional[CellRange]
    merged_regions: list[MergedRegion]
    candidate_tables: list[TableCandidate]
    empty_fraction: float   # fraction of used_range cells that are empty
    notes: list[str]


@dataclass
class WorkbookProfile:
    filename: str
    sheet_profiles: list[SheetProfile]
    primary_table: Optional[TableCandidate]   # highest-scoring candidate
    gp_name_candidates: list[str]             # extracted from top rows
    report_date: Optional[date]
    currency: Optional[str]
    warnings: list[str]

    # ── Convenience accessors ──────────────────────────────────────────

    def best_sheet(self) -> Optional[str]:
        """Name of the sheet containing the primary deal table."""
        return self.primary_table.sheet_name if self.primary_table else None

    def visible_sheets(self) -> list[str]:
        return [sp.name for sp in self.sheet_profiles if sp.state == "visible"]

    def all_tables(self) -> list[TableCandidate]:
        return [t for sp in self.sheet_profiles for t in sp.candidate_tables]

    def to_dict(self) -> dict:
        """JSON-serialisable dict (converts date objects to ISO strings)."""
        import json
        def _coerce(obj):
            if isinstance(obj, (date, datetime)):
                return obj.isoformat()
            return str(obj)
        return json.loads(json.dumps(asdict(self), default=_coerce))


# ═══════════════════════════════════════════════════════════════════════════════
# Configuration
# ═══════════════════════════════════════════════════════════════════════════════

_MAX_SCAN_ROWS   = 2000    # scan cap per sheet (rows)
_MAX_SCAN_COLS   = 300     # scan cap per sheet (columns)
_MIN_TABLE_ROWS  = 3       # fewer data rows → not a deal table
_MIN_TABLE_COLS  = 2       # fewer cols → skip
_DENSITY_THRESH  = 0.20    # ≥20% of cols filled → "occupied" row
_MAX_BLANK_GAP   = 4       # consecutive blank rows allowed inside a table block
_MAX_HDR_ROWS    = 5       # look at most this many rows for header detection

# PE / deal-level terminology — matched against detected headers
_DEAL_KEYWORDS: set[str] = {
    "company", "investment", "portfolio", "fund", "deal", "asset",
    "entry", "exit", "invested", "realized", "unrealised", "unrealized",
    "moic", "irr", "ebitda", "revenue", "sales", "enterprise", "equity",
    "sector", "geography", "region", "country", "status", "hold",
    "vintage", "date", "return", "multiple", "gross", "net", "value",
    "capital", "cost", "debt", "proceeds", "write", "acquisition",
}

_CURRENCY_RE = re.compile(
    r'\b(USD|EUR|GBP|CHF|JPY|CAD|AUD|SEK|NOK|DKK|CNY|INR|SGD|HKD)\b',
    re.IGNORECASE,
)
_DATE_MIN_YEAR = 1990


# ═══════════════════════════════════════════════════════════════════════════════
# Table grouping — consolidated vs per-fund layout detection
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class TableGroup:
    """
    A set of deal tables that should be read together.

    layout == "consolidated": one sheet holds every deal (group has 1 table).
    layout == "per_fund":      deals are sharded across sibling tabs that share
                               the same schema (group has ≥2 tables) and must be
                               concatenated to recover the full deal set.
    """
    tables: list[TableCandidate]
    layout: str   # "consolidated" | "per_fund"

    def sheet_names(self) -> list[str]:
        return [t.sheet_name for t in self.tables]


def _header_tokens(headers: list[str]) -> set[str]:
    """Normalise header labels into a bag of comparable tokens."""
    toks: set[str] = set()
    for h in headers:
        for tok in re.split(r"[^a-z0-9]+", str(h).lower()):
            if len(tok) >= 2:
                toks.add(tok)
    return toks


def _schemas_match(
    a: TableCandidate, b: TableCandidate,
    col_ratio_min: float, jaccard_min: float,
) -> bool:
    """
    True when two candidate tables look like the *same kind* of table.

    Two gates, both must pass:
      1. Column-count proximity — same template ⇒ near-identical width. This is
         what separates a wide consolidated sheet (e.g. 80 cols) from narrow
         per-fund summary/metric tabs (27–48 cols).
      2. Header token overlap (Jaccard) — guards against two unrelated tables
         that happen to have a similar column count.
    """
    ca, cb = a.col_count, b.col_count
    if max(ca, cb) == 0:
        return False
    if min(ca, cb) / max(ca, cb) < col_ratio_min:
        return False
    ta, tb = _header_tokens(a.sample_headers), _header_tokens(b.sample_headers)
    if not ta or not tb:
        return False
    jac = len(ta & tb) / len(ta | tb)
    return jac >= jaccard_min


def group_candidate_tables(
    profile: WorkbookProfile,
    col_ratio_min: float = 0.75,
    jaccard_min: float = 0.5,
) -> Optional[TableGroup]:
    """
    Decide whether the workbook is consolidated (one deal sheet) or per-fund
    (deals sharded across same-schema sibling tabs).

    Strategy: take the profiler's primary table and gather the best deal-like
    candidate from every *other* sheet that shares its schema. If any siblings
    match, the layout is per-fund and all matching tables form the group;
    otherwise the primary stands alone (consolidated).

    Returns None when there is no primary table at all.
    """
    primary = profile.primary_table
    if primary is None:
        return None

    # Best deal-like candidate per sheet (one table per sheet avoids counting
    # stacked sub-tables on the same tab as separate shards).
    best_per_sheet: list[TableCandidate] = []
    for sp in profile.sheet_profiles:
        deal_cands = [t for t in sp.candidate_tables if t.likely_deal_table]
        if deal_cands:
            best_per_sheet.append(max(deal_cands, key=lambda t: t.score))

    siblings = [
        t for t in best_per_sheet
        if t.sheet_name != primary.sheet_name
        and _schemas_match(primary, t, col_ratio_min, jaccard_min)
    ]

    if siblings:
        group = [primary] + siblings
        # Stable, readable order: by sheet position in the workbook.
        order = {sp.name: sp.index for sp in profile.sheet_profiles}
        group.sort(key=lambda t: order.get(t.sheet_name, 0))
        return TableGroup(tables=group, layout="per_fund")

    return TableGroup(tables=[primary], layout="consolidated")


# ═══════════════════════════════════════════════════════════════════════════════
# Public API
# ═══════════════════════════════════════════════════════════════════════════════

def profile_workbook(file_bytes: bytes, filename: str = "upload.xlsx") -> WorkbookProfile:
    """
    Profile a GP Excel workbook.

    Loads the workbook, inspects every sheet, detects tables and metadata,
    and returns a structured WorkbookProfile.

    Parameters
    ----------
    file_bytes : bytes   Raw Excel file bytes
    filename   : str     Original filename (for display only, not parsed)

    Returns
    -------
    WorkbookProfile
    """
    warnings: list[str] = []

    # Non-read_only so merged_cells.ranges is available.
    # data_only=True resolves formula values (where Excel has cached them).
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        msg = str(exc)
        # Detect legacy .xls (binary) format — openpyxl only supports .xlsx
        if "not a zip" in msg.lower() or filename.lower().endswith(".xls") and not filename.lower().endswith(".xlsx"):
            msg = (
                f"Legacy .xls format is not supported. "
                f"Please open the file in Excel and save it as .xlsx, then re-upload."
            )
        return WorkbookProfile(
            filename=filename, sheet_profiles=[], primary_table=None,
            gp_name_candidates=[], report_date=None, currency=None,
            warnings=[msg],
        )

    sheet_profiles: list[SheetProfile] = []
    gp_name_candidates: list[str] = []
    report_date: Optional[date] = None
    currency: Optional[str] = None

    for idx, name in enumerate(wb.sheetnames):
        try:
            ws    = wb[name]
            state = _sheet_state(wb, name)
            sp    = _profile_sheet(ws, name, idx, state)
            sheet_profiles.append(sp)

            # Extract workbook-level metadata from the first visible sheet
            if state == "visible":
                gp_cands, rdate, cur = _extract_meta(ws)
                gp_name_candidates.extend(gp_cands)
                if report_date is None:
                    report_date = rdate
                if currency is None:
                    currency = cur

        except Exception as exc:
            warnings.append(f"Sheet '{name}': {exc}")
            sheet_profiles.append(SheetProfile(
                name=name, index=idx, state="unknown",
                used_range=None, merged_regions=[], candidate_tables=[],
                empty_fraction=1.0, notes=[f"Profiling error: {exc}"],
            ))

    wb.close()

    # The filename is the most reliable GP-name signal (raw files follow a
    # "<GP Name> <descriptor>…" convention); put it at the front of candidates.
    fn_gp = _gp_from_filename(filename)
    ordered = ([fn_gp] if fn_gp else []) + gp_name_candidates

    # Deduplicate GP name candidates (preserve first-seen order)
    seen: set[str] = set()
    unique_gp: list[str] = [
        c for c in ordered if c and not (c in seen or seen.add(c))  # type: ignore[func-returns-value]
    ]

    all_tables = [t for sp in sheet_profiles for t in sp.candidate_tables]
    primary    = max(all_tables, key=lambda t: t.score) if all_tables else None

    return WorkbookProfile(
        filename=filename,
        sheet_profiles=sheet_profiles,
        primary_table=primary,
        gp_name_candidates=unique_gp[:5],
        report_date=report_date,
        currency=currency,
        warnings=warnings,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet Profiling
# ═══════════════════════════════════════════════════════════════════════════════

def _sheet_state(wb, name: str) -> str:
    state = getattr(wb[name], "sheet_state", "visible")
    return state if state in ("hidden", "veryHidden") else "visible"


def _profile_sheet(ws, name: str, idx: int, state: str) -> SheetProfile:
    notes: list[str] = []

    # ── Sample sheet content ──────────────────────────────────────────
    grid, r_min, r_max, c_min, c_max = _sample_sheet(ws)

    if not grid:
        return SheetProfile(
            name=name, index=idx, state=state,
            used_range=None, merged_regions=[], candidate_tables=[],
            empty_fraction=1.0, notes=["Sheet is empty"],
        )

    used_range   = CellRange(r_min, r_max, c_min, c_max)
    empty_frac   = 1.0 - len(grid) / max(used_range.area, 1)
    merged       = _get_merged_regions(ws, grid)
    tables       = _detect_tables(grid, used_range, merged, name)

    tables.sort(key=lambda t: t.score, reverse=True)

    if not tables:
        notes.append("No candidate data tables detected")
    if state != "visible":
        notes.append(f"Sheet is {state}")
    if empty_frac > 0.85:
        notes.append(f"Sparse sheet ({empty_frac:.0%} empty)")

    return SheetProfile(
        name=name, index=idx, state=state,
        used_range=used_range,
        merged_regions=merged,
        candidate_tables=tables,
        empty_fraction=empty_frac,
        notes=notes,
    )


# ── Grid sampling ──────────────────────────────────────────────────────────────

def _sample_sheet(ws):
    """
    Read cell values into a sparse dict {(row, col): value}.

    Skips completely blank cells and cells containing only dash/hyphen.
    Returns (grid, min_row, max_row, min_col, max_col).
    """
    grid: dict[tuple[int, int], Any] = {}
    r_min = c_min = 10 ** 9
    r_max = c_max = 0

    for row in ws.iter_rows(max_row=_MAX_SCAN_ROWS, max_col=_MAX_SCAN_COLS):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            sv = str(v).strip()
            if not sv or sv in ("-", "–", "—", "N/A", "n/a", "#N/A"):
                continue
            r, c = cell.row, cell.column
            grid[(r, c)] = v
            r_min = min(r_min, r); r_max = max(r_max, r)
            c_min = min(c_min, c); c_max = max(c_max, c)

    if not grid:
        return {}, None, None, None, None
    return grid, r_min, r_max, c_min, c_max


# ── Merged cells ───────────────────────────────────────────────────────────────

def _get_merged_regions(ws, grid: dict) -> list[MergedRegion]:
    regions: list[MergedRegion] = []
    try:
        for mr in ws.merged_cells.ranges:
            val = grid.get((mr.min_row, mr.min_col))
            regions.append(MergedRegion(
                min_row=mr.min_row, max_row=mr.max_row,
                min_col=mr.min_col, max_col=mr.max_col,
                value=str(val).strip() if val is not None else "",
            ))
    except Exception:
        pass
    return regions


# ═══════════════════════════════════════════════════════════════════════════════
# Table Detection
# ═══════════════════════════════════════════════════════════════════════════════

def _detect_tables(
    grid: dict,
    used_range: CellRange,
    merged: list[MergedRegion],
    sheet_name: str,
) -> list[TableCandidate]:
    """
    Locate candidate tables by building a row-density profile and finding
    "dense blocks" — runs of rows where a meaningful fraction of columns
    have values.  Each block is analysed for headers, orientation, and scored.
    """
    if not grid or used_range is None:
        return []

    col_min, col_max = used_range.min_col, used_range.max_col
    total_cols = max(col_max - col_min + 1, 1)

    # Adaptive density threshold: wide ILPA-style sheets (100+ columns) have
    # many optional columns left blank per row, naturally dropping row density.
    # Scale down so a row needs at least ~12 filled cells to qualify; never
    # go below 0.12 (avoids noise in truly sparse sheets).
    density_thresh = max(0.12, min(_DENSITY_THRESH, 12 / max(total_cols, 60)))

    # ── Row density profile ──────────────────────────────────────────
    row_density: dict[int, float] = {}
    for r in range(used_range.min_row, min(used_range.max_row + 1, _MAX_SCAN_ROWS + 1)):
        filled = sum(1 for c in range(col_min, col_max + 1) if (r, c) in grid)
        row_density[r] = filled / total_cols

    # ── Find dense blocks (tolerating short blank gaps) ──────────────
    blocks: list[tuple[int, int]] = []
    block_start: Optional[int]    = None
    last_dense_row: Optional[int] = None
    blank_run   = 0

    for r in range(used_range.min_row, used_range.max_row + 1):
        dense = row_density.get(r, 0) >= density_thresh
        if dense:
            if block_start is None:
                block_start = r
            last_dense_row = r
            blank_run = 0
        else:
            if block_start is not None:
                blank_run += 1
                if blank_run > _MAX_BLANK_GAP:
                    # Close the block at the last dense row seen
                    if last_dense_row is not None and last_dense_row >= block_start:
                        blocks.append((block_start, last_dense_row))
                    block_start    = None
                    last_dense_row = None
                    blank_run      = 0

    # Close any block that extends to the end of the sheet
    # (the old +2 sentinel approach missed this case).
    if block_start is not None and last_dense_row is not None and last_dense_row >= block_start:
        blocks.append((block_start, last_dense_row))

    # ── Analyse each block ───────────────────────────────────────────
    candidates: list[TableCandidate] = []

    for (b_start, b_end) in blocks:
        # Trim trailing sparse rows
        while b_end > b_start and row_density.get(b_end, 0) < density_thresh:
            b_end -= 1
        if b_end - b_start + 1 < _MIN_TABLE_ROWS:
            continue

        # Column extent of this block specifically
        block_cells = {(r, c): v for (r, c), v in grid.items()
                       if b_start <= r <= b_end}
        if not block_cells:
            continue
        b_col_min = min(c for _, c in block_cells)
        b_col_max = max(c for _, c in block_cells)
        if b_col_max - b_col_min + 1 < _MIN_TABLE_COLS:
            continue

        table_bounds = CellRange(b_start, b_end, b_col_min, b_col_max)

        # ── Header detection ─────────────────────────────────────────
        header_rows = _detect_header_rows(block_cells, table_bounds)
        first_data_row = max(header_rows) + 1
        if first_data_row > b_end:
            continue

        # ── Orientation detection ────────────────────────────────────
        orientation = _detect_orientation(block_cells, table_bounds, header_rows)

        # ── Data-region metrics ──────────────────────────────────────
        data_cells = {k: v for k, v in block_cells.items() if k[0] >= first_data_row}
        data_area  = max((b_end - first_data_row + 1) * table_bounds.col_count, 1)
        density    = len(data_cells) / data_area
        has_date   = _has_date_column(data_cells)
        has_num    = _has_numeric_column(data_cells)

        # ── Sample headers ───────────────────────────────────────────
        sample_hdrs = _extract_headers(
            block_cells, header_rows, b_col_min, b_col_max, merged
        )

        # ── Scoring ──────────────────────────────────────────────────
        row_count = b_end - first_data_row + 1
        col_count = table_bounds.col_count
        score = _score_table(
            row_count=row_count, col_count=col_count, density=density,
            sample_headers=sample_hdrs, has_date=has_date, has_num=has_num,
            orientation=orientation,
        )

        likely_deal = _is_likely_deal_table(sample_hdrs, row_count)

        # ── Notes ────────────────────────────────────────────────────
        notes: list[str] = []
        if len(header_rows) > 1:
            notes.append(f"Multi-row header ({len(header_rows)} rows)")
        if orientation == "col_per_record":
            notes.append("Transposed layout (each row is a field)")
        merged_in_hdr = [m for m in merged if m.min_row in header_rows]
        if merged_in_hdr:
            notes.append(f"{len(merged_in_hdr)} merged cell group(s) in header")

        candidates.append(TableCandidate(
            sheet_name=sheet_name,
            bounds=table_bounds,
            header_rows=header_rows,
            first_data_row=first_data_row,
            last_data_row=b_end,
            orientation=orientation,
            score=score,
            sample_headers=sample_hdrs[:15],
            row_count=row_count,
            col_count=col_count,
            density=density,
            has_date_col=has_date,
            has_numeric_col=has_num,
            likely_deal_table=likely_deal,
            notes=notes,
        ))

    return candidates


# ── Header detection ───────────────────────────────────────────────────────────

def _detect_header_rows(cells: dict, table_bounds: CellRange) -> list[int]:
    """
    Identify header rows at the top of a table block.

    Strategy:
      1. Score the first few rows: a "header row" has high string fraction
         and low numeric fraction.
      2. Collect consecutive header rows (for multi-row headers).
      3. If the first row is numeric-heavy, treat it as data and use the
         last all-text row before it as the sole header.

    Always returns at least one row.
    """
    r_min  = table_bounds.min_row
    r_max  = table_bounds.max_row
    cols   = range(table_bounds.min_col, table_bounds.max_col + 1)
    n_cols = max(table_bounds.col_count, 1)

    def string_frac(r: int) -> float:
        vals = [cells[(r, c)] for c in cols if (r, c) in cells]
        return sum(1 for v in vals if isinstance(v, str)) / max(len(vals), 1)

    def numeric_frac(r: int) -> float:
        vals = [cells[(r, c)] for c in cols if (r, c) in cells]
        return sum(1 for v in vals
                   if isinstance(v, (int, float)) and not isinstance(v, bool)
                   ) / max(len(vals), 1)

    def filled_frac(r: int) -> float:
        return sum(1 for c in cols if (r, c) in cells) / n_cols

    header_rows: list[int] = []
    for r in range(r_min, min(r_min + _MAX_HDR_ROWS, r_max)):
        sf = string_frac(r)
        nf = numeric_frac(r)
        ff = filled_frac(r)
        # A header row is dense, predominantly text, and not predominantly numbers.
        if ff >= 0.20 and sf >= 0.45 and nf < 0.40:
            header_rows.append(r)
        else:
            # Stop at the first row that looks like data — unless it's very sparse
            if ff >= 0.20:
                break

    # Fallback: if nothing detected, use first row
    if not header_rows:
        return [r_min]

    return header_rows


# ── Orientation detection ──────────────────────────────────────────────────────

def _detect_orientation(
    cells: dict, bounds: CellRange, header_rows: list[int]
) -> str:
    """
    Distinguish between:
      "row_per_record"  — normal tabular layout (each row is a deal/company)
      "col_per_record"  — transposed (each column is a deal/company, each row a field)

    Heuristic: if the first data column is overwhelmingly strings (labels)
    and other columns are overwhelmingly numeric, the table is transposed.
    """
    first_data_row = max(header_rows) + 1
    if first_data_row > bounds.max_row:
        return "unknown"

    col_min = bounds.min_col

    # Inspect the first data column
    first_col_vals = [cells[(r, col_min)]
                      for r in range(first_data_row, bounds.max_row + 1)
                      if (r, col_min) in cells]
    if not first_col_vals:
        return "unknown"

    first_col_str_frac = (
        sum(1 for v in first_col_vals if isinstance(v, str)) / len(first_col_vals)
    )

    # Inspect the first data row (values across columns)
    first_data_vals = [cells[(first_data_row, c)]
                       for c in range(col_min + 1, bounds.max_col + 1)
                       if (first_data_row, c) in cells]
    if not first_data_vals:
        return "row_per_record"

    first_row_num_frac = (
        sum(1 for v in first_data_vals
            if isinstance(v, (int, float)) and not isinstance(v, bool))
        / len(first_data_vals)
    )

    # If first column is mostly strings AND first data row is mostly numbers
    # → transposed layout
    if first_col_str_frac > 0.75 and first_row_num_frac > 0.5:
        return "col_per_record"

    return "row_per_record"


# ── Header value extraction ────────────────────────────────────────────────────

def _extract_headers(
    cells: dict,
    header_rows: list[int],
    col_min: int,
    col_max: int,
    merged: list[MergedRegion],
) -> list[str]:
    """
    Build a list of header labels, one per column.

    For multi-row headers, concatenate values from all header rows
    in a column (e.g. "Region | Country").
    For merged cells spanning multiple columns, assign the merged label
    to each spanned column.
    """
    # Build a merged-cell lookup: (row, col) → effective value
    merged_lookup: dict[tuple[int, int], str] = {}
    for mr in merged:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if mr.value:
                    merged_lookup[(r, c)] = mr.value

    headers: list[str] = []
    for c in range(col_min, col_max + 1):
        parts: list[str] = []
        for r in header_rows:
            v = merged_lookup.get((r, c)) or cells.get((r, c))
            if v is not None:
                sv = str(v).strip()
                if sv and sv not in ("-",):
                    parts.append(sv)
        if parts:
            # Deduplicate consecutive identical parts (same merged label repeated)
            deduped = [parts[0]]
            for p in parts[1:]:
                if p != deduped[-1]:
                    deduped.append(p)
            headers.append(" | ".join(deduped))
        else:
            headers.append("")

    # Strip trailing empty headers
    while headers and not headers[-1]:
        headers.pop()

    return [h for h in headers if h]


# ── Column type heuristics ─────────────────────────────────────────────────────

def _has_date_column(cells: dict) -> bool:
    for v in cells.values():
        if isinstance(v, (date, datetime)):
            return True
        if isinstance(v, str) and re.search(r'\b(19|20)\d{2}\b', v):
            return True
    return False


def _has_numeric_column(cells: dict) -> bool:
    return any(
        isinstance(v, (int, float)) and not isinstance(v, bool)
        for v in cells.values()
    )


# ── Scoring ────────────────────────────────────────────────────────────────────

def _score_table(
    row_count: int,
    col_count: int,
    density: float,
    sample_headers: list[str],
    has_date: bool,
    has_num: bool,
    orientation: str,
) -> float:
    """
    Composite 0–100 score; higher = more likely to be the main deal table.

    Component weights:
      Row count       0–25  (log scale; 10 rows = ~19, 100 rows = ~25)
      Column count    0–20  (wider tables preferred, cap at 10 cols → 20 pts)
      Density         0–15  (denser = more reliable data)
      Date column     0–10  (deal tables always have dates)
      Numeric column  0–10  (deal tables always have financials)
      Keyword hits    0–20  (PE terminology in headers)
      Orientation      −10  (penalty for transposed — usually not deal-level)
    """
    score = 0.0

    # Row count (log scale)
    score += min(25.0, 8 * math.log1p(row_count))

    # Column count
    score += min(20.0, 2.0 * col_count)

    # Density
    score += 15.0 * min(density, 1.0)

    # Date / numeric presence
    if has_date:
        score += 10.0
    if has_num:
        score += 10.0

    # PE keyword match
    header_text = " ".join(sample_headers).lower()
    kw_hits = sum(1 for kw in _DEAL_KEYWORDS if kw in header_text)
    score += min(20.0, 2.0 * kw_hits)

    # Orientation penalty
    if orientation == "col_per_record":
        score -= 10.0

    return max(0.0, min(100.0, score))


def _is_likely_deal_table(headers: list[str], row_count: int) -> bool:
    """True when headers strongly suggest a deal-level (one row per company) table."""
    if row_count < _MIN_TABLE_ROWS:
        return False
    text  = " ".join(headers).lower()
    hits  = sum(1 for kw in _DEAL_KEYWORDS if kw in text)
    return hits >= 3


# ═══════════════════════════════════════════════════════════════════════════════
# Metadata Extraction
# ═══════════════════════════════════════════════════════════════════════════════

_FN_DESCRIPTORS = re.compile(
    r"\b(consolidated|master|data\s*sheet|data|template|portfolio|workbook|report|"
    r"spreadsheet|deal\s*level|gross|track\s*record|value\s*creation|analysis|"
    r"equity|compan(y|ies)|all\s*funds|segmented|q[1-4]\b|fy\s*20\d\d|20\d\d)\b",
    re.IGNORECASE,
)


def _gp_from_filename(filename: str) -> str:
    """
    Derive the GP name from the upload filename, which conventionally leads with
    the GP name: '<GP Name> <descriptor>…' or '<GP Name> - <descriptor>…'.
    Cuts at the first ' - ' separator or descriptor/date keyword.
    """
    name = str(filename or "").replace("\\", "/").split("/")[-1]   # basename only
    name = re.sub(r"\.(xlsx|xlsm|xls)$", "", name, flags=re.I)
    name = name.strip().lstrip("[]").strip()
    if " - " in name:
        name = name.split(" - ")[0].strip()
    m = _FN_DESCRIPTORS.search(name)
    if m:
        name = name[:m.start()].strip()
    name = name.strip(" -_")
    return name if 1 <= len(name) <= 60 else ""


_ASOF_LABEL_RE = re.compile(
    r"\b(as\s+of|as\s+at|data\s+as\s+of|report(ing)?\s+date|track\s+record\s+date|"
    r"valuation\s+date|data\s+date)\b", re.IGNORECASE)


def _date_from_filename(filename: str) -> Optional[date]:
    """Report date from filename conventions: 12.31.2025 / 2025.09.30 /
    2025-09-30 / Q3 2025 / 2025 Q3."""
    name = str(filename or "").replace("\\", "/").split("/")[-1]
    for pat, order in [
        (r"(\d{4})[._-](\d{1,2})[._-](\d{1,2})", "ymd"),
        (r"(\d{1,2})[._-](\d{1,2})[._-](\d{4})", "mdy"),
    ]:
        for m in re.finditer(pat, name):
            a, b, c = (int(x) for x in m.groups())
            y, mo, d = (a, b, c) if order == "ymd" else (c, a, b)
            if mo > 12 and d <= 12:
                mo, d = d, mo
            try:
                if 1990 <= y <= 2100:
                    return date(y, mo, d)
            except ValueError:
                continue
    m = re.search(r"Q([1-4])\s*[' ]?(\d{4})|(\d{4})\s*[' ]?Q([1-4])", name, re.IGNORECASE)
    if m:
        q = int(m.group(1) or m.group(4)); y = int(m.group(2) or m.group(3))
        from calendar import monthrange
        mo = q * 3
        return date(y, mo, monthrange(y, mo)[1])
    return None


def detect_track_record_date(file_bytes: bytes, filename: str,
                             fallback: Optional[date] = None) -> Optional[date]:
    """
    The as-of date the GP reports for the track record. Priority:
      1. a date next to an "as of"/"report date"-style label in the top rows
         of any visible sheet;  2. a date embedded in the filename;
      3. the caller's fallback (e.g. profiler first-seen date).
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        for ws in wb.worksheets:
            if getattr(ws, "sheet_state", "visible") != "visible":
                continue
            rows = []
            for row in ws.iter_rows(min_row=1, max_row=12, max_col=25):
                rows.append(list(row))
                if len(rows) >= 12:
                    break
            for ri, row in enumerate(rows):
                for ci, cell in enumerate(row):
                    v = cell.value
                    if not isinstance(v, str) or not _ASOF_LABEL_RE.search(v):
                        continue
                    inline = _parse_date_string(
                        _ASOF_LABEL_RE.split(v)[-1].strip(" :,-"))
                    if inline:
                        wb.close(); return inline
                    neighbours = []
                    if ci + 1 < len(row): neighbours.append(row[ci + 1])
                    if ci + 2 < len(row): neighbours.append(row[ci + 2])
                    if ri + 1 < len(rows) and ci < len(rows[ri + 1]):
                        neighbours.append(rows[ri + 1][ci])
                    for nb in neighbours:
                        d = _coerce_date(nb.value)
                        if d is None and isinstance(nb.value, str):
                            d = _parse_date_string(nb.value)
                        if d and d.year >= _DATE_MIN_YEAR:
                            wb.close(); return d
        wb.close()
    except Exception:
        pass
    return _date_from_filename(filename) or fallback


# File-level monetary-unit declarations in sheet banner rows, e.g.
# "($ in thousands)", "USD '000", "in millions", "$mm".
_UNIT_BANNER_K_RE = re.compile(
    r"\bin\s+thousands\b"
    r"|\(\s*(?:[$€£]|USD|EUR|GBP)?\s*(?:in\s+)?['’`]?000s?\s*\)"
    r"|(?:[$€£]|\b(?:USD|EUR|GBP))\s*(?:in\s+)?['’`]000s?\b"
    r"|\bin\s+['’`]?000s\b",
    re.IGNORECASE)
_UNIT_BANNER_M_RE = re.compile(
    r"\bin\s+millions\b"
    r"|\(\s*(?:[$€£]|USD|EUR|GBP)?\s*(?:in\s+)?millions?\s*\)"
    r"|(?:[$€£]|\b(?:USD|EUR|GBP))\s*mm\b",
    re.IGNORECASE)


def detect_unit_banner(file_bytes: bytes) -> Optional[str]:
    """
    File-level monetary-unit declaration from the banner rows (top 12) of
    visible sheets: "k" (thousands), "m" (millions), or None. Conflicting
    declarations across sheets → None (fall back to value magnitudes).
    """
    found: set[str] = set()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        for ws in wb.worksheets:
            if getattr(ws, "sheet_state", "visible") != "visible":
                continue
            for row in ws.iter_rows(max_row=12, values_only=True):
                for v in row:
                    if not isinstance(v, str):
                        continue
                    if _UNIT_BANNER_K_RE.search(v):
                        found.add("k")
                    if _UNIT_BANNER_M_RE.search(v):
                        found.add("m")
        wb.close()
    except Exception:
        return None
    return found.pop() if len(found) == 1 else None


def _extract_meta(ws) -> tuple[list[str], Optional[date], Optional[str]]:
    """
    Scan the first 10 rows × 20 columns for GP name candidates, report date,
    and currency.  Returns (gp_names, report_date, currency).
    """
    gp_names: list[str] = []
    report_date: Optional[date] = None
    currency: Optional[str] = None

    for row in ws.iter_rows(max_row=10, max_col=20, values_only=True):
        for v in row:
            if v is None:
                continue
            sv = str(v).strip()
            if not sv:
                continue

            # Currency — first hit wins
            if currency is None:
                m = _CURRENCY_RE.search(sv)
                if m:
                    currency = m.group(1).upper()

            # Date — first plausible date wins
            if report_date is None:
                d = _coerce_date(v)
                if d and d.year >= _DATE_MIN_YEAR:
                    report_date = d
                elif isinstance(v, str):
                    d = _parse_date_string(sv)
                    if d:
                        report_date = d

            # GP name candidate — non-numeric strings of reasonable length
            if (isinstance(v, str) and 4 <= len(sv) <= 80
                    and not isinstance(v, (int, float))
                    and not _CURRENCY_RE.search(sv)
                    and _coerce_date(v) is None):
                # Strip common trailing words
                name = re.split(
                    r'\s+(equity|portfolio|workbook|track|record|data|report|'
                    r'management|capital|partners|fund|group|llp|llc|lp|gp|'
                    r'holdings|investments|advisors|associates|asset|private)\b',
                    sv, maxsplit=1, flags=re.IGNORECASE,
                )[0].strip()
                if 3 < len(name) <= 60:
                    gp_names.append(name)

    return gp_names, report_date, currency


def _coerce_date(v: Any) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _parse_date_string(s: str) -> Optional[date]:
    """Try common date string formats."""
    s = s.strip()[:20]
    for fmt in (
        "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y",
        "%d %b %Y", "%d-%b-%Y", "%b %d, %Y", "%B %d, %Y",
        "%d %B %Y", "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    # "Q2 2025" → use last day of the quarter as a proxy
    m = re.match(r'Q(\d)\s+(\d{4})', s, re.IGNORECASE)
    if m:
        q, yr = int(m.group(1)), int(m.group(2))
        month = q * 3
        try:
            from calendar import monthrange
            last_day = monthrange(yr, month)[1]
            return date(yr, month, last_day)
        except ValueError:
            pass
    return None
