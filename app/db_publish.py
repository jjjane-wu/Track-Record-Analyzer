"""
db_publish.py — Publish a *verified* Deal Level Input workbook to the database.

The database is deliberately simple: a folder of CSV snapshots, one file per
GP per as-of date ("GP_2 - 2025-09-30.csv"). Point the folder at a
OneDrive-synced SharePoint library and every published file uploads itself;
Power BI reads the whole folder with its SharePoint/Folder connector and
combines the files into one long table.

Publishing is a separate, human-triggered step — never automatic. The app's
parsed output may contain mapping or data errors, so the analyst first
downloads the Deal Level Input workbook, corrects and verifies it in Excel,
and only then publishes that file here. The workbook is the single source of
truth: the same verified file feeds both TR-Analyzer.xlsm and the database.

Re-publishing the same GP + as-of date overwrites its snapshot (idempotent —
a correction replaces the old rows, nothing duplicates). A new as-of date
creates a new snapshot, preserving the history of the track record over time.

CLI (for scripted use):
    python app/db_publish.py "path/to/[12-Aug-26 - GP_2] - Gross Deal Level Input.xlsx"
    python app/db_publish.py input.xlsx --dir "~/OneDrive/TR Database" --by "Jane"
"""

from __future__ import annotations

import io
import json
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_output import INPUT_COLS   # single source of truth for the schema

ROOT_DIR    = Path(__file__).resolve().parent.parent
DB_DIR      = ROOT_DIR / "database"
CONFIG_PATH = DB_DIR / "db_config.json"

# Record keys (transformer numbering, via INPUT_COLS) by value type.
_NUMERIC_KEYS = {16, 17, 18, 20, 35, 36, 37, 39, 42, 46, 47, 49, 52}
_DATE_KEYS    = {6, 7}

_ALLOWED_STATUS = {"realized", "unrealized"}


# Old column names that later releases renamed — normalised on read so
# previously downloaded input files keep working everywhere.
_LEGACY_HEADERS = {"fund currency": "Deal Currency"}


def _clean_header(h: Any) -> str:
    """Collapse newlines/extra spaces: 'Realized\\nValue' -> 'Realized Value'."""
    out = " ".join(str(h).split()) if h is not None else ""
    return _LEGACY_HEADERS.get(out.lower(), out)


# Cleaned expected headers, in schema order, with their record keys.
EXPECTED_HEADERS: list[tuple[str, int]] = [
    (_clean_header(h), k) for h, k in INPUT_COLS
]
_HEADER_TO_KEY = {h.lower(): k for h, k in EXPECTED_HEADERS}


# ═══════════════════════════════════════════════════════════════════════════
# Reading the Deal Level Input workbook (our own fixed format)
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class ParsedInput:
    gp: str = ""
    as_of: date | None = None
    currency: str = ""
    headers: list[str] = field(default_factory=list)   # cleaned, as found
    rows: list[dict[str, Any]] = field(default_factory=list)
    issues: list[str] = field(default_factory=list)    # structural warnings
    sheet_name: str = ""
    source_name: str = ""
    header_row: int = 0                                # 1-based Excel row


def _as_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str) and v.strip():
        try:
            return pd.to_datetime(v.strip(), dayfirst=False).date()
        except Exception:
            return None
    return None


def read_input_workbook(source: bytes | str | Path,
                        source_name: str = "") -> ParsedInput:
    """Parse a Deal Level Input workbook (path or raw bytes) into ParsedInput.

    The layout is our own generated format (meta labels in column B, header
    row starting at 'Company' in column B, data rows below) — but the file may
    have been hand-corrected in Excel, so positions are located by label
    rather than assumed.
    """
    if isinstance(source, (str, Path)):
        p = Path(source)
        data = p.read_bytes()
        source_name = source_name or p.name
    else:
        data = source

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    try:
        ws = wb["Deal Level Inputs"] if "Deal Level Inputs" in wb.sheetnames \
             else wb[wb.sheetnames[0]]
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        parsed = ParsedInput(sheet_name=ws.title, source_name=source_name)
    finally:
        wb.close()

    def cell(r: int, c: int) -> Any:          # 0-based, blank-safe
        if 0 <= r < len(grid) and 0 <= c < len(grid[r]):
            return grid[r][c]
        return None

    # ── Meta block: labels in column B, values in column C ──────────────
    for r in range(min(15, len(grid))):
        label = _clean_header(cell(r, 1)).lower()
        v = cell(r, 2)
        if label == "gp name" and v is not None:
            parsed.gp = str(v).strip()
        elif label == "track record date":
            parsed.as_of = _as_date(v)
        elif label == "currency" and v is not None:
            parsed.currency = str(v).strip()

    # ── Header row: 'Company' in column B ───────────────────────────────
    hdr_r = next((r for r in range(min(25, len(grid)))
                  if _clean_header(cell(r, 1)).lower() == "company"), None)
    if hdr_r is None:
        parsed.issues.append(
            "Could not find the header row (no 'Company' cell in column B) — "
            "is this really a Deal Level Input file?")
        return parsed
    parsed.header_row = hdr_r + 1

    headers: list[str] = []
    c = 1
    while True:
        h = _clean_header(cell(hdr_r, c))
        if not h:
            break
        headers.append(h)
        c += 1
    parsed.headers = headers

    expected = [h for h, _ in EXPECTED_HEADERS]
    missing = [h for h in expected if h.lower() not in {x.lower() for x in headers}]
    unknown = [h for h in headers if h.lower() not in _HEADER_TO_KEY]
    if missing:
        parsed.issues.append(
            f"{len(missing)} expected column(s) not found (left blank in the "
            f"database): {', '.join(missing)}")
    if unknown:
        parsed.issues.append(
            f"{len(unknown)} unrecognised column(s) ignored: {', '.join(unknown)}")

    # ── Data rows: below the header until the first fully blank row ─────
    span = len(headers)
    for r in range(hdr_r + 1, len(grid)):
        vals = [cell(r, 1 + j) for j in range(span)]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in vals):
            break
        row: dict[str, Any] = {}
        for h, v in zip(headers, vals):
            if isinstance(v, str):
                v = v.strip()
            row[h] = v
        parsed.rows.append(row)

    return parsed


# ═══════════════════════════════════════════════════════════════════════════
# Validation — the publish gate
# ═══════════════════════════════════════════════════════════════════════════

def validate(parsed: ParsedInput) -> tuple[list[str], list[str]]:
    """Return (errors, warnings). Errors block publishing; warnings don't."""
    errors: list[str] = []
    warnings: list[str] = list(parsed.issues)

    if not parsed.gp:
        errors.append("GP Name is missing (cell C3 of the input file).")
    if parsed.as_of is None:
        errors.append("Track Record Date is missing or not a date (cell C4).")
    if not parsed.rows:
        errors.append("No deal rows found below the header row.")
    if not parsed.currency:
        warnings.append("Currency is missing (cell C5) — Deal Currency blanks "
                        "cannot be back-filled.")

    def col_of(key: int) -> str | None:
        for h, k in EXPECTED_HEADERS:
            if k == key:
                return next((x for x in parsed.headers
                             if x.lower() == h.lower()), None)
        return None

    numeric_cols = [c for k in _NUMERIC_KEYS if (c := col_of(k))]
    date_cols    = [c for k in _DATE_KEYS if (c := col_of(k))]
    company_col, fund_col = col_of(1), col_of(2)
    status_col, tvpi_col, irr_col = col_of(5), col_of(20), col_of(35)
    ic_col = col_of(16)

    bad_numeric: list[str] = []
    bad_dates: list[str] = []
    seen: dict[tuple[str, str], int] = {}
    n_blank_company = n_blank_fund = n_bad_status = 0
    n_bad_tvpi = n_big_irr = n_no_ic = 0

    for i, row in enumerate(parsed.rows):
        xl_row = parsed.header_row + 1 + i

        company = str(row.get(company_col) or "").strip() if company_col else ""
        if not company:
            n_blank_company += 1
        fund = str(row.get(fund_col) or "").strip() if fund_col else ""
        if not fund:
            n_blank_fund += 1
        if company and fund:
            key = (company.lower(), fund.lower())
            if key in seen:
                warnings.append(
                    f"Row {xl_row}: duplicate deal '{company}' in '{fund}' "
                    f"(first seen row {seen[key]}).")
            else:
                seen[key] = xl_row

        for c in numeric_cols:
            v = row.get(c)
            if v is None or v == "" or isinstance(v, (int, float)):
                continue
            bad_numeric.append(f"{c} row {xl_row}: '{v}'")
        for c in date_cols:
            v = row.get(c)
            if v in (None, "") or _as_date(v) is not None:
                continue
            bad_dates.append(f"{c} row {xl_row}: '{v}'")

        if status_col:
            s = str(row.get(status_col) or "").strip().lower()
            if s and s not in _ALLOWED_STATUS:
                n_bad_status += 1
        if tvpi_col and isinstance(row.get(tvpi_col), (int, float)) \
                and row[tvpi_col] < 0:
            n_bad_tvpi += 1
        if irr_col and isinstance(row.get(irr_col), (int, float)) \
                and abs(row[irr_col]) > 3:
            n_big_irr += 1
        if ic_col and row.get(ic_col) in (None, "", 0):
            n_no_ic += 1

    if n_blank_company:
        errors.append(f"{n_blank_company} row(s) have no Company name.")
    if bad_numeric:
        shown = "; ".join(bad_numeric[:8])
        more = f" … and {len(bad_numeric) - 8} more" if len(bad_numeric) > 8 else ""
        errors.append(f"Non-numeric text in numeric column(s) — fix in Excel "
                      f"before publishing: {shown}{more}")
    if bad_dates:
        shown = "; ".join(bad_dates[:8])
        more = f" … and {len(bad_dates) - 8} more" if len(bad_dates) > 8 else ""
        errors.append(f"Unreadable date(s): {shown}{more}")

    if n_blank_fund:
        warnings.append(f"{n_blank_fund} row(s) have no Fund.")
    if n_bad_status:
        warnings.append(f"{n_bad_status} row(s) have a Status other than "
                        "Realized/Unrealized.")
    if n_bad_tvpi:
        warnings.append(f"{n_bad_tvpi} row(s) have a negative Gross TVPI.")
    if n_big_irr:
        warnings.append(f"{n_big_irr} row(s) have |Gross IRR| > 300% — check "
                        "whether IRR was entered in percentage points "
                        "(25 instead of 0.25).")
    if n_no_ic:
        warnings.append(f"{n_no_ic} row(s) have no Total Invested Capital.")

    return errors, warnings


# ═══════════════════════════════════════════════════════════════════════════
# Long table + publishing
# ═══════════════════════════════════════════════════════════════════════════

def to_long_table(parsed: ParsedInput,
                  published_by: str = "",
                  published_at: datetime | None = None) -> pd.DataFrame:
    """One row per deal; GP/as-of stamped on every row; tidy typed columns."""
    published_at = published_at or datetime.now()
    fc_col = next((h for h, k in EXPECTED_HEADERS if k == 90), "Deal Currency")

    out_rows: list[dict[str, Any]] = []
    for row in parsed.rows:
        rec: dict[str, Any] = {
            "GP Name": parsed.gp,
            "Track Record Date": parsed.as_of.isoformat() if parsed.as_of else "",
        }
        for h, key in EXPECTED_HEADERS:
            src = next((x for x in parsed.headers if x.lower() == h.lower()), None)
            v = row.get(src) if src else None
            if key in _DATE_KEYS:
                d = _as_date(v)
                v = d.isoformat() if d else ""
            elif v is None:
                v = ""
            if h == fc_col and v == "" and parsed.currency:
                v = parsed.currency          # back-fill from the meta block
            rec[h] = v
        rec["Source File"]  = parsed.source_name
        rec["Published By"] = published_by
        rec["Published At"] = published_at.strftime("%Y-%m-%d %H:%M:%S")
        out_rows.append(rec)

    return pd.DataFrame(out_rows)


def snapshot_filename(parsed: ParsedInput) -> str:
    gp = re.sub(r"[\\/:*?\"<>|]", "-", parsed.gp).strip() or "GP"
    as_of = parsed.as_of.isoformat() if parsed.as_of else "undated"
    return f"{gp} - {as_of}.csv"


def publish(parsed: ParsedInput, db_dir: str | Path,
            published_by: str = "") -> tuple[Path, bool]:
    """Write the snapshot CSV into db_dir. Returns (path, replaced_existing).

    Same GP + as-of overwrites its previous snapshot (a correction replaces,
    never duplicates). The write goes through a temp file + atomic rename so
    a sync client never sees a half-written file.
    """
    db_dir = Path(db_dir).expanduser()
    db_dir.mkdir(parents=True, exist_ok=True)
    target = db_dir / snapshot_filename(parsed)
    replaced = target.exists()

    df = to_long_table(parsed, published_by=published_by)
    tmp = target.with_suffix(".csv.tmp")
    df.to_csv(tmp, index=False, encoding="utf-8-sig")
    os.replace(tmp, target)
    return target, replaced


def list_snapshots(db_dir: str | Path) -> pd.DataFrame:
    """Inventory of the database folder: one row per snapshot CSV."""
    db_dir = Path(db_dir).expanduser()
    rows = []
    if db_dir.is_dir():
        for p in sorted(db_dir.glob("*.csv")):
            try:
                df = pd.read_csv(p, dtype=str, keep_default_na=False)
                first = df.iloc[0] if len(df) else {}
                rows.append({
                    "File": p.name,
                    "GP": first.get("GP Name", ""),
                    "As of": first.get("Track Record Date", ""),
                    "Deals": len(df),
                    "Published": first.get("Published At", ""),
                    "By": first.get("Published By", ""),
                })
            except Exception as e:
                rows.append({"File": p.name, "GP": f"(unreadable: {e})",
                             "As of": "", "Deals": 0, "Published": "", "By": ""})
    return pd.DataFrame(rows,
                        columns=["File", "GP", "As of", "Deals", "Published", "By"])


# ═══════════════════════════════════════════════════════════════════════════
# Config — where the database folder lives
# ═══════════════════════════════════════════════════════════════════════════

def default_deals_dir() -> Path:
    return DB_DIR / "deals"


def load_db_config() -> dict:
    try:
        cfg = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        cfg = {}
    cfg.setdefault("deals_dir", str(default_deals_dir()))
    return cfg


def save_db_config(deals_dir: str | Path) -> None:
    DB_DIR.mkdir(parents=True, exist_ok=True)
    cfg = load_db_config()
    cfg["deals_dir"] = str(Path(deals_dir).expanduser())
    CONFIG_PATH.write_text(json.dumps(cfg, indent=2), encoding="utf-8")


# ═══════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════

def _main(argv: list[str]) -> int:
    import argparse
    ap = argparse.ArgumentParser(
        description="Publish a verified Deal Level Input workbook to the "
                    "database folder (one CSV snapshot per GP per as-of date).")
    ap.add_argument("input", help="Path to the verified Deal Level Input .xlsx")
    ap.add_argument("--dir", default=None,
                    help="Database folder (default: the configured folder)")
    ap.add_argument("--by", default="", help="Analyst name for provenance")
    ap.add_argument("--allow-warnings", action="store_true",
                    help="Publish even when validation warnings exist "
                         "(errors always block)")
    args = ap.parse_args(argv)

    parsed = read_input_workbook(args.input)
    errors, warns = validate(parsed)
    for w in warns:
        print(f"WARNING: {w}")
    for e in errors:
        print(f"ERROR:   {e}")
    if errors:
        print("Not published — fix the errors in Excel and re-run.")
        return 1
    if warns and not args.allow_warnings:
        print(f"Not published — {len(warns)} warning(s) above. Re-run with "
              "--allow-warnings after confirming they are fine.")
        return 2

    db_dir = args.dir or load_db_config()["deals_dir"]
    path, replaced = publish(parsed, db_dir, published_by=args.by)
    verb = "Replaced snapshot" if replaced else "Published"
    print(f"{verb}: {path}  ({parsed.gp}, as of {parsed.as_of}, "
          f"{len(parsed.rows)} deals)")
    return 0


if __name__ == "__main__":
    raise SystemExit(_main(sys.argv[1:]))
