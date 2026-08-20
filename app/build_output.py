"""
build_output.py — Build the Segmented Track Record output from a BLANK workbook.

Produces a clean 3-tab workbook (no heavy template, no external links):
  1. Deal Level Inputs — the clean input data (values).
  2. Deal List        — analysis-ready deal table: input columns are formulas
                        linking to Deal Level Inputs; computed columns (Vintage,
                        Total Value, MOIC, buckets, …) are formulas. An Excel
                        Table named "DealList" — the source of every pivot.
  3. Return & Loss Ratios — REAL pivot tables (Count / MOIC / Loss Ratio by
                        Sector, Geography, Process Type, GP Role, Exit Type,
                        buckets, Vintage, COI, Fund), refresh-on-open, mirroring
                        the Segmented Track Record Analysis Output template.

openpyxl builds the two data sheets; the pivots are injected as raw OOXML
(openpyxl cannot create pivots). The pivot XML structure is Excel-verified:
  - workbook <pivotCaches> must sit at the END of workbook.xml (child order!)
  - the cache ships POPULATED pivotCacheRecords (one <r> per deal)
  - with 2+ data fields, the σ-Values field (x=-2) must be on the column axis
"""

from __future__ import annotations

import io
import re
import zipfile
from datetime import date, datetime
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Color, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import StrRef

from transformer import (
    REV_LABELS, EBITDA_LABELS, ENTRY_MULT_LABELS,
    EBITDA_MARGIN_LABELS, ENTRY_EV_LABELS, IC_LABELS, HP_LABELS,
)
from deal_list_spec import DL_COLS, DL_FORMATS, HEADER_BLOCK, TAG_ROW


# ═══════════════════════════════════════════════════════════════════════════════
# Column layouts
# ═══════════════════════════════════════════════════════════════════════════════

# Deal Level Inputs: (header, transformer record key).
INPUT_COLS: list[tuple[str, Any]] = [
    ("Company", 1), ("Fund", 2), ("Status", 5), ("Fund Currency", 90),
    ("Inv. Date", 6), ("Exit Date", 7),
    ("Sector", 11), ("Geography", 12), ("Total Invested Capital (mlns)", 16),
    ("Realized\nValue", 17), ("Current\nValue", 18), ("Transaction Type", 29),
    ("GP Role", 30), ("Process Type", 31), ("Sourcing Partner", 32),
    ("Exit Type", 33), ("COI Deal (Yes/No)", 34), ("Gross TVPI", 20), ("Gross\nIRR", 35),
    ("Valuation Method", 55), ("Entry LTM\nRevenue", 36), ("Entry LTM\nEBITDA", 37),
    ("Entry\nNet Debt", 39), ("Entry Enterprise\nValue", 42), ("Exit LTM\nRevenue", 46),
    ("Exit LTM\nEBITDA", 47), ("Exit\nNet Debt", 49), ("Exit Enterprise Value", 52),
]

# Template column widths for the Deal Level Inputs tab (letter → width),
# re-lettered after the 'Data as of' column was removed (user request).
_INPUT_WIDTHS = {
    "A": 8.8, "B": 30.5, "C": 15.5, "E": 12.5, "F": 15.8, "H": 28.5,
    "I": 20.0, "J": 18.2, "L": 13.8, "M": 18.5, "Q": 26.2, "R": 16.0,
    "U": 16.2, "V": 17.5, "W": 17.3, "X": 13.5, "Y": 17.5, "Z": 13.5,
    "AA": 20.5, "AB": 18.2, "AC": 17.5,
}
_INPUT_BLUE = Font(color="0000FF", size=10)          # classic "input cell" blue
_HDR_VALUE_FONT = Font(bold=True, size=10, color="0000FF")

# Template cell styling (IO templates): INPUT cells are light blue with blue
# text ("Accent1, Lighter 80%" = theme 4 tint 0.8); formula cells are plain
# white with black 10pt; five key computed Deal List columns carry a light
# grey fill ("Background 1, Darker 5%" = theme 0 tint -0.05).
_INPUT_FILL = PatternFill("solid", fgColor=Color(theme=4, tint=0.8))
_GRAY_FILL = PatternFill("solid", fgColor=Color(theme=0, tint=-0.05))
_CALC_FONT = Font(size=10)
_GRAY_COLS = {"Total\nValue", "Gross\nMOIC", "Performing\n(1=Underperform)",
              "InvCapital in Loss Position", "Impaired\nValue"}


# The Deal List schema is the template's own 80-column layout (see
# deal_list_spec.py, extracted verbatim from the Segmented Analysis template).
DEAL_LIST_COLS = DL_COLS

# Dispersion bucket labels (match the Deal List helper-table CONCATENATEs)
_MOIC_BUCKET_LABELS = ["<=1.0x", "1.0x - 2.0x", "2.0x - 3.0x", ">=3.0x"]
_IRR_BUCKET_LABELS = ["<=0%", "0% - 10%", "10% - 20%", ">=20%"]

# Canonical (ordered) item lists for bucket dimensions so pivot rows appear in
# bucket order rather than first-seen order.
_CANONICAL_ORDER: dict[str, list[str]] = {
    "Revenue Buckets": REV_LABELS,
    "EBITDA Buckets": EBITDA_LABELS,
    "Entry Multiple Bucket": ENTRY_MULT_LABELS,
    "Entry EBITDA Margin Bucket": EBITDA_MARGIN_LABELS,
    "Entry Enterprise\nValueBuckets": ENTRY_EV_LABELS,
    "Initial InvCap Buckets": IC_LABELS,
    "Hold Period Buckets": HP_LABELS,
    "Status": ["Realized", "Unrealized"],
    "MOIC Buckets": _MOIC_BUCKET_LABELS,
    "IRR Buckets": _IRR_BUCKET_LABELS,
}

# Report filters shipped on every pivot (template: Fund / Status / Hold
# Period Buckets page fields). A pivot whose row axis IS one of these drops
# that filter (a field cannot sit on two axes).
_PAGE_FIELD_HEADERS = ("Fund", "Status", "Hold Period Buckets")

# The Return & Loss Ratios breakdowns, in the template's order.
# (title, axis field header, extra_datafields)
#   extra "impaired" → Count/MOIC/Impaired Loss Ratio instead of Loss Ratio
#   extra "with_ic"  → adds Sum of Initial Invested Capital as a 4th data field
PIVOT_SPECS: list[tuple[str, str, str]] = [
    ("Gross Returns & Loss Ratios by Sector",                    "Sector",                ""),
    ("Gross Returns & Loss Ratios by Geography",                 "Geography",             ""),
    ("Gross Returns & Loss Ratios by Process Type",              "Process Type",          ""),
    ("Gross Returns & Loss Ratios by GP Role",                   "GP Role",               ""),
    ("Gross Returns & Loss Ratios by Exit Type",                 "Exit Type",             ""),
    ("Gross Returns & Loss Ratios by Entry Revenue",             "Revenue Buckets",       ""),
    ("Gross Returns & Loss Ratios by Entry EBITDA",              "EBITDA Buckets",        ""),
    ("Gross Returns & Loss Ratios by Entry EBITDA Multiple",     "Entry Multiple Bucket", ""),
    ("Impaired Loss Ratios by Entry Enterprise Value",           "Entry Enterprise\nValueBuckets", "impaired"),
    ("Gross Returns & Loss Ratios by Vintage",                   "Vintage",               ""),
    ("Capital Deployment & Returns by Vintage",                  "Vintage",               "with_ic"),
    ("Gross Returns & Loss Ratios by Entry EBITDA Margin",       "Entry EBITDA Margin Bucket", ""),
    ("Gross Returns & Loss Ratios by Initial Invested Capital",  "Initial InvCap Buckets", ""),
    ("Gross Returns & Loss Ratios by COI Deals",                 "COI Deal (Yes/No)",     ""),
    ("Gross Returns & Loss Ratios by Fund",                      "Fund",                  ""),
]

_HDR_FILL = PatternFill("solid", fgColor="1F4E78")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_SECTION_FONT = Font(bold=True, size=11, color="1F4E78")
_INPUT_START = 7          # Deal Level Inputs data start row (header row 6)
_DL_START = 14            # Deal List data rows (tag row 12, table header row 13)
_PIVOT_GAP = 7            # rows between a section title and the pivot body:
                          # 3 shipped report filters + separator sit at the
                          # bottom; the spare rows above keep extra filters
                          # from overwriting the heading (EWL)

# Debug hook (bisection): headers listed here are degraded to plain string
# cache fields and their pivots skipped.
_DEBUG_GROUP_DISABLE: set = set()
_DEBUG_NO_CHARTS = False
_DEBUG_NO_LABELS = False

_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


# ═══════════════════════════════════════════════════════════════════════════════
# Small helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _val(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, float) and v != v:
        return None
    if isinstance(v, str) and v.strip() in ("", "None", "nan", "NaT"):
        return None
    return v


def _cell_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return None if s.lower() in ("", "none", "nan", "nat", "-") else s


def _cell_num(v: Any):
    if v is None or isinstance(v, bool):
        return None
    try:
        f = float(v)
        return None if (f != f or f in (float("inf"), float("-inf"))) else f
    except (ValueError, TypeError):
        return None


def _numstr(x: float) -> str:
    """Plain-decimal string for XML — never scientific notation."""
    if x == int(x):
        return str(int(x))
    s = repr(x)
    if "e" in s or "E" in s:
        s = f"{x:.10f}".rstrip("0").rstrip(".")
    return s


def _esc(s: str) -> str:
    return (s.replace("&", "&amp;").replace("<", "&lt;")
             .replace(">", "&gt;").replace('"', "&quot;"))


# Charts carry their data as LITERALS (strLit/numLit), not cell references:
# a filtered pivot reshuffles its cells, and range-bound charts would then
# show wrong slices (user rule: filtering a pivot must never change a chart —
# charts always show the full, unfiltered distribution).

def _str_lit(vals: list) -> str:
    pts = "".join(f'<c:pt idx="{i}"><c:v>{_esc(str(v))}</c:v></c:pt>'
                  for i, v in enumerate(vals))
    return f'<c:strLit><c:ptCount val="{len(vals)}"/>{pts}</c:strLit>'


def _num_lit(vals: list, fmt: str = "General") -> str:
    pts = "".join(f'<c:pt idx="{i}"><c:v>{_numstr(float(v))}</c:v></c:pt>'
                  for i, v in enumerate(vals) if v is not None)
    return (f'<c:numLit><c:formatCode>{_esc(fmt)}</c:formatCode>'
            f'<c:ptCount val="{len(vals)}"/>{pts}</c:numLit>')


def _rec_value(rec: dict, header: str, key: int) -> Any:
    """Record value for a Deal List column — must MATCH what the sheet formula
    computes, since Excel re-reads the sheet on refresh (and current Excel
    honours refreshOnLoad, so any mismatch becomes visible on open).

    The aggregate columns are therefore recomputed here with the SHEET's
    semantics (Total Value = Realized + Current, MOIC = TV/IC, …) rather than
    taken from the transformer record, which may carry a GP-reported total."""
    if header == "Vintage":
        d = rec.get(6)                            # deal entry date
        return d.year if isinstance(d, (date, datetime)) else "n/a"
    if header == "Exit Year":
        d = rec.get(7)                            # deal exit date
        return d.year if isinstance(d, (date, datetime)) else "n/a"
    if header in _SHEET_CALC:
        return _SHEET_CALC[header](rec)
    return rec.get(key)


def _sheet_tv(rec: dict):
    """Sheet 'Total\\nValue': IF(AND(Realized="",Current=""),"",N(R)+N(S))."""
    a, b = _cell_num(rec.get(17)), _cell_num(rec.get(18))
    return None if (a is None and b is None) else (a or 0.0) + (b or 0.0)


def _sheet_moic(rec: dict):
    """Sheet 'Gross\\nMOIC': IFERROR(TV/IC, "n/a")."""
    tv, ic = _sheet_tv(rec), _cell_num(rec.get(16))
    return (tv / ic) if (tv is not None and ic) else None


def _moic_below_1(rec: dict) -> bool:
    # sheet comparison MOIC<1: "n/a" (text) compares FALSE in Excel
    m = _sheet_moic(rec)
    return m is not None and m < 1


def _sheet_moic_bucket(rec: dict) -> str:
    m = _sheet_moic(rec)
    if m is None:
        return "n/a"
    return (_MOIC_BUCKET_LABELS[0] if m <= 1 else
            _MOIC_BUCKET_LABELS[1] if m <= 2 else
            _MOIC_BUCKET_LABELS[2] if m <= 3 else _MOIC_BUCKET_LABELS[3])


def _sheet_irr_bucket(rec: dict) -> str:
    v = _cell_num(rec.get(35))                    # Gross IRR (decimal)
    if v is None:
        return "n/a"
    return (_IRR_BUCKET_LABELS[0] if v <= 0 else
            _IRR_BUCKET_LABELS[1] if v <= 0.1 else
            _IRR_BUCKET_LABELS[2] if v <= 0.2 else _IRR_BUCKET_LABELS[3])


_SHEET_CALC = {
    "Total\nValue": _sheet_tv,
    "Gross\nMOIC": _sheet_moic,
    "MOIC Buckets": _sheet_moic_bucket,
    "IRR Buckets": _sheet_irr_bucket,
    "Total IC mlns for Buckets": lambda rec: _cell_num(rec.get(16)) or 0.0,
    "Performing\n(1=Underperform)": lambda rec: 1 if _moic_below_1(rec) else "-",
    "InvCapital in Loss Position":
        lambda rec: (_cell_num(rec.get(16)) or 0.0) if _moic_below_1(rec) else 0.0,
    "Impaired\nValue":
        lambda rec: ((1 - _sheet_moic(rec)) * (_cell_num(rec.get(16)) or 0.0)
                     if _moic_below_1(rec) else 0.0),
}


def _input_col_letter(header: str) -> str:
    for i, (h, _k) in enumerate(INPUT_COLS):
        if h == header:
            return get_column_letter(2 + i)
    raise KeyError(header)


def _dl_field_index(header: str) -> int:
    for i, (h, *_r) in enumerate(DEAL_LIST_COLS):
        if h == header:
            return i
    raise KeyError(header)


# ═══════════════════════════════════════════════════════════════════════════════
# Data sheets (openpyxl)
# ═══════════════════════════════════════════════════════════════════════════════

def _write_mini_toc(ws, entries: list[tuple[str, int]], start_row: int) -> None:
    """Small in-sheet list of contents (template style: numbered light-blue
    cells, banded bordered labels), each label an internal link to its
    section's row on this sheet."""
    thin = Side(style="thin", color="7F7F7F")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    num_fill = PatternFill("solid", fgColor=Color(theme=4, tint=0.6))
    for i, (label, target_row) in enumerate(entries):
        r = start_row + i
        num = ws.cell(row=r, column=2, value=i + 1)
        num.fill = num_fill
        num.font = Font(bold=True, size=10)
        num.alignment = Alignment(horizontal="center")
        num.border = box
        lbl = ws.cell(row=r, column=3, value=label)
        lbl.font = Font(size=10)
        lbl.border = box
        if i % 2 == 1:
            lbl.fill = _GRAY_FILL
        lbl.hyperlink = Hyperlink(ref=lbl.coordinate,
                                  location=f"'{ws.title}'!B{target_row}",
                                  display=label)


# One-line TOC description per known tab (unknown / future tabs stay blank).
# Keep in sync with vba/modToc.bas TabBlurb.
_TOC_BLURBS = {
    "Deal Level Inputs": "The cleaned deal data — the single source of truth every other tab reads",
    "Deal List": "Every deal as plain values plus the full per-deal analytics; the blue threshold tables set the bucket boundaries",
    "Return & Loss Ratios": "Pooled MOIC and loss ratio across 15 cuts - sector, geography, vintage, fund, entry size, exit type - with a chart per cut",
    "Return Dispersion": "MOIC and IRR distributions: deal count, % of invested capital and average return per bucket",
    "Portfolio Construction": "Capital mix by fund x sector / geography, plus deal-count attribute breakdowns",
    "Vintage Perf by Sector": "Invested capital, MOIC and loss ratio by vintage, with vintage x sector count and MOIC matrices",
    "Deployment & Exits": "Capital deployment pacing (vintage x fund) and realization pacing (fund x exit year, realized deals)",
    "Underperforming Assets": "Deals below the performance threshold, with their share of capital and value",
    "Partner Attribution": "Returns and capital by sourcing partner",
    "Op Performance": "Operating metrics (revenue / EBITDA growth, margins) for realized deals",
    "Op Performance - Unrealized": "Operating metrics for the unrealized portfolio",
}


def _write_toc(wb) -> None:
    """Table of Contents: numbered, banded list of internal links to every
    other tab (first sheet of the workbook), each with a one-line blurb."""
    ws = wb["Table of Contents"]
    ws["B2"] = "Table of Contents"
    ws["B2"].font = Font(bold=True, size=16)

    thin = Side(style="thin", color="7F7F7F")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    band = PatternFill("solid", fgColor=Color(theme=0, tint=-0.05))
    link_font = Font(color="0563C1", underline="single", size=10)

    row = 4
    n = 0
    for sheet in wb.worksheets:
        if sheet.title == ws.title:
            continue
        n += 1
        num = ws.cell(row=row, column=2, value=n)
        num.fill = _HDR_FILL
        num.font = Font(bold=True, color="FFFFFF", size=10)
        num.alignment = Alignment(horizontal="center")
        num.border = box
        lbl = ws.cell(row=row, column=3, value=sheet.title)
        lbl.font = link_font
        lbl.border = box
        if n % 2 == 0:
            lbl.fill = band
        lbl.hyperlink = Hyperlink(ref=lbl.coordinate,
                                  location=f"'{sheet.title}'!A1",
                                  display=sheet.title)
        desc = ws.cell(row=row, column=4, value=_TOC_BLURBS.get(sheet.title, ""))
        desc.font = Font(size=10, color="595959")
        if n % 2 == 0:
            desc.fill = band
        row += 1

    ws.column_dimensions["B"].width = 4.5
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 78


def _write_inputs(ws, records: list[dict], gp: str, currency: str,
                  track_record_date: date | None) -> None:
    ws["B2"] = "Deal Level Input"; ws["B2"].font = Font(bold=True, size=16)

    # Compact meta block (EWL layout: no indicator/spacer rows)
    ws["B3"] = "GP Name";            ws["C3"] = gp
    ws["B4"] = "Track Record Date";  ws["C4"] = track_record_date or date.today()
    ws["C4"].number_format = "d-mmm-yy"
    ws["B5"] = "Currency";           ws["C5"] = currency
    for lbl, val in (("B3", "C3"), ("B4", "C4"), ("B5", "C5")):
        ws[lbl].font = Font(size=10)
        ws[val].font = _INPUT_BLUE
        ws[val].fill = _INPUT_FILL
        ws[val].alignment = Alignment(horizontal="center")

    hdr_row = _INPUT_START - 1
    for j, (hdr, _key) in enumerate(INPUT_COLS):
        cell = ws.cell(row=hdr_row, column=2 + j, value=hdr)
        cell.font = _HDR_FONT; cell.fill = _HDR_FILL
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws.row_dimensions[hdr_row].height = 35.5
    for col, w in _INPUT_WIDTHS.items():
        ws.column_dimensions[col].width = w

    for i, rec in enumerate(records):
        r = _INPUT_START + i
        for j, (_hdr, key) in enumerate(INPUT_COLS):
            col = 2 + j
            v = _val(rec.get(key))
            cell = ws.cell(row=r, column=col, value=v)
            if isinstance(v, (date, datetime)):
                cell.number_format = "d-mmm-yy"
                cell.alignment = Alignment(horizontal="center")
            cell.font = _INPUT_BLUE
            cell.fill = _INPUT_FILL

    n = len(records)
    last_col = get_column_letter(1 + len(INPUT_COLS))
    ref = f"B{_INPUT_START - 1}:{last_col}{max(_INPUT_START + n - 1, _INPUT_START)}"
    tbl = Table(displayName="GrossDealLevelInput", ref=ref)
    # template tables carry no table style (no banding) — cell fills rule
    tbl.tableStyleInfo = TableStyleInfo(showRowStripes=True)
    ws.add_table(tbl)


def _dl_src_formula(spec: str, r: int, input_row: int) -> Any:
    if spec.startswith("in0:"):
        col = _input_col_letter(spec[4:])
        src = f"'Deal Level Inputs'!{col}{input_row}"
        return f'=IF({src}="",0,{src})'       # blank -> explicit 0 (EWL)
    if spec.startswith("in:"):
        col = _input_col_letter(spec[3:])
        src = f"'Deal Level Inputs'!{col}{input_row}"
        return f'=IF({src}="","",{src})'      # blank-safe link
    if spec.startswith("FT:"):
        return "=" + spec[3:]                 # template formula, verbatim
    if spec.startswith("F:"):
        return "=" + spec[2:].replace("{r}", str(r))
    return None


def _write_deal_list(ws, records: list[dict], gp: str,
                     track_record_date, currency: str) -> None:
    n = len(records)
    last_row = max(_DL_START + n - 1, _DL_START)

    # Header block (template rows 1–11: title, GP header, bucket helper tables,
    # count/loss-ratio helpers). "{LAST}" resolves to the last data row.
    for cell_spec in HEADER_BLOCK:
        v = cell_spec["v"]
        if isinstance(v, str) and v.startswith("="):
            v = v.replace("{LAST}", str(last_row))
        c = ws[cell_spec["ref"]]
        c.value = v
        if cell_spec.get("fmt"):
            c.number_format = cell_spec["fmt"]
        if cell_spec.get("bold"):
            c.font = Font(bold=True, size=16 if cell_spec["ref"] == "B2" else 10)
        elif isinstance(v, (int, float)):
            # adjustable bucket thresholds — input styling (template)
            c.font = _INPUT_BLUE
            c.fill = _INPUT_FILL
        else:
            c.font = _CALC_FONT
    # GP / as-of / currency — linked to the Deal Level Inputs tab.
    # linked (not typed-in) cells — plain formula look, black on white
    ws["C4"] = "='Deal Level Inputs'!C3"
    ws["C5"] = "='Deal Level Inputs'!C4"; ws["C5"].number_format = "d-mmm-yy"
    ws["C6"] = "='Deal Level Inputs'!C5"
    for ref in ("C4", "C5", "C6"):
        ws[ref].font = _CALC_FONT

    # Row 12: Input/Formula/Entry/Exit column tags (EWL edit)
    tag_row = _DL_START - 2
    for col_letter, tag in TAG_ROW:
        c = ws[f"{col_letter}{tag_row}"]
        c.value = tag
        c.font = Font(size=9, italic=True, color="808080")

    # Column headers (row 13) + data rows (14…)
    hdr_row = _DL_START - 1
    for j, (hdr, *_r) in enumerate(DEAL_LIST_COLS):
        cell = ws.cell(row=hdr_row, column=2 + j, value=hdr)
        cell.font = _HDR_FONT; cell.fill = _HDR_FILL
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws.row_dimensions[hdr_row].height = 35.5

    for i in range(n):
        r = _DL_START + i
        input_row = _INPUT_START + i
        for j, (hdr, spec, _key, _kind) in enumerate(DEAL_LIST_COLS):
            cell = ws.cell(row=r, column=2 + j,
                           value=_dl_src_formula(spec, r, input_row))
            fmt = DL_FORMATS.get(j)
            if fmt:
                cell.number_format = fmt
            # every Deal List data cell is a formula (links included) — plain
            # white/black; the only blue "input" cells on this tab are the
            # bucket-threshold numbers in the header block
            if hdr in _GRAY_COLS:               # key computed columns — grey
                cell.font = _CALC_FONT
                cell.fill = _GRAY_FILL
            else:
                cell.font = _CALC_FONT

    last_col = get_column_letter(1 + len(DEAL_LIST_COLS))
    tbl = Table(displayName="DealLevelInput", ref=f"B{_DL_START - 1}:{last_col}{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(showRowStripes=True)
    ws.add_table(tbl)


# ═══════════════════════════════════════════════════════════════════════════════
# Pivot plan — shared by the title writer (openpyxl) and the XML injector
# ═══════════════════════════════════════════════════════════════════════════════

def _g255(s):
    """Excel caps pivot cache item text at 255 chars (longer items ⇒
    "PivotTable cache" repair). Excel itself truncates on refresh, so
    truncated items are the refresh-stable representation."""
    return s if s is None or len(s) <= 255 else s[:255]


def _grouped_items(records: list[dict], header: str, key: int, kind: str) -> list:
    """Distinct axis items for a grouped field, canonically ordered."""
    if kind == "gn":     # numeric+text — numbers ascending, distinct strings last
        raw = [_rec_value(r, header, key) for r in records]
        nums = sorted({int(v) for v in (_cell_num(x) for x in raw) if v is not None})
        strs = sorted({_g255(s) for s in (_cell_str(x) for x in raw
                                          if _cell_num(x) is None) if s is not None})
        return list(nums) + strs
    present = []
    seen = set()
    for rec in records:
        s = _g255(_cell_str(_rec_value(rec, header, key)))
        if s and s not in seen:
            seen.add(s); present.append(s)
    canon = _CANONICAL_ORDER.get(header)
    if canon:
        ordered = [l for l in canon if l in seen]
        ordered += [s for s in present if s not in ordered]
        return ordered
    return present


def plan_pivots(records: list[dict]) -> list[dict]:
    """Compute each pivot's items, blank flag, data fields, and sheet anchor."""
    # Page-filter fields: cache item count per field (incl. the blank item) —
    # the pivotField <items> list must cover every shared item.
    page_meta = []
    for ph in _PAGE_FIELD_HEADERS:
        pidx = _dl_field_index(ph)
        _h, _s, pkey, pkind = DEAL_LIST_COLS[pidx]
        pitems = _grouped_items(records, ph, pkey, pkind)
        pblank = False if pkind == "gn" else any(
            _cell_str(_rec_value(r, ph, pkey)) is None for r in records)
        page_meta.append((ph, pidx, len(pitems) + (1 if pblank else 0)))

    plan = []
    # rows 5.. hold the in-sheet contents list; first section heading follows
    anchor = 5 + len(PIVOT_SPECS) + 3
    for title, header, extra in PIVOT_SPECS:
        idx = _dl_field_index(header)
        _h, _spec, key, kind = DEAL_LIST_COLS[idx]
        items = _grouped_items(records, header, key, kind)
        has_blank = False if kind == "gn" else any(
            _cell_str(_rec_value(r, header, key)) is None for r in records)
        # blank items are HIDDEN in the pivot (h="1" item filter) — only the
        # labelled items render; grand totals cover the visible items only
        n_axis = len(items)
        if extra == "with_ic":
            datafields = ["ic_sum", "count", "moic", "loss"]
        elif extra == "impaired":
            datafields = ["count", "moic", "impaired"]
        else:
            datafields = ["count", "moic", "loss"]
        empty = header in _DEBUG_GROUP_DISABLE   # all-blank axes still get a pivot
        height = 1 if empty else (n_axis + 3)     # Values + header + items + grand
        plan.append({
            "name": f"RL{len(plan) + 1}",
            "title": title, "header": header, "field_idx": idx, "kind": kind,
            "items": items, "has_blank": has_blank, "datafields": datafields,
            "anchor": anchor, "top": anchor + _PIVOT_GAP, "height": height,
            "empty": empty,
            "page_fields": [(ph, pi, pc) for ph, pi, pc in page_meta
                            if ph != header],
        })
        # title + filter head-room + max(pivot, chart) + gap
        anchor += 1 + _PIVOT_GAP + max(height, 15) + 3
    return plan


def _rl_toc_label(title: str) -> str:
    head, _, dim = title.partition(" by ")
    label = f"By {dim}" if dim else title
    if head not in ("Gross Returns & Loss Ratios", ""):
        label += f" ({head})"
    return label


def _write_rl_titles(ws, plan: list[dict]) -> None:
    ws["B2"] = "Return & Loss Ratios"; ws["B2"].font = _TITLE_FONT
    ws["B3"] = "All figures computed from the Deal List (pivot tables refresh on open)."
    ws["B3"].font = Font(italic=True, size=9)
    _write_mini_toc(ws, [(_rl_toc_label(p["title"]), p["anchor"] - 1) for p in plan],
                    start_row=5)
    for p in plan:
        c = ws.cell(row=p["anchor"] - 1, column=2, value=p["title"])
        c.font = _SECTION_FONT



_MOIC_FMT = "0.0\\x;\\(0.0\\x\\)"

# Chart geometry (template chart10): anchored at col K, ~11 cols × 14 rows.
_CHART_C0, _CHART_C1, _CHART_ROWS = 10, 21, 14


def _pivot_graph_cols(p: dict):
    """(count, moic, loss) column letters; loss covers the impaired variant."""
    dfs = p["datafields"]
    col_of = lambda f: get_column_letter(4 + dfs.index(f))
    loss_key = "impaired" if "impaired" in dfs else "loss"
    return col_of("count"), col_of("moic"), col_of(loss_key)


_DF_CAPTION = {"count": "Count", "moic": "MOIC", "loss": "Loss Ratio",
               "impaired": "Impaired Loss Ratio", "ic_sum": "Total Invested Capital"}
_PCT_FMT = "0%"


def _render_pivot_cells(ws, plan: list[dict], records: list[dict]) -> None:
    """Write each pivot's RENDERED output into the sheet (Excel's own saved
    layout: 'Values' caption, header row, item rows, Grand Total). Mac Excel
    does not reliably honour refreshOnLoad, so the tab must display without a
    refresh — exactly like the template file, whose sheets carry the values."""
    for p in plan:
        if p.get("empty"):
            continue
        header, key, kind = p["header"], None, p["kind"]
        for h, _s, k, kd in DEAL_LIST_COLS:
            if h == header:
                key = k
                break
        items = list(p["items"])                 # blank item hidden from pivot

        def _stats(match):
            n = tv = ic = lo = im = ii = 0.0
            cnt = 0
            for rec in records:
                rv = _rec_value(rec, header, key)
                v = (int(_cell_num(rv)) if (kind == "gn" and _cell_num(rv) is not None)
                     else _cell_str(rv))
                if v != match:
                    continue
                cnt += 1
                # sheet-consistent aggregates (same math as the calc fields
                # evaluate after a refresh reads the Deal List)
                tv += _cell_num(_sheet_tv(rec)) or 0
                ic += _cell_num(rec.get(16)) or 0
                lo += _cell_num(_rec_value(rec, "InvCapital in Loss Position", 0)) or 0
                im += _cell_num(_rec_value(rec, "Impaired\nValue", 0)) or 0
                ii += _cell_num(rec.get(16)) or 0
            return cnt, tv, ic, lo, im, ii

        top = p["top"]
        dfs = p["datafields"]
        # report-filter rows (Excel's saved layout: field name + "(All)"
        # immediately above the body, one separator row between)
        page_fields = p.get("page_fields", [])
        for k, (ph, _pi, _pc) in enumerate(page_fields):
            fr = top - 1 - len(page_fields) + k
            ws.cell(row=fr, column=3, value=ph).font = Font(size=10)
            ws.cell(row=fr, column=4, value="(All)").font = Font(size=10)
        ws.cell(row=top, column=4, value="Values").font = Font(size=10)
        hdr = ws.cell(row=top + 1, column=3, value="Row Labels"); hdr.font = Font(size=10)
        for j, f in enumerate(dfs):
            ws.cell(row=top + 1, column=4 + j, value=_DF_CAPTION[f]).font = Font(size=10)

        totals = [0.0] * 6
        # capture the full-data chart series alongside (charts carry literals)
        cats, moics, losses = [], [], []
        use_impaired = "impaired" in dfs
        for i, item in enumerate(items):
            r = top + 2 + i
            cnt, tv, ic, lo, im, ii = _stats(item)
            for t_i, val in enumerate((cnt, tv, ic, lo, im, ii)):
                totals[t_i] += val
            ws.cell(row=r, column=3, value=item)
            _fill_pivot_row(ws, r, dfs, cnt, tv, ic, lo, im, ii)
            cats.append(f"{item}\n{cnt}")
            moics.append((tv / ic) if ic else None)
            losses.append(((im if use_impaired else lo) / ic) if ic else None)
        p["chart_cats"], p["chart_moic"], p["chart_loss"] = cats, moics, losses
        gr = top + 2 + len(items)
        g = ws.cell(row=gr, column=3, value="Grand Total"); g.font = Font(bold=True, size=10)
        _fill_pivot_row(ws, gr, dfs, *totals, bold=True)


def _fill_pivot_row(ws, r, dfs, cnt, tv, ic, lo, im, ii, bold=False):
    vals = {"count": (cnt, "0"),
            "moic": ((tv / ic) if ic else None, _MOIC_FMT),
            "loss": ((lo / ic) if ic else None, _PCT_FMT),
            "impaired": ((im / ic) if ic else None, _PCT_FMT),
            "ic_sum": (ii, "#,##0")}
    for j, f in enumerate(dfs):
        v, fmt = vals[f]
        c = ws.cell(row=r, column=4 + j, value=v)
        c.number_format = fmt
        if bold:
            c.font = Font(bold=True, size=10)


def _write_rl_graphics(ws, plan: list[dict]) -> None:
    """Graph Label helper cells (template: H column, 'label\ncount' strings)."""
    if _DEBUG_NO_LABELS:
        return
    for p in plan:
        if p.get("empty"):
            continue
        # blank axis item (always the LAST item row) gets no label — it must
        # not appear in the charts (EWL: "no blank items showing")
        n_vis = len(p["items"])
        if n_vis == 0:
            continue
        top = p["top"]
        first, last = top + 2, top + 1 + n_vis
        count_col, _m, _l = _pivot_graph_cols(p)
        ws.cell(row=top + 1, column=8, value="Graph Label").font = Font(bold=True, size=10)
        for r in range(first, last + 1):
            ws.cell(row=r, column=8,
                    value=f"=CONCATENATE(C{r},CHAR(10),{count_col}{r})")


# ═══════════════════════════════════════════════════════════════════════════════
# Pivot XML (Excel-verified structure)
# ═══════════════════════════════════════════════════════════════════════════════

_CALC_FIELDS = [
    ("CalcMOIC", "'Total_x000a_Value'/'Total Invested Capital (mlns)'"),
    ("CalcLossRatio", "'InvCapital in Loss Position'/'Total Invested Capital (mlns)'"),
    ("CalcImpairedLossRatio", "'Impaired_x000a_Value'/'Total Invested Capital (mlns)'"),
    # IC-weighted ratios (template cache 3/4 calc fields, verbatim semantics)
    ("CalcICWeightedRevenueCAGR", "WghtdRevCAGR/AdjInvCapRevenueCAGR"),
    ("CalcWeightedEBITDACAGR", "WghtdEBITDAcagr/AdjInvCapEBITDAcagr"),
    ("CalcWeightedEntryMargin", "'Wgtd Entry EBITDA Margin'/AdjInvCapEBITDAMargin"),
    ("CalcWeightedExitMargin", "'Wgtd Current EBITDA Margin'/AdjInvCapEBITDAMargin"),
    ("CalcWghtdEntryMultiple", "WgtdEntryMultiple/AdjInvCapMultiples"),
    ("CalcWghtExitMultiple", "WghtdExitMultiple/AdjInvCapMultiples"),
    ("CalcWghtdEntryLeverage", "WghtdEntryLeverage/AdjInvCapLeverage"),
    ("CalcWghtdExitLeverage", "WghtdExitLeverage/AdjInvCapLeverage"),
    ("CalcWghtedHoldPeriod", "'Wghtd Hold Period'/'Total Invested Capital (mlns)'"),
]


def _fname(header: str) -> str:
    """cacheField/dataField name encoding: OOXML escapes control characters
    as _xHHHH_ (e.g. newline -> _x000a_) — Excel rejects real newlines here."""
    return _esc(header.replace("\n", "_x000a_").replace("\r", "_x000d_"))


def _build_cache_parts(records: list[dict]) -> tuple[bytes, bytes, dict]:
    """Cache definition + populated records. Returns (def_xml, rec_xml, item_maps)."""
    n_fields = len(DEAL_LIST_COLS) + len(_CALC_FIELDS)
    item_maps: dict[int, dict] = {}              # field_idx -> {value: item_idx}

    cf = []
    for idx, (header, _spec, key, kind) in enumerate(DEAL_LIST_COLS):
        if header in _DEBUG_GROUP_DISABLE:
            kind = "s"
        if kind == "b":
            # header-only column (manual entry) — sheet cells are blank, so
            # the cache must ship blanks too (Excel re-reads the sheet on
            # refresh and the two must agree)
            cf.append(f'<cacheField name="{_fname(header)}" numFmtId="0">'
                      f'<sharedItems containsBlank="1"/></cacheField>')
        elif kind in ("g", "gn"):
            items = _grouped_items(records, header, key, kind)
            has_blank = False if kind == "gn" else any(
                _cell_str(_rec_value(r, header, key)) is None for r in records)
            item_maps[idx] = {v: i for i, v in enumerate(items)}
            cnt = len(items) + (1 if has_blank else 0)
            blank = ' containsBlank="1"' if has_blank else ""
            if kind == "gn":
                nums = [v for v in items if not isinstance(v, str)]
                strs = [v for v in items if isinstance(v, str)]
                gn_blank = any(_rec_value(r, header, key) is None for r in records)
                body = "".join(
                    (f'<s v="{_esc(v)}"/>' if isinstance(v, str) else f'<n v="{_numstr(float(v))}"/>')
                    for v in items)
                if nums:
                    # numeric (optionally mixed with strings) — template shape
                    mixed = (' containsMixedTypes="1"' if strs else
                             ' containsSemiMixedTypes="0" containsString="0"')
                    mnmx = (f' minValue="{_numstr(float(min(nums)))}"'
                            f' maxValue="{_numstr(float(max(nums)))}"')
                    attrs = (f'{mixed} containsNumber="1" containsInteger="1"'
                             f'{mnmx} count="{len(items)}"')
                elif strs:
                    # no numeric values in the data ("n/a" only) — declaring
                    # containsNumber with zero <n> items = cache repair
                    if gn_blank:
                        body += "<m/>"
                    attrs = ((' containsBlank="1"' if gn_blank else "")
                             + f' count="{len(items) + (1 if gn_blank else 0)}"')
                else:
                    # no values at all — all-blank grouped field
                    body = "<m/>"
                    attrs = ' containsBlank="1" count="1"'
            else:
                body = "".join(f'<s v="{_esc(v)}"/>' for v in items)
                attrs = f'{blank} count="{cnt}"'
            body += "<m/>" if has_blank else ""
            cf.append(f'<cacheField name="{_fname(header)}" numFmtId="0">'
                      f'<sharedItems{attrs}>{body}</sharedItems></cacheField>')
        elif kind == "dt":
            ds = [v for v in (_rec_value(r, header, key) for r in records)
                  if isinstance(v, (date, datetime))]
            dblank = ' containsBlank="1"' if len(ds) < len(records) else ""
            if ds:
                _iso = lambda d: (d if isinstance(d, datetime) else datetime(d.year, d.month, d.day)).isoformat()
                cf.append(f'<cacheField name="{_fname(header)}" numFmtId="0">'
                          f'<sharedItems containsNonDate="0" containsDate="1" containsString="0"{dblank} '
                          f'minDate="{_iso(min(ds))}" maxDate="{_iso(max(ds))}"/></cacheField>')
            else:
                cf.append(f'<cacheField name="{_fname(header)}" numFmtId="0">'
                          f'<sharedItems containsBlank="1"/></cacheField>')
        elif kind == "n":
            # value source MUST match the records writer (_rec_value applies
            # _SHEET_CALC) — a field that is raw-blank but sheet-computed
            # (e.g. 'Total IC mlns for Buckets' = 0 when IC missing) would
            # otherwise declare blank-only while the records ship numbers,
            # which Excel repairs away as "PivotTable cache"
            nums = [x for x in (_cell_num(_rec_value(r, header, key))
                                for r in records) if x is not None]
            if nums:
                cf.append(f'<cacheField name="{_fname(header)}" numFmtId="0">'
                          f'<sharedItems containsString="0" containsBlank="1" containsNumber="1" '
                          f'minValue="{_numstr(min(nums))}" maxValue="{_numstr(max(nums))}"/></cacheField>')
            else:
                cf.append(f'<cacheField name="{_fname(header)}" numFmtId="0">'
                          f'<sharedItems containsBlank="1"/></cacheField>')
        else:
            cf.append(f'<cacheField name="{_fname(header)}" numFmtId="0">'
                      f'<sharedItems containsBlank="1"/></cacheField>')
    for name, formula in _CALC_FIELDS:
        cf.append(f'<cacheField name="{name}" numFmtId="0" formula="{_esc(formula)}" databaseField="0"/>')

    cache_def = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<pivotCacheDefinition xmlns="{_NS}" xmlns:r="{_R}" r:id="rId1" '
        'refreshOnLoad="1" refreshedBy="TR Automation" createdVersion="8" '
        f'refreshedVersion="8" minRefreshableVersion="3" recordCount="{len(records)}">'
        '<cacheSource type="worksheet"><worksheetSource name="DealLevelInput"/></cacheSource>'
        f'<cacheFields count="{n_fields}">{"".join(cf)}</cacheFields>'
        '</pivotCacheDefinition>')

    rows_xml = []
    for rec in records:
        cells = []
        for idx, (header, _spec, key, kind) in enumerate(DEAL_LIST_COLS):
            if header in _DEBUG_GROUP_DISABLE:
                kind = "s"
            if kind == "b":
                cells.append("<m/>")
                continue
            rv = _rec_value(rec, header, key)
            if kind in ("g", "gn"):
                if kind == "gn":
                    n = _cell_num(rv)
                    v = int(n) if n is not None else _g255(_cell_str(rv))
                else:
                    v = _g255(_cell_str(rv))
                imap = item_maps[idx]
                cells.append(f'<x v="{imap[v]}"/>' if v in imap else "<m/>")
            elif kind == "dt":
                if isinstance(rv, (date, datetime)):
                    dd = rv if isinstance(rv, datetime) else datetime(rv.year, rv.month, rv.day)
                    cells.append(f'<d v="{dd.isoformat()}"/>')
                else:
                    cells.append("<m/>")
            elif kind == "n":
                x = _cell_num(rv)
                cells.append(f'<n v="{_numstr(x)}"/>' if x is not None else "<m/>")
            else:
                s = _g255(_cell_str(rv))
                cells.append(f'<s v="{_esc(s)}"/>' if s is not None else "<m/>")
        rows_xml.append("<r>" + "".join(cells) + "</r>")
    cache_rec = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                 f'<pivotCacheRecords xmlns="{_NS}" xmlns:r="{_R}" count="{len(records)}">'
                 f'{"".join(rows_xml)}</pivotCacheRecords>')
    return cache_def.encode("utf8"), cache_rec.encode("utf8"), item_maps


_DF_XML = {
    "count":    '<dataField name="Count" fld="0" subtotal="count" baseField="0" baseItem="0"/>',
    "ic_sum":   '<dataField name="Total Invested Capital" fld="{ic}" baseField="0" baseItem="0"/>',
    "moic":     '<dataField name="MOIC" fld="{moic}" baseField="0" baseItem="0" numFmtId="217"/>',
    "loss":     '<dataField name="Loss Ratio" fld="{loss}" baseField="0" baseItem="0" numFmtId="9"/>',
    "impaired": '<dataField name="Impaired Loss Ratio" fld="{imp}" baseField="0" baseItem="0" numFmtId="9"/>',
}


def _build_pivot_table_xml(p: dict) -> bytes:
    n_base = len(DEAL_LIST_COLS)
    n_fields = n_base + len(_CALC_FIELDS)
    fld_ids = {"ic": _dl_field_index("Total Invested Capital (mlns)"),
               "moic": n_base, "loss": n_base + 1, "imp": n_base + 2}

    n_axis = len(p["items"])                     # visible items (blank hidden)
    datafields = p["datafields"]
    n_df = len(datafields)

    # which fields need dataField="1"
    df_src = {0}
    if "ic_sum" in datafields:
        df_src.add(fld_ids["ic"])
    if "moic" in datafields:
        df_src.add(fld_ids["moic"])
    if "loss" in datafields:
        df_src.add(fld_ids["loss"])
    if "impaired" in datafields:
        df_src.add(fld_ids["imp"])

    page_fields = p.get("page_fields", [])
    page_idx = {pi: pc for _ph, pi, pc in page_fields}

    pfs = []
    for i in range(n_fields):
        if i == p["field_idx"]:
            items = "".join(f'<item x="{k}"/>' for k in range(n_axis))
            n_items = n_axis
            if p["has_blank"]:                   # blank cache item: hidden
                items += f'<item h="1" x="{n_axis}"/>'
                n_items += 1
            items += '<item t="default"/>'
            pfs.append(f'<pivotField axis="axisRow" showAll="0">'
                       f'<items count="{n_items + 1}">{items}</items></pivotField>')
        elif i in page_idx:
            cnt = page_idx[i]
            items = "".join(f'<item x="{k}"/>' for k in range(cnt)) + '<item t="default"/>'
            pfs.append(f'<pivotField axis="axisPage" showAll="0">'
                       f'<items count="{cnt + 1}">{items}</items></pivotField>')
        elif i in df_src:
            pfs.append('<pivotField dataField="1" showAll="0"/>')
        else:
            pfs.append('<pivotField showAll="0"/>')

    row_items = "".join('<i><x v="%d"/></i>' % k for k in range(n_axis)) + '<i t="grand"><x/></i>'
    top = p["top"]
    last_col = get_column_letter(3 + n_df)       # C=labels + n data columns
    page_attr = (f' rowPageCount="{len(page_fields)}" colPageCount="1"'
                 if page_fields else "")
    loc = (f'<location ref="C{top}:{last_col}{top + n_axis + 2}" '
           f'firstHeaderRow="1" firstDataRow="2" firstDataCol="1"{page_attr}/>')
    ci = "<i><x/></i>" + "".join(f'<i i="{k}"><x v="{k}"/></i>' for k in range(1, n_df))
    col_block = ('<colFields count="1"><field x="-2"/></colFields>'
                 f'<colItems count="{n_df}">{ci}</colItems>')
    page_block = ("" if not page_fields else
                  f'<pageFields count="{len(page_fields)}">'
                  + "".join(f'<pageField fld="{pi}" hier="-1"/>'
                            for _ph, pi, _pc in page_fields)
                  + '</pageFields>')
    dfs = "".join(_DF_XML[d].format(**fld_ids) for d in datafields)

    xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<pivotTableDefinition xmlns="{_NS}" xmlns:r="{_R}" name="{p["name"]}" '
        'cacheId="1" applyNumberFormats="0" applyBorderFormats="0" applyFontFormats="0" '
        'applyPatternFormats="0" applyAlignmentFormats="0" applyWidthHeightFormats="1" '
        'dataCaption="Values" updatedVersion="8" minRefreshableVersion="3" createdVersion="8" '
        'indent="0" outline="1" outlineData="1" multipleFieldFilters="0">'
        + loc
        + f'<pivotFields count="{n_fields}">{"".join(pfs)}</pivotFields>'
        + f'<rowFields count="1"><field x="{p["field_idx"]}"/></rowFields>'
        + f'<rowItems count="{n_axis + 1}">{row_items}</rowItems>'
        + col_block
        + page_block
        + f'<dataFields count="{n_df}">{dfs}</dataFields>'
        + '<pivotTableStyleInfo name="PivotStyleLight16" showRowHeaders="1" showColHeaders="1" '
          'showRowStripes="0" showColStripes="0" showLastColumn="1"/>'
        + '</pivotTableDefinition>')
    return xml.encode("utf8")


# ═══════════════════════════════════════════════════════════════════════════════
# Return Dispersion tab (template sheet 3 — two sections per user spec:
# Gross MOIC dispersion on top, Gross IRR dispersion below, one chart each)
# ═══════════════════════════════════════════════════════════════════════════════
# Pivot shape (template pivotTable19): compact header (no σ-Values row —
# location firstHeaderRow="0" firstDataRow="1"), grandTotalCaption="Total",
# axis field caption renamed ("MOIC"/"IRR"), dataFields Count (numFmt 1) +
# "% IC" = Sum of Total Invested Capital showDataAs="percentOfCol" (numFmt 9),
# the standard 3 report filters, Graph Label helper col F, and a clustered
# column chart with per-point colours (red → grey → light/dark green).

# (title, axis header, caption, averaged metric header)
# Each dispersion pivot shows Count | % IC | Average of the metric (user spec):
# % IC = capital share per bucket, the average = the actual metric level.
RD_SPECS: list[tuple[str, str, str, str]] = [
    ("Gross MOIC", "MOIC Buckets", "MOIC", "Gross\nMOIC"),
    ("Gross IRR",  "IRR Buckets",  "IRR",  "Gross\nIRR"),
]
# render format + pivot numFmtId for each averaged metric
_RD_AVG_STYLE = {
    "Gross\nMOIC": ("Avg Gross MOIC", _MOIC_FMT, "217"),
    "Gross\nIRR":  ("Avg Gross IRR", "0.0%", "218"),
}


def plan_rd(records: list[dict]) -> list[dict]:
    page_meta = []
    for ph in _PAGE_FIELD_HEADERS:
        pidx = _dl_field_index(ph)
        _h, _s, pkey, pkind = DEAL_LIST_COLS[pidx]
        pitems = _grouped_items(records, ph, pkey, pkind)
        pblank = False if pkind == "gn" else any(
            _cell_str(_rec_value(r, ph, pkey)) is None for r in records)
        page_meta.append((ph, pidx, len(pitems) + (1 if pblank else 0)))

    plan = []
    # rows 8.. hold the in-sheet contents list (below the meta block)
    anchor = 8 + len(RD_SPECS) + 3
    for title, header, caption, metric_header in RD_SPECS:
        idx = _dl_field_index(header)
        _h, _s, key, kind = DEAL_LIST_COLS[idx]
        m_idx = _dl_field_index(metric_header)
        _mh, _ms, m_key, _mk = DEAL_LIST_COLS[m_idx]
        items = _grouped_items(records, header, key, kind)
        # per bucket: count, IC share of the visible total, mean of the metric
        stats = []
        total_ic = 0.0
        for it in items:
            cnt = 0
            ic = 0.0
            mvals: list[float] = []
            for rec in records:
                if _cell_str(_rec_value(rec, header, key)) == it:
                    cnt += 1
                    ic += _cell_num(rec.get(16)) or 0.0
                    v = _cell_num(_rec_value(rec, metric_header, m_key))
                    if v is not None:
                        mvals.append(v)
            stats.append([it, cnt, ic, mvals])
            total_ic += ic
        rows = [(it, cnt,
                 (ic / total_ic) if total_ic else None,
                 (sum(mvals) / len(mvals)) if mvals else None)
                for it, cnt, ic, mvals in stats]
        all_m = [v for _i, _c, _ic, mvals in stats for v in mvals]
        total_avg = (sum(all_m) / len(all_m)) if all_m else None
        height = len(items) + 3                   # Values + header + items + Total
        plan.append({
            "type": "rd", "name": f"RD{len(plan) + 1}", "title": title,
            "header": header, "caption": caption, "field_idx": idx,
            "metric_header": metric_header, "metric_idx": m_idx,
            "items": items, "rows": rows,
            "total_count": sum(c for _i, c, _p, _a in rows),
            "total_avg": total_avg,
            "page_fields": page_meta,
            "anchor": anchor, "top": anchor + _PIVOT_GAP, "height": height,
        })
        anchor += 1 + _PIVOT_GAP + max(height, _CHART_ROWS + 1) + 3
    return plan


def _write_rd_sheet(ws, rd_plan: list[dict]) -> None:
    ws["B2"] = "Return Dispersion"; ws["B2"].font = _TITLE_FONT
    ws["B4"] = "Sponsor/GP:"; ws["C4"] = "='Deal List'!$C$4"
    ws["B5"] = "As of Date:"; ws["C5"] = "='Deal List'!$C$5"
    ws["C5"].number_format = "d-mmm-yy"
    ws["B6"] = "Currency:";   ws["C6"] = "='Deal List'!$C$6"
    for ref in ("B4", "B5", "B6", "C4", "C5", "C6"):
        ws[ref].font = _CALC_FONT
    _write_mini_toc(ws, [(p["title"], p["anchor"] - 1) for p in rd_plan],
                    start_row=8)

    for p in rd_plan:
        ws.cell(row=p["anchor"] - 1, column=2, value=p["title"]).font = _SECTION_FONT
        top = p["top"]
        for k, (ph, _pi, _pc) in enumerate(p["page_fields"]):
            fr = top - 1 - len(p["page_fields"]) + k
            ws.cell(row=fr, column=2, value=ph).font = Font(size=10)
            ws.cell(row=fr, column=3, value="(All)").font = Font(size=10)
        avg_hdr, avg_fmt, _avg_id = _RD_AVG_STYLE[p["metric_header"]]
        # RLR-proven refresh-stable layout: σ-Values caption row, then header
        ws.cell(row=top, column=3, value="Values").font = Font(size=10)
        ws.cell(row=top + 1, column=2, value=p["caption"]).font = Font(size=10)
        ws.cell(row=top + 1, column=3, value="Count").font = Font(size=10)
        ws.cell(row=top + 1, column=4, value="% IC").font = Font(size=10)
        ws.cell(row=top + 1, column=5, value=avg_hdr).font = Font(size=10)
        ws.cell(row=top + 1, column=7, value="Graph Label").font = Font(bold=True, size=10)
        for i, (it, cnt, pct, avg) in enumerate(p["rows"]):
            r = top + 2 + i
            ws.cell(row=r, column=2, value=it).font = Font(size=10)
            c = ws.cell(row=r, column=3, value=cnt)
            c.number_format = "0"; c.font = Font(size=10)
            c = ws.cell(row=r, column=4, value=pct)
            c.number_format = "0%"; c.font = Font(size=10)
            c = ws.cell(row=r, column=5, value=avg)
            c.number_format = avg_fmt; c.font = Font(size=10)
            ws.cell(row=r, column=7,
                    value=f"=CONCATENATE(B{r},CHAR(10),C{r})")
        gr = top + 2 + len(p["rows"])
        g = ws.cell(row=gr, column=2, value="Total"); g.font = Font(bold=True, size=10)
        c = ws.cell(row=gr, column=3, value=p["total_count"])
        c.number_format = "0"; c.font = Font(bold=True, size=10)
        c = ws.cell(row=gr, column=4, value=(1 if p["total_count"] else None))
        c.number_format = "0%"; c.font = Font(bold=True, size=10)
        c = ws.cell(row=gr, column=5, value=p["total_avg"])
        c.number_format = avg_fmt; c.font = Font(bold=True, size=10)
        ws.cell(row=gr, column=7, value=f"=CONCATENATE(B{gr},CHAR(10),C{gr})")


def _build_rd_pivot_xml(p: dict) -> bytes:
    n_fields = len(DEAL_LIST_COLS) + len(_CALC_FIELDS)
    ic_fld = _dl_field_index("Total Invested Capital (mlns)")
    avg_hdr, _avg_fmt, avg_id = _RD_AVG_STYLE[p["metric_header"]]
    avg_fld = p["metric_idx"]
    n_axis = len(p["items"])
    page_idx = {pi: pc for _ph, pi, pc in p["page_fields"]}

    pfs = []
    for i in range(n_fields):
        if i == p["field_idx"]:
            items = "".join(f'<item x="{k}"/>' for k in range(n_axis)) + '<item t="default"/>'
            pfs.append(f'<pivotField axis="axisRow" showAll="0">'
                       f'<items count="{n_axis + 1}">{items}</items></pivotField>')
        elif i in page_idx:
            cnt = page_idx[i]
            items = "".join(f'<item x="{k}"/>' for k in range(cnt)) + '<item t="default"/>'
            pfs.append(f'<pivotField axis="axisPage" showAll="0">'
                       f'<items count="{cnt + 1}">{items}</items></pivotField>')
        elif i in (0, ic_fld, avg_fld):
            pfs.append('<pivotField dataField="1" showAll="0"/>')
        else:
            pfs.append('<pivotField showAll="0"/>')

    top = p["top"]
    row_items = "".join('<i><x v="%d"/></i>' % k for k in range(n_axis)) + '<i t="grand"><x/></i>'
    xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        + _pc_root(p["name"], ' grandTotalCaption="Total"'
                             f' rowHeaderCaption="{_esc(p["caption"])}"')
        + f'<location ref="B{top}:E{top + n_axis + 2}" '
          'firstHeaderRow="1" firstDataRow="2" firstDataCol="1" '
          'rowPageCount="3" colPageCount="1"/>'
        + f'<pivotFields count="{n_fields}">{"".join(pfs)}</pivotFields>'
        + f'<rowFields count="1"><field x="{p["field_idx"]}"/></rowFields>'
        + f'<rowItems count="{n_axis + 1}">{row_items}</rowItems>'
        + '<colFields count="1"><field x="-2"/></colFields>'
        + '<colItems count="3"><i><x/></i><i i="1"><x v="1"/></i>'
          '<i i="2"><x v="2"/></i></colItems>'
        + '<pageFields count="3">'
        + "".join(f'<pageField fld="{pi}" hier="-1"/>' for _ph, pi, _pc in p["page_fields"])
        + '</pageFields>'
        + '<dataFields count="3">'
          '<dataField name="Count" fld="0" subtotal="count" baseField="0" baseItem="0" numFmtId="1"/>'
          f'<dataField name="% IC" fld="{ic_fld}" showDataAs="percentOfCol" '
          'baseField="0" baseItem="0" numFmtId="9"/>'
          f'<dataField name="{_esc(avg_hdr)}" fld="{avg_fld}" subtotal="average" '
          f'baseField="0" baseItem="0" numFmtId="{avg_id}"/>'
        + '</dataFields>'
        + _PC_STYLE + '</pivotTableDefinition>')
    return xml.encode("utf8")


# ═══════════════════════════════════════════════════════════════════════════════
# Portfolio Construction tab (template sheet 4)
# ═══════════════════════════════════════════════════════════════════════════════
# Layout mirrors the IO template: two Fund×dimension matrix pivots (Sum of
# Total Invested Capital shown as % of row, Status report filter, no grand
# column) each with a percent-stacked column chart, then a "Deal Count
# Attributes" section of Count-of-Company pivots (sorted descending) with a
# pie chart each; a bare total-count pivot sits mid-section (no chart — the
# template's sixth pie points at an empty range, a stale artifact).

PC_MATRIX_SPECS: list[tuple[str, str]] = [
    ("Invested Capital by Fund and Sector",    "Sector"),
    ("Invested Capital by Fund and Geography", "Geography"),
]
# None = the bare total-count pivot
PC_COUNT_ORDER: list = ["Sector", "Geography", None,
                        "Transaction Type", "GP Role", "Process Type"]

_PC_PCT_FMT = '0%;\\(0%\\);"-"'
_PC_CHART_ROWS = 11          # template stacked/pie anchors span ~10-11 rows
# template stacked-series fill palette (cycled)
_PC_STACK_FILLS = [
    '<a:schemeClr val="accent1"><a:lumMod val="20000"/><a:lumOff val="80000"/></a:schemeClr>',
    '<a:schemeClr val="tx2"><a:lumMod val="60000"/><a:lumOff val="40000"/></a:schemeClr>',
    '<a:srgbClr val="FFCCCC"/>',
    '<a:schemeClr val="accent3"><a:lumMod val="20000"/><a:lumOff val="80000"/></a:schemeClr>',
    '<a:schemeClr val="accent6"><a:lumMod val="20000"/><a:lumOff val="80000"/></a:schemeClr>',
    '<a:schemeClr val="accent4"><a:lumMod val="40000"/><a:lumOff val="60000"/></a:schemeClr>',
]


def _pc_matrix_stats(records: list[dict], col_key: int):
    """(cell, row, col, total) invested-capital sums keyed by fund/dimension."""
    cell: dict = {}; row: dict = {}; col: dict = {}; total = 0.0
    for rec in records:
        f = _cell_str(rec.get(2))
        c = _cell_str(rec.get(col_key))
        ic = _cell_num(rec.get(16)) or 0.0
        cell[(f, c)] = cell.get((f, c), 0.0) + ic
        row[f] = row.get(f, 0.0) + ic
        col[c] = col.get(c, 0.0) + ic
        total += ic
    return cell, row, col, total


def plan_pc(records: list[dict]) -> list[dict]:
    """Plan the Portfolio Construction pivots (geometry + items)."""
    status_idx = _dl_field_index("Status")
    _h, _s, skey, skind = DEAL_LIST_COLS[status_idx]
    status_items = _grouped_items(records, "Status", skey, skind)
    status_blank = any(_cell_str(_rec_value(r, "Status", skey)) is None for r in records)
    status_cnt = len(status_items) + (1 if status_blank else 0)

    plan: list[dict] = []
    # rows 8.. hold the in-sheet contents list (below the meta block)
    anchor = 8 + len(PC_MATRIX_SPECS) + len(PC_COUNT_ORDER) + 3
    for title, col_header in PC_MATRIX_SPECS:
        cidx = _dl_field_index(col_header)
        _ch, _cs, ckey, ckind = DEAL_LIST_COLS[cidx]
        fidx = _dl_field_index("Fund")
        _fh, _fs, fkey, fkind = DEAL_LIST_COLS[fidx]

        funds = _grouped_items(records, "Fund", fkey, fkind)
        fund_blank = any(_cell_str(r.get(fkey)) is None for r in records)
        cell_s, _row_s, col_s, _total = _pc_matrix_stats(records, ckey)
        col_cache_items = _grouped_items(records, col_header, ckey, ckind)
        # template look: columns ascending by grand-total share
        cols = sorted(col_cache_items, key=lambda c: col_s.get(c, 0.0))
        col_blank = any(_cell_str(r.get(ckey)) is None for r in records)

        # visible-only percentages: blank fund row and blank column are HIDDEN
        # in the pivot, so every denominator covers labelled items only
        matrix_rows = []
        for f in funds:
            denom = sum(cell_s.get((f, c), 0.0) for c in cols)
            matrix_rows.append(
                (f, [((cell_s.get((f, c), 0.0) / denom) if denom else None)
                     for c in cols]))
        gtot = sum(cell_s.get((f, c), 0.0) for f in funds for c in cols)
        matrix_grand = [((sum(cell_s.get((f, c), 0.0) for f in funds) / gtot)
                         if gtot else None) for c in cols]

        n_rows = len(funds)
        height = n_rows + 3
        plan.append({
            "type": "matrix", "name": f"PC{len(plan) + 1}", "title": title,
            "col_header": col_header, "col_idx": cidx, "col_key": ckey,
            "fund_idx": fidx, "fund_key": fkey,
            "funds": funds, "fund_blank": fund_blank,
            "cols": cols, "col_blank": col_blank,
            "col_cache_items": col_cache_items,
            "matrix_rows": matrix_rows, "matrix_grand": matrix_grand,
            "status_idx": status_idx, "status_cnt": status_cnt,
            "anchor": anchor, "top": anchor + _PIVOT_GAP, "height": height,
        })
        anchor += 1 + _PIVOT_GAP + max(height, _PC_CHART_ROWS + 1) + 3

    plan.append({"type": "section", "title": "Deal Count Attributes",
                 "anchor": anchor})
    top = anchor + 5                              # 5 blank rows below heading
    for header in PC_COUNT_ORDER:
        if header is None:
            plan.append({"type": "total", "name": f"PC{len(plan)}",
                         "top": top, "height": 2})
            top += 2 + 6
            continue
        cidx = _dl_field_index(header)
        _ch, _cs, ckey, ckind = DEAL_LIST_COLS[cidx]
        # "Count of Company" over the SHEET counts every data row: the Company
        # column holds ""-string link formulas, which pivot counts treat as
        # non-blank — so blank-company records count too (refresh-consistent).
        counts: dict = {}
        for rec in records:
            counts[_cell_str(rec.get(ckey))] = counts.get(_cell_str(rec.get(ckey)), 0) + 1
        items = [v for v in sorted((k for k in counts if k is not None),
                                   key=lambda k: -counts[k])]
        has_blank = None in counts
        n_rows = len(items)                      # blank row hidden
        height = n_rows + 2
        plan.append({
            "type": "count", "name": f"PC{len(plan)}", "header": header,
            "field_idx": cidx, "key": ckey, "items": items,
            "cache_items": _grouped_items(records, header, ckey, ckind),
            "has_blank": has_blank, "counts": counts,
            "top": top, "height": height,
        })
        top += max(height, _PC_CHART_ROWS + 1) + 6
    return plan


def _write_pc_sheet(ws, pc_plan: list[dict], records: list[dict]) -> None:
    ws["B2"] = "Portfolio Construction"; ws["B2"].font = _TITLE_FONT
    ws["B4"] = "Sponsor/GP:"; ws["C4"] = "='Deal List'!$C$4"
    ws["B5"] = "As of Date:"; ws["C5"] = "='Deal List'!$C$5"
    ws["C5"].number_format = "d-mmm-yy"
    ws["B6"] = "Currency:";   ws["C6"] = "='Deal List'!$C$6"
    for ref in ("B4", "B5", "B6"):
        ws[ref].font = Font(size=10)
    for ref in ("C4", "C5", "C6"):                # links — plain formula look
        ws[ref].font = _CALC_FONT
    ws.column_dimensions["C"].width = 26

    toc_entries = []
    for p in pc_plan:
        if p["type"] == "matrix":
            label = p["title"].replace("Invested Capital by", "By") + " (Invested Capital)"
            toc_entries.append((label, p["anchor"] - 1))
        elif p["type"] == "count":
            toc_entries.append((f"By {p['header']} (Deal Count)", p["top"]))
        elif p["type"] == "total":
            toc_entries.append(("Total Deal Count", p["top"]))
    _write_mini_toc(ws, toc_entries, start_row=8)

    for p in pc_plan:
        if p["type"] == "section":
            c = ws.cell(row=p["anchor"] - 1, column=2, value=p["title"])
            c.font = _SECTION_FONT
            continue

        if p["type"] == "matrix":
            c = ws.cell(row=p["anchor"] - 1, column=2, value=p["title"])
            c.font = _SECTION_FONT
            top = p["top"]
            ws.cell(row=top - 2, column=3, value="Status").font = Font(size=10)
            ws.cell(row=top - 2, column=4, value="(All)").font = Font(size=10)
            ws.cell(row=top, column=3,
                    value="Sum of Total Invested Capital (mlns)").font = Font(size=10)
            ws.cell(row=top, column=4, value="Column Labels").font = Font(size=10)
            ws.cell(row=top + 1, column=3, value="Fund").font = Font(size=10)
            for k, cv in enumerate(p["cols"]):
                ws.cell(row=top + 1, column=4 + k, value=cv).font = Font(size=10)
            for i, (fv, pcts) in enumerate(p["matrix_rows"]):
                r = top + 2 + i
                ws.cell(row=r, column=3, value=fv).font = Font(size=10)
                for k, v in enumerate(pcts):
                    cc = ws.cell(row=r, column=4 + k, value=v)
                    cc.number_format = _PC_PCT_FMT; cc.font = Font(size=10)
            gr = top + 2 + len(p["matrix_rows"])
            g = ws.cell(row=gr, column=3, value="Grand Total")
            g.font = Font(bold=True, size=10)
            for k, v in enumerate(p["matrix_grand"]):
                cc = ws.cell(row=gr, column=4 + k, value=v)
                cc.number_format = _PC_PCT_FMT; cc.font = Font(bold=True, size=10)

        elif p["type"] == "count":
            top = p["top"]
            ws.cell(row=top, column=3, value="Row Labels").font = Font(size=10)
            ws.cell(row=top, column=4, value="Count of Company").font = Font(size=10)
            for i, iv in enumerate(p["items"]):
                r = top + 1 + i
                ws.cell(row=r, column=3, value=iv).font = Font(size=10)
                cc = ws.cell(row=r, column=4, value=p["counts"].get(iv, 0))
                cc.number_format = "0"; cc.font = Font(size=10)
            gr = top + 1 + len(p["items"])
            g = ws.cell(row=gr, column=3, value="Grand Total")
            g.font = Font(bold=True, size=10)
            # grand total covers VISIBLE items only (blank row is hidden)
            cc = ws.cell(row=gr, column=4,
                         value=sum(p["counts"][i] for i in p["items"]))
            cc.number_format = "0"; cc.font = Font(bold=True, size=10)

        elif p["type"] == "total":
            top = p["top"]
            n = len(records)                      # every row counts (see above)
            ws.cell(row=top, column=3, value="Count of Company").font = Font(size=10)
            cc = ws.cell(row=top + 1, column=3, value=n)
            cc.number_format = "0"; cc.font = Font(size=10)


_PC_STYLE = ('<pivotTableStyleInfo name="PivotStyleLight16" showRowHeaders="1" '
             'showColHeaders="1" showRowStripes="0" showColStripes="0" showLastColumn="1"/>')


def _pc_root(name: str, extra: str = "") -> str:
    return (f'<pivotTableDefinition xmlns="{_NS}" xmlns:r="{_R}" name="{name}" '
            'cacheId="1" applyNumberFormats="0" applyBorderFormats="0" applyFontFormats="0" '
            'applyPatternFormats="0" applyAlignmentFormats="0" applyWidthHeightFormats="1" '
            'dataCaption="Values" updatedVersion="8" minRefreshableVersion="3" createdVersion="8" '
            f'indent="0" outline="1" outlineData="1" multipleFieldFilters="0"{extra}>')


def _build_pc_matrix_xml(p: dict) -> bytes:
    """Fund (rows) × dimension (cols) Sum-of-IC %-of-row pivot, Status filter,
    no grand column (template pivotTable21/20 shape on a worksheet cache)."""
    n_fields = len(DEAL_LIST_COLS) + len(_CALC_FIELDS)
    ic_fld = _dl_field_index("Total Invested Capital (mlns)")
    n_rows = len(p["funds"])                      # blank fund row hidden
    n_cols = len(p["cols"])                       # blank column hidden

    # column items: display order (ascending share) → cache item indices
    cache_pos = {v: i for i, v in enumerate(p["col_cache_items"])}
    col_x = [cache_pos[c] for c in p["cols"]]

    pfs = []
    for i in range(n_fields):
        if i == p["fund_idx"]:
            items = "".join(f'<item x="{k}"/>' for k in range(n_rows))
            n_items = n_rows
            if p["fund_blank"]:
                items += f'<item h="1" x="{n_rows}"/>'
                n_items += 1
            items += '<item t="default"/>'
            pfs.append(f'<pivotField axis="axisRow" showAll="0">'
                       f'<items count="{n_items + 1}">{items}</items></pivotField>')
        elif i == p["col_idx"]:
            items = "".join(f'<item x="{k}"/>' for k in col_x)
            n_items = len(col_x)
            if p["col_blank"]:
                items += f'<item h="1" x="{len(p["col_cache_items"])}"/>'
                n_items += 1
            items += '<item t="default"/>'
            pfs.append(f'<pivotField axis="axisCol" showAll="0">'
                       f'<items count="{n_items + 1}">{items}</items></pivotField>')
        elif i == p["status_idx"]:
            items = "".join(f'<item x="{k}"/>' for k in range(p["status_cnt"])) + '<item t="default"/>'
            pfs.append(f'<pivotField axis="axisPage" showAll="0">'
                       f'<items count="{p["status_cnt"] + 1}">{items}</items></pivotField>')
        elif i == ic_fld:
            pfs.append('<pivotField dataField="1" showAll="0"/>')
        else:
            pfs.append('<pivotField showAll="0"/>')

    top = p["top"]
    last_col = get_column_letter(3 + n_cols)
    row_items = ("".join('<i><x v="%d"/></i>' % k for k in range(n_rows))
                 + '<i t="grand"><x/></i>').replace('<x v="0"/>', "<x/>", 1)
    col_items = ("".join('<i><x v="%d"/></i>' % k for k in range(n_cols))
                 ).replace('<x v="0"/>', "<x/>", 1)
    xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        + _pc_root(p["name"], ' colGrandTotals="0"')
        + f'<location ref="C{top}:{last_col}{top + n_rows + 2}" '
          'firstHeaderRow="1" firstDataRow="2" firstDataCol="1" '
          'rowPageCount="1" colPageCount="1"/>'
        + f'<pivotFields count="{n_fields}">{"".join(pfs)}</pivotFields>'
        + f'<rowFields count="1"><field x="{p["fund_idx"]}"/></rowFields>'
        + f'<rowItems count="{n_rows + 1}">{row_items}</rowItems>'
        + f'<colFields count="1"><field x="{p["col_idx"]}"/></colFields>'
        + f'<colItems count="{n_cols}">{col_items}</colItems>'
        + f'<pageFields count="1"><pageField fld="{p["status_idx"]}" hier="-1"/></pageFields>'
        + '<dataFields count="1"><dataField name="Sum of Total Invested Capital (mlns)" '
          f'fld="{ic_fld}" showDataAs="percentOfRow" baseField="0" baseItem="0" numFmtId="192"/>'
          '</dataFields>'
        + _PC_STYLE + '</pivotTableDefinition>')
    return xml.encode("utf8")


def _build_pc_count_xml(p: dict) -> bytes:
    """Single-dimension Count-of-Company pivot (template pivotTable27 shape;
    items shipped in descending-count order, blank last — manual order, so a
    refresh keeps the blank out of the pie ranges)."""
    n_fields = len(DEAL_LIST_COLS) + len(_CALC_FIELDS)
    n_rows = len(p["items"])                      # blank row hidden
    cache_pos = {v: i for i, v in enumerate(p["cache_items"])}
    order_x = [cache_pos[v] for v in p["items"]]

    pfs = []
    for i in range(n_fields):
        if i == p["field_idx"]:
            items = "".join(f'<item x="{k}"/>' for k in order_x)
            n_items = len(order_x)
            if p["has_blank"]:
                items += f'<item h="1" x="{len(p["cache_items"])}"/>'
                n_items += 1
            items += '<item t="default"/>'
            pfs.append(f'<pivotField axis="axisRow" showAll="0">'
                       f'<items count="{n_items + 1}">{items}</items></pivotField>')
        elif i == 0:
            pfs.append('<pivotField dataField="1" showAll="0"/>')
        else:
            pfs.append('<pivotField showAll="0"/>')

    top = p["top"]
    row_items = ("".join('<i><x v="%d"/></i>' % k for k in range(n_rows))
                 + '<i t="grand"><x/></i>').replace('<x v="0"/>', "<x/>", 1)
    xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        + _pc_root(p["name"])
        + f'<location ref="C{top}:D{top + n_rows + 1}" '
          'firstHeaderRow="1" firstDataRow="1" firstDataCol="1"/>'
        + f'<pivotFields count="{n_fields}">{"".join(pfs)}</pivotFields>'
        + f'<rowFields count="1"><field x="{p["field_idx"]}"/></rowFields>'
        + f'<rowItems count="{n_rows + 1}">{row_items}</rowItems>'
        + '<dataFields count="1"><dataField name="Count of Company" fld="0" '
          'subtotal="count" baseField="0" baseItem="0"/></dataFields>'
        + _PC_STYLE + '</pivotTableDefinition>')
    return xml.encode("utf8")


def _build_pc_total_xml(p: dict) -> bytes:
    """Bare grand-count pivot (template pivotTable23 shape)."""
    n_fields = len(DEAL_LIST_COLS) + len(_CALC_FIELDS)
    pfs = ['<pivotField dataField="1" showAll="0"/>'] + \
          ['<pivotField showAll="0"/>'] * (n_fields - 1)
    top = p["top"]
    xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        + _pc_root(p["name"])
        + f'<location ref="C{top}:C{top + 1}" '
          'firstHeaderRow="1" firstDataRow="1" firstDataCol="0"/>'
        + f'<pivotFields count="{n_fields}">{"".join(pfs)}</pivotFields>'
        + '<rowItems count="1"><i/></rowItems><colItems count="1"><i/></colItems>'
        + '<dataFields count="1"><dataField name="Count of Company" fld="0" '
          'subtotal="count" baseField="0" baseItem="0"/></dataFields>'
        + _PC_STYLE + '</pivotTableDefinition>')
    return xml.encode("utf8")


# ═══════════════════════════════════════════════════════════════════════════════
# Extra analysis tabs (template sheets 5-10): Vintage Perf by Sector,
# Underperforming Assets, Partner Attribution, Op Performance,
# Op Performance - Unrealized, Deployment & Exits
# ═══════════════════════════════════════════════════════════════════════════════
# The Op Performance sections aggregate the IC-weighted metrics whose Deal
# List columns were removed per EWL — the transformer still computes them per
# record, so those sections ship as COMPUTED tables (template math: ratio of
# weighted sums), not pivots. Everything else is real pivots on the shared
# cache, in our refresh-stable layout.

_EX_MONEY_FMT = "#,##0"
_OP_X_FMT = "0.0\\x"
# IC-weighted sections, template calc-field semantics (caption, hdr1, hdr2,
# kind, cell format). The OPU tab renames section 1's caption.
_OP_SECTIONS = [
    ("IC Wghted Rev/EBITDA CAGRs", "Sum of CalcICWeightedRevenueCAGR",
     "Sum of CalcWeightedEBITDACAGR", "cagr", "0%"),
    ("IC Weighted EBITDA Margins", "Entry Margin", "Exit Margin",
     "margin", "0%"),
    ("IC Weighted Multiples", "Sum of CalcWghtdEntryMultiple",
     "Sum of CalcWghtExitMultiple", "mult", _OP_X_FMT),
    ("IC Weighted Leverage", "Entry", "Exit", "lev", _OP_X_FMT),
]


def _op_cagr(entry, exit_, hp):
    """Template CAGR: (exit/entry)^(1/hp)-1, IFERROR → None."""
    if entry is None or exit_ is None or hp is None or not hp or not entry:
        return None
    try:
        r = exit_ / entry
        if r <= 0:
            return None
        return r ** (1.0 / hp) - 1.0
    except Exception:
        return None


def _op_icb(rec):
    return _cell_num(_rec_value(rec, "Total IC mlns for Buckets", 0)) or 0.0


def _sc_rev_cagr(rec):
    return _op_cagr(_cell_num(rec.get(36)), _cell_num(rec.get(46)), _cell_num(rec.get(9)))


def _sc_eb_cagr(rec):
    return _op_cagr(_cell_num(rec.get(37)), _cell_num(rec.get(47)), _cell_num(rec.get(9)))


def _sc_prod(metric):
    def f(rec):
        v = metric(rec)
        return None if v is None else _op_icb(rec) * v
    return f


def _sc_adj_mult(rec):
    a, b = _cell_num(rec.get(43)), _cell_num(rec.get(53))
    return 0.0 if (a == 0 or b == 0) else _op_icb(rec)


def _sc_adj_lev(rec):
    a, b = _cell_num(rec.get(40)), _cell_num(rec.get(50))
    return _op_icb(rec) if (a is not None and b is not None) else 0.0


def _sc_wprod(adj, key):
    def f(rec):
        v = _cell_num(rec.get(key))
        return None if v is None else adj(rec) * v
    return f


_SHEET_CALC.update({
    "AdjInvCapRevenueCAGR": _op_icb,
    "Revenue CAGR": _sc_rev_cagr,
    "WghtdRevCAGR": _sc_prod(_sc_rev_cagr),
    "AdjInvCapEBITDAcagr": _op_icb,
    "EBITDA CAGR": _sc_eb_cagr,
    "WghtdEBITDAcagr": _sc_prod(_sc_eb_cagr),
    "AdjInvCapEBITDAMargin": _op_icb,
    "Wgtd Entry EBITDA Margin": _sc_wprod(_op_icb, 38),
    "Wgtd Current EBITDA Margin": _sc_wprod(_op_icb, 48),
    "Wghtd Hold Period": _sc_wprod(_op_icb, 9),
    "AdjInvCapMultiples": _sc_adj_mult,
    "WgtdEntryMultiple": _sc_wprod(_sc_adj_mult, 43),
    "WghtdExitMultiple": _sc_wprod(_sc_adj_mult, 53),
    "AdjInvCapLeverage": _sc_adj_lev,
    "WghtdEntryLeverage": _sc_wprod(_sc_adj_lev, 40),
    "WghtdExitLeverage": _sc_wprod(_sc_adj_lev, 50),
})


def _op_pair(subset, kind):
    """(entry_ratio, exit_ratio) with the template's calc-field math:
    Σ(weight×metric where metric computes) / Σ(weight over the scope)."""
    num1 = num2 = den = 0.0
    for rec in subset:
        icb = _cell_num(_rec_value(rec, "Total IC mlns for Buckets", 0)) or 0.0
        if kind == "cagr":
            a = _op_cagr(_cell_num(rec.get(36)), _cell_num(rec.get(46)),
                         _cell_num(rec.get(9)))
            b = _op_cagr(_cell_num(rec.get(37)), _cell_num(rec.get(47)),
                         _cell_num(rec.get(9)))
            w = icb                               # unconditional denominator
        elif kind == "margin":
            a, b = _cell_num(rec.get(38)), _cell_num(rec.get(48))
            w = icb
        elif kind == "mult":
            a, b = _cell_num(rec.get(43)), _cell_num(rec.get(53))
            w = 0.0 if (a == 0 or b == 0) else icb   # template: zero-test only
        else:                                     # leverage
            a, b = _cell_num(rec.get(40)), _cell_num(rec.get(50))
            w = icb if (a is not None and b is not None) else 0.0
        den += w
        if a is not None:
            num1 += w * a
        if b is not None:
            num2 += w * b
    if not den:
        return None, None
    return num1 / den, num2 / den


def _op_hp(subset):
    """CalcWghtedHoldPeriod: Σ(icb×hp) / Σ(Total Invested Capital)."""
    num = den = 0.0
    for rec in subset:
        icb = _cell_num(_rec_value(rec, "Total IC mlns for Buckets", 0)) or 0.0
        hp = _cell_num(rec.get(9))
        if hp is not None:
            num += icb * hp
        den += _cell_num(rec.get(16)) or 0.0
    return (num / den) if den else None


def _ex_group(records, pred):
    """(count, Σic, pooled MOIC, pooled loss ratio, Σtv) over matching records."""
    cnt = 0; ic = tv = lo = 0.0
    for rec in records:
        if not pred(rec):
            continue
        cnt += 1
        ic += _cell_num(rec.get(16)) or 0.0
        tv += _cell_num(_sheet_tv(rec)) or 0.0
        lo += _cell_num(_rec_value(rec, "InvCapital in Loss Position", 0)) or 0.0
    return cnt, ic, ((tv / ic) if ic else None), ((lo / ic) if ic else None), tv


def _gn_match(rec, header, key, item):
    rv = _rec_value(rec, header, key)
    n = _cell_num(rv)
    return (int(n) if n is not None else _cell_str(rv)) == item


def _df_xml(name, fld, numfmt, subtotal=None, show_as=None):
    extra = ""
    if subtotal:
        extra += f' subtotal="{subtotal}"'
    if show_as:
        extra += f' showDataAs="{show_as}"'
    return (f'<dataField name="{_esc(name)}" fld="{fld}"{extra} '
            f'baseField="0" baseItem="0" numFmtId="{numfmt}"/>')


def _page_field_xml(p):
    fld, sel = p[0], (p[3] if len(p) > 3 else None)
    item = (' item="%d"' % sel) if sel is not None else ""
    return f'<pageField fld="{fld}"{item} hier="-1"/>'


def _bx_pivot_fields(n_fields, axis_specs, page_specs, df_flds):
    """axis_specs: {idx: (axis, items_xml, count)}; page_specs: {idx: (count, hidden_set)}"""
    pfs = []
    for i in range(n_fields):
        if i in axis_specs:
            axis, items, cnt = axis_specs[i]
            pfs.append(f'<pivotField axis="{axis}" showAll="0">'
                       f'<items count="{cnt + 1}">{items}<item t="default"/></items></pivotField>')
        elif i in page_specs:
            cnt, hidden = page_specs[i]
            items = "".join(('<item h="1" x="%d"/>' % k) if k in hidden
                            else ('<item x="%d"/>' % k) for k in range(cnt))
            df_attr = ' dataField="1"' if i in df_flds else ""
            pfs.append(f'<pivotField axis="axisPage"{df_attr} showAll="0">'
                       f'<items count="{cnt + 1}">{items}<item t="default"/></items></pivotField>')
        elif i in df_flds:
            pfs.append('<pivotField dataField="1" showAll="0"/>')
        else:
            pfs.append('<pivotField showAll="0"/>')
    return "".join(pfs)


def _bx_axis_pivot_xml(name, top, left, axis_idx, n_axis, hidden_blank,
                       pages, dfs, df_flds, order_x=None, row_x=None,
                       extra_root="") -> bytes:
    """Single row axis + σ-values columns (our refresh-stable shape).
    pages: (fld, item_count, hidden_set[, selected_idx[, label]]) tuples;
    order_x = display order of cache item indices; row_x = the visible row
    subset (pivotField item positions), defaults to all."""
    n_fields = len(DEAL_LIST_COLS) + len(_CALC_FIELDS)
    items = "".join(f'<item x="{k}"/>'
                    for k in (order_x if order_x is not None else range(n_axis)))
    n_items = n_axis
    if hidden_blank is not None:
        items += f'<item h="1" x="{hidden_blank}"/>'
        n_items += 1
    axis_specs = {axis_idx: ("axisRow", items, n_items)}
    page_specs = {p[0]: (p[1], p[2]) for p in pages}
    n_df = len(dfs)
    lcol = get_column_letter(left + n_df)
    rows = list(row_x) if row_x is not None else list(range(n_axis))
    if not rows:
        return None                               # empty axis ⇒ invalid pivot
    n_rows = len(rows)
    row_items = "".join('<i><x v="%d"/></i>' % k for k in rows) + '<i t="grand"><x/></i>'
    ci = "<i><x/></i>" + "".join(f'<i i="{k}"><x v="{k}"/></i>' for k in range(1, n_df))
    sigma = n_df >= 2                            # σ-Values col axis only for 2+
    fdr = 2 if sigma else 1
    col_block = ('<colFields count="1"><field x="-2"/></colFields>'
                 f'<colItems count="{n_df}">{ci}</colItems>') if sigma else ""
    page_attr = f' rowPageCount="{len(pages)}" colPageCount="1"' if pages else ""
    xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        + _pc_root(name, extra_root)
        + f'<location ref="{get_column_letter(left)}{top}:{lcol}{top + n_rows + fdr}"  '
          f'firstHeaderRow="1" firstDataRow="{fdr}" firstDataCol="1"{page_attr}/>'
        + f'<pivotFields count="{n_fields}">'
        + _bx_pivot_fields(n_fields, axis_specs, page_specs, df_flds)
        + '</pivotFields>'
        + f'<rowFields count="1"><field x="{axis_idx}"/></rowFields>'
        + f'<rowItems count="{n_rows + 1}">{row_items}</rowItems>'
        + col_block
        + (('<pageFields count="%d">' % len(pages))
           + "".join(_page_field_xml(p) for p in pages)
           + '</pageFields>' if pages else "")
        + f'<dataFields count="{n_df}">{"".join(dfs)}</dataFields>'
        + _PC_STYLE + '</pivotTableDefinition>')
    return xml.encode("utf8")


def _bx_matrix_xml(name, top, left, row_idx, n_rows, row_hidden,
                   col_idx, col_x, col_hidden, df, df_flds,
                   pages=(), grand_col=True, row_x=None, n_cols_all=None,
                   extra_root="") -> bytes:
    """Row axis × column axis, one dataField. row_x = rendered row subset
    (pivotField keeps every item; no-data items auto-hide on refresh via
    showAll="0"). col_x = rendered column subset out of n_cols_all items.
    row_hidden / col_hidden: int or list of item positions flagged h="1"."""
    n_fields = len(DEAL_LIST_COLS) + len(_CALC_FIELDS)

    def _hid(h):
        return [] if h is None else ([h] if isinstance(h, int) else list(h))

    hid_r, hid_c = _hid(row_hidden), _hid(col_hidden)
    n_ca = n_cols_all if n_cols_all is not None else len(col_x)

    def _axis_items(n, hid):
        ks = sorted(set(range(n)) | set(hid))
        return ("".join(('<item h="1" x="%d"/>' if k in hid else '<item x="%d"/>') % k
                        for k in ks), len(ks))

    r_items, nr = _axis_items(n_rows, hid_r)
    c_items, nc = _axis_items(n_ca, hid_c)
    axis_specs = {row_idx: ("axisRow", r_items, nr),
                  col_idx: ("axisCol", c_items, nc)}
    page_specs = {p[0]: (p[1], p[2]) for p in pages}
    rows = list(row_x) if row_x is not None else list(range(n_rows))
    if not rows or not col_x:
        return None                               # empty axis ⇒ invalid pivot
    n_cols = len(col_x)
    last = get_column_letter(left + n_cols + (1 if grand_col else 0))
    row_items = "".join('<i><x v="%d"/></i>' % k for k in rows) + '<i t="grand"><x/></i>'
    col_items = "".join('<i><x v="%d"/></i>' % k for k in col_x)
    if grand_col:
        col_items += '<i t="grand"><x/></i>'
    page_attr = f' rowPageCount="{len(pages)}" colPageCount="1"' if pages else ""
    xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        + _pc_root(name, extra_root + ("" if grand_col else ' colGrandTotals="0"'))
        + f'<location ref="{get_column_letter(left)}{top}:{last}{top + len(rows) + 2}" '
          f'firstHeaderRow="1" firstDataRow="2" firstDataCol="1"{page_attr}/>'
        + f'<pivotFields count="{n_fields}">'
        + _bx_pivot_fields(n_fields, axis_specs, page_specs, df_flds)
        + '</pivotFields>'
        + f'<rowFields count="1"><field x="{row_idx}"/></rowFields>'
        + f'<rowItems count="{len(rows) + 1}">{row_items}</rowItems>'
        + f'<colFields count="1"><field x="{col_idx}"/></colFields>'
        + f'<colItems count="{n_cols + (1 if grand_col else 0)}">{col_items}</colItems>'
        + (('<pageFields count="%d">' % len(pages))
           + "".join(_page_field_xml(p) for p in pages)
           + '</pageFields>' if pages else "")
        + f'<dataFields count="1">{df}</dataFields>'
        + _PC_STYLE + '</pivotTableDefinition>')
    return xml.encode("utf8")


def _bx_two_level_xml(name, top, left, f1_idx, f2_idx, row_tuples,
                      n1, n2, pages, dfs, df_flds) -> bytes:
    """Two row fields (parent > child); row_tuples = [(p_x, [c_x, ...]), …]."""
    n_fields = len(DEAL_LIST_COLS) + len(_CALC_FIELDS)
    i1 = "".join(f'<item x="{k}"/>' for k in range(n1))
    i2 = "".join(f'<item x="{k}"/>' for k in range(n2))
    axis_specs = {f1_idx: ("axisRow", i1, n1), f2_idx: ("axisRow", i2, n2)}
    page_specs = {p[0]: (p[1], p[2]) for p in pages}
    n_df = len(dfs)
    lcol = get_column_letter(left + n_df)
    ri, n_body = [], 0
    for p_x, kids in row_tuples:
        ri.append(f'<i><x v="{p_x}"/></i>'); n_body += 1
        for c_x in kids:
            ri.append(f'<i r="1"><x v="{c_x}"/></i>'); n_body += 1
    ri.append('<i t="grand"><x/></i>')
    ci = "<i><x/></i>" + "".join(f'<i i="{k}"><x v="{k}"/></i>' for k in range(1, n_df))
    xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        + _pc_root(name)
        + f'<location ref="{get_column_letter(left)}{top}:{lcol}{top + n_body + 2}" '
          f'firstHeaderRow="1" firstDataRow="2" firstDataCol="1" '
          f'rowPageCount="{len(pages)}" colPageCount="1"/>'
        + f'<pivotFields count="{n_fields}">'
        + _bx_pivot_fields(n_fields, axis_specs, page_specs, df_flds)
        + '</pivotFields>'
        + f'<rowFields count="2"><field x="{f1_idx}"/><field x="{f2_idx}"/></rowFields>'
        + f'<rowItems count="{n_body + 1}">{"".join(ri)}</rowItems>'
        + '<colFields count="1"><field x="-2"/></colFields>'
        + f'<colItems count="{n_df}">{ci}</colItems>'
        + ('<pageFields count="%d">' % len(pages))
        + "".join(_page_field_xml(p) for p in pages)
        + '</pageFields>'
        + f'<dataFields count="{n_df}">{"".join(dfs)}</dataFields>'
        + _PC_STYLE + '</pivotTableDefinition>')
    return xml.encode("utf8")


def _fld_meta(header):
    idx = _dl_field_index(header)
    _h, _s, key, kind = DEAL_LIST_COLS[idx]
    return idx, key, kind


def plan_extra(records: list[dict]) -> dict:
    """Data + geometry for the six extra tabs."""
    ex: dict = {}
    n_base = len(DEAL_LIST_COLS)
    calc_moic, calc_loss = n_base, n_base + 1

    def page_spec(header, selected=None):
        idx, key, kind = _fld_meta(header)
        items = _grouped_items(records, header, key, kind)
        blank = (kind != "gn") and any(
            _cell_str(_rec_value(r, header, key)) is None for r in records)
        sel = items.index(selected) if (selected in items) else None
        return (idx, len(items) + (1 if blank else 0), set(), sel,
                str(selected) if sel is not None else "(All)")

    v_idx, v_key, _vk = _fld_meta("Vintage")
    vintages = _grouped_items(records, "Vintage", v_key, "gn")
    f_idx, f_key, _fk = _fld_meta("Fund")
    funds = _grouped_items(records, "Fund", f_key, "g")
    fund_blank = any(_cell_str(r.get(f_key)) is None for r in records)
    s_idx, s_key, _sk = _fld_meta("Sector")
    sectors = _grouped_items(records, "Sector", s_key, "g")
    sector_blank = any(_cell_str(r.get(s_key)) is None for r in records)

    def in_v(rec, v): return _gn_match(rec, "Vintage", v_key, v)
    def in_f(rec, f): return _cell_str(rec.get(f_key)) == f
    def in_s(rec, s): return _cell_str(rec.get(s_key)) == s

    tot = _ex_group(records, lambda r: True)

    # ── Vintage Perf by Sector ────────────────────────────────────────────
    vs = {"sections": []}
    anchor = 8 + 4 + 3
    rows = []
    for v in vintages:
        cnt, ic, moic, loss, _tv = _ex_group(records, lambda r, v=v: in_v(r, v))
        rows.append((v, cnt, ic, moic, loss))
    vs["sections"].append({
        "type": "vpivot", "title": "Invested Capital & MOIC by Vintage",
        "anchor": anchor, "top": anchor + 8, "rows": rows,
        "total": (tot[0], tot[1], tot[2], tot[3]),
        "axis_idx": v_idx, "n_axis": len(vintages),
        "pages": [page_spec("Fund"), page_spec("Sector"),
                  page_spec("Status"), page_spec("Hold Period Buckets")],
    })
    anchor = vs["sections"][0]["top"] + len(vintages) + 3 + 4

    def matrix_section(title, value, pages):
        nonlocal anchor
        cells, rowt, colt = {}, {}, {}
        for v in vintages:
            for s in sectors:
                cnt, _ic, moic, _lo, _tv = _ex_group(
                    records, lambda r, v=v, s=s: in_v(r, v) and in_s(r, s))
                cells[(v, s)] = cnt if value == "count" else moic
        for v in vintages:
            cnt, _ic, moic, _lo, _tv = _ex_group(records, lambda r, v=v: in_v(r, v))
            rowt[v] = cnt if value == "count" else moic
        for s in sectors:
            cnt, _ic, moic, _lo, _tv = _ex_group(records, lambda r, s=s: in_s(r, s))
            colt[s] = cnt if value == "count" else moic
        sec = {"type": "vmatrix", "title": title, "value": value,
               "anchor": anchor, "top": anchor + (7 if pages else 2),
               "cells": cells, "rowt": rowt, "colt": colt,
               "grand": tot[0] if value == "count" else tot[2],
               "pages": pages}
        vs["sections"].append(sec)
        anchor = sec["top"] + len(vintages) + 3 + 3

    matrix_section("Deal Count by Vintage and Sector", "count", [])
    matrix_section("Pooled MOIC by Vintage and Sector", "moic", [])
    matrix_section("Deal Count by Vintage and Sector (filter by Fund)", "count",
                   [page_spec("Fund")])
    vs["_vintages"], vs["_sectors"] = vintages, sectors
    ex["Vintage Perf by Sector"] = vs
    ex["_dims"] = {"vintages": vintages, "sectors": sectors, "funds": funds,
                   "sector_blank": sector_blank, "fund_blank": fund_blank,
                   "v_idx": v_idx, "s_idx": s_idx, "f_idx": f_idx,
                   "calc_moic": calc_moic, "calc_loss": calc_loss}

    # ── Underperforming Assets ────────────────────────────────────────────  [EWL: tabs removed — kept for later]  [EWL: tabs removed — kept for later]
    # p_idx, p_key, _pk = _fld_meta("Performing\n(1=Underperform)")
    # perf_items = _grouped_items(records, "Performing\n(1=Underperform)", p_key, "gn")
    # c_idx, c_key, _ck = _fld_meta("Company")
    # companies = _grouped_items(records, "Company", c_key, "g")
    # c_pos = {v: i for i, v in enumerate(companies)}
    # f_pos = {v: i for i, v in enumerate(funds)}
    # loss_recs = [r for r in records
    #              if _rec_value(r, "Performing\n(1=Underperform)", 0) == 1]
    # ua_rows = []
    # for f in funds:
    #     frecs = [r for r in loss_recs if in_f(r, f)]
    #     if not frecs:
    #         continue
    #     kids = []
    #     for r in sorted(frecs, key=lambda r: _cell_str(r.get(1)) or ""):
    #         comp = _cell_str(r.get(1))
    #         ic = _cell_num(r.get(16)) or 0.0
    #         imp = _cell_num(_rec_value(r, "Impaired\nValue", 0)) or 0.0
    #         kids.append((comp, (1, _sheet_moic(r), _cell_num(r.get(18)) or 0.0,
    #                             (imp / ic) if ic else None, _cell_num(r.get(9)))))
    #     _cnt, fic, fmoic, _lo, _tv = _ex_group(loss_recs, lambda r, f=f: in_f(r, f))
    #     cur = sum(_cell_num(r.get(18)) or 0.0 for r in frecs)
    #     imp = sum(_cell_num(_rec_value(r, "Impaired\nValue", 0)) or 0.0 for r in frecs)
    #     hps = [_cell_num(r.get(9)) for r in frecs if _cell_num(r.get(9)) is not None]
    #     ua_rows.append((f, (len(frecs), fmoic, cur, (imp / fic) if fic else None,
    #                         (sum(hps) / len(hps)) if hps else None), kids))
    # tic = sum(_cell_num(r.get(16)) or 0.0 for r in loss_recs)
    # ttv = sum(_cell_num(_sheet_tv(r)) or 0.0 for r in loss_recs)
    # tcur = sum(_cell_num(r.get(18)) or 0.0 for r in loss_recs)
    # timp = sum(_cell_num(_rec_value(r, "Impaired\nValue", 0)) or 0.0 for r in loss_recs)
    # thps = [_cell_num(r.get(9)) for r in loss_recs if _cell_num(r.get(9)) is not None]
    # ex["Underperforming Assets"] = {
    #     "top": 15, "rows": ua_rows,
    #     "total": (len(loss_recs), (ttv / tic) if tic else None, tcur,
    #               (timp / tic) if tic else None,
    #               (sum(thps) / len(thps)) if thps else None),
    #     "p_idx": p_idx,
    #     "n_perf": len(perf_items),
    #     "sel_perf": perf_items.index(1) if 1 in perf_items else None,
    #     "f_idx": f_idx, "c_idx": c_idx,
    #     "n_funds": len(funds) + (1 if fund_blank else 0),
    #     "n_comp": len(companies) + (1 if any(_cell_str(r.get(1)) is None
    #                                          for r in records) else 0),
    #     "tuples": [(f_pos[f], [c_pos[c] for c, _v in kids if c in c_pos])
    #                for f, _agg, kids in ua_rows],
    # }

    # ── Partner Attribution ──────────────────────────────────────────────
    # pa_idx, pa_key, _pak = _fld_meta("Sourcing Partner")
    # partners_cache = _grouped_items(records, "Sourcing Partner", pa_key, "g")
    # partner_blank = any(_cell_str(r.get(pa_key)) is None for r in records)
    # partners = sorted(partners_cache, key=lambda s: s.lower())
    # pa_pos = {v: i for i, v in enumerate(partners_cache)}
    # pa_rows = []
    # for p in partners:
    #     cnt, _ic, moic, _lo, _tv = _ex_group(
    #         records, lambda r, p=p: _cell_str(r.get(pa_key)) == p)
    #     pa_rows.append((p, cnt, moic))
    # ex["Partner Attribution"] = {
    #     "top": 11, "rows": pa_rows, "total": (tot[0], tot[2]),
    #     "axis_idx": pa_idx, "order_x": [pa_pos[p] for p in partners],
    #     "hidden_blank": len(partners_cache) if partner_blank else None,
    # }

    # ── Op Performance (+ Unrealized) ────────────────────────────────────
    # def op_tables(subset):
    #     out = []
    #     for title, h1, h2, kind, fmt in _OP_SECTIONS:
    #         rows = []
    #         for f in funds:
    #             fsub = [r for r in subset if in_f(r, f)]
    #             if not fsub:
    #                 continue                      # template omits empty funds
    #             a, b = _op_pair(fsub, kind)
    #             rows.append((f, a, b))
    #         out.append({"title": title, "h1": h1, "h2": h2, "kind": kind,
    #                     "fmt": fmt, "rows": rows,
    #                     "total": _op_pair(subset, kind)})
    #     return out

    # f_pos_all = {v: i for i, v in enumerate(funds)}
    # Grand totals must match the pivot's refresh scope: the blank-fund item
    # is hidden, so totals aggregate fund-visible records only.
    # vis = [r for r in records if any(in_f(r, f) for f in funds)]
    # op_tabs = op_tables(vis)
    # for t in op_tabs:
    #     t["row_x"] = [f_pos_all[f] for f, _a, _b in t["rows"]]
    # ex["Op Performance"] = {
    #     "tables": op_tabs, "funds": funds,
    #     "f_idx": f_idx, "n_funds": len(funds), "fund_blank": fund_blank,
    #     "pages": [page_spec("Sector"), page_spec("Status"),
    #               page_spec("Hold Period Buckets")],
    # }

    # unreal = [r for r in vis if _cell_str(r.get(5)) == "Unrealized"]
    # opu_tables = op_tables(unreal)
    # opu_tables[0]["title"] = "Growth CAGRs"
    # for t in opu_tables:
    #     t["row_x"] = [f_pos_all[f] for f, _a, _b in t["rows"]]
    # rv_rows = [(f,
    #             sum(_cell_num(r.get(17)) or 0.0 for r in records if in_f(r, f)),
    #             sum(_cell_num(r.get(18)) or 0.0 for r in records if in_f(r, f)))
    #            for f in funds]
    # ex["Op Performance - Unrealized"] = {
    #     "tables": opu_tables, "funds": funds,
    #     "rv_rows": rv_rows,
    #     "rv_total": (sum(v for _f, v, _c in rv_rows), sum(c for _f, _v, c in rv_rows)),
    #     "f_idx": f_idx, "n_funds": len(funds), "fund_blank": fund_blank,
    #     "hp_rows": [(f, _op_hp([r for r in unreal if in_f(r, f)]))
    #                 for f in funds
    #                 if any(in_f(r, f) for r in unreal)],
    #     "hp_total": _op_hp(unreal),
    #     "hp_row_x": [f_pos_all[f] for f in funds
    #                  if any(in_f(r, f) for r in unreal)],
    #     "pages_all": [page_spec("Sector"), page_spec("Status"),
    #                   page_spec("Hold Period Buckets")],
    #     "pages_unreal": [page_spec("Sector"),
    #                      page_spec("Status", selected="Unrealized"),
    #                      page_spec("Hold Period Buckets")],
    # }

    # ── Deployment & Exits ───────────────────────────────────────────────
    e_idx, e_key, _ek = _fld_meta("Exit Year")
    exit_years = _grouped_items(records, "Exit Year", e_key, "gn")

    def in_e(rec, y): return _gn_match(rec, "Exit Year", e_key, y)

    # fund-visible records (the blank-fund pivot item is hidden everywhere);
    # defined here since the Op block that used to own it is commented out
    vis = [r for r in records if any(in_f(r, f) for f in funds)]

    def grid(subset, rowvals, colvals, rpred, cpred, key):
        """cells + totals derived from the same visible grid, so written
        grand totals always equal the pivot's visible-only refresh values."""
        cells = {}
        for rv in rowvals:
            rsub = [r for r in subset if rpred(r, rv)]
            for cv in colvals:
                cells[(rv, cv)] = sum(
                    (_cell_num(r.get(key)) or 0.0) if key else 1
                    for r in rsub if cpred(r, cv))
        rowt = {rv: sum(cells[(rv, cv)] for cv in colvals) for rv in rowvals}
        colt = {cv: sum(cells[(rv, cv)] for rv in rowvals) for cv in colvals}
        return cells, rowt, colt, sum(rowt.values())

    # sections 1-2: vintage × fund over fund-visible records
    vis_vintages = [v for v in vintages if any(in_v(r, v) for r in vis)]
    ic_cells, _rt, ic_colt, _g = grid(vis, vis_vintages, funds, in_v, in_f, 16)
    cnt_cells, cnt_rowt, cnt_colt, cnt_g = grid(vis, vis_vintages, funds,
                                                in_v, in_f, None)

    # sections 3-4: fund × exit year, Status pre-filtered to Realized
    realized = [r for r in vis if _cell_str(r.get(5)) == "Realized"]
    yr_items = [y for y in exit_years if not isinstance(y, str)]
    vis_years = [y for y in yr_items
                 if any(in_e(r, y) for r in realized)]
    ex_funds = [f for f in funds
                if any(in_f(r, f) and any(in_e(r, y) for y in vis_years)
                       for r in realized)]
    exp_cells, exp_rowt, exp_colt, exp_g = grid(realized, ex_funds, vis_years,
                                                in_f, in_e, 14)
    exc_cells, _ert, exc_colt, exc_g = grid(realized, ex_funds, vis_years,
                                            in_f, in_e, None)
    ex["Deployment & Exits"] = {
        "vintages": vintages, "vis_vintages": vis_vintages, "funds": funds,
        "exit_years": exit_years, "vis_years": vis_years, "ex_funds": ex_funds,
        "v_idx": v_idx, "f_idx": f_idx, "e_idx": e_idx, "fund_blank": fund_blank,
        "v_row_x": [vintages.index(v) for v in vis_vintages],
        "f_row_x": [funds.index(f) for f in ex_funds],
        "y_col_x": [exit_years.index(y) for y in vis_years],
        "y_hidden": [i for i, y in enumerate(exit_years) if isinstance(y, str)],
        "ic_pct": {k: ((v / ic_colt[k[1]]) if ic_colt.get(k[1]) else 0.0)
                   for k, v in ic_cells.items()},
        "cnt": (cnt_cells, cnt_rowt, cnt_colt, cnt_g),
        "exits_pct": ({k: ((v / exp_rowt[k[0]]) if exp_rowt.get(k[0]) else 0.0)
                       for k, v in exp_cells.items()},
                      {y: ((exp_colt[y] / exp_g) if exp_g else 0.0)
                       for y in vis_years}),
        "exits_cnt": (exc_cells, exc_colt, exc_g),
        "status_page": page_spec("Status", selected="Realized"),
    }
    return ex


_F10 = lambda: Font(size=10)
_F10B = lambda: Font(bold=True, size=10)


def _ex_meta(ws, title):
    ws["B2"] = title; ws["B2"].font = _TITLE_FONT
    ws["B4"] = "Sponsor/GP:"; ws["C4"] = "='Deal List'!$C$4"
    ws["B5"] = "As of Date:"; ws["C5"] = "='Deal List'!$C$5"
    ws["C5"].number_format = "d-mmm-yy"
    ws["B6"] = "Currency:";   ws["C6"] = "='Deal List'!$C$6"
    for ref in ("B4", "B5", "B6", "C4", "C5", "C6"):
        ws[ref].font = _CALC_FONT


def _ex_cell(ws, r, c, v, fmt=None, bold=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = _F10B() if bold else _F10()
    if fmt:
        cell.number_format = fmt
    return cell


_PAGE_NAMES = {}


def _page_label(idx):
    return DEAL_LIST_COLS[idx][0].replace("\n", " ")


def _ex_filters(ws, pages, top, label_col=3):
    for k, p in enumerate(pages):
        fr = top - 1 - len(pages) + k
        _ex_cell(ws, fr, label_col, _page_label(p[0]))
        sel = p[3] if len(p) > 3 else None
        lbl = (p[4] if len(p) > 4 and sel is not None else
               ("(All)" if sel is None else str(sel)))
        _ex_cell(ws, fr, label_col + 1, lbl)


def _write_vs_sheet(ws, vs) -> None:
    _ex_meta(ws, "Vintage Performance & Invested Capital by Sector")
    _write_mini_toc(ws, [(s["title"], s["anchor"] - 1) for s in vs["sections"]],
                    start_row=8)
    for s in vs["sections"]:
        _ex_cell(ws, s["anchor"] - 1, 2, s["title"]).font = _SECTION_FONT
        top = s["top"]
        if s["type"] == "vpivot":
            _ex_filters(ws, s["pages"], top)
            _ex_cell(ws, top, 4, "Values")
            _ex_cell(ws, top + 1, 3, "Row Labels")
            for j, h in enumerate(("Count", "Invested Capital", "MOIC", "Loss Ratio")):
                _ex_cell(ws, top + 1, 4 + j, h)
            _ex_cell(ws, top + 1, 9, "Graph Label", bold=True)
            for i, (v, cnt, ic, moic, loss) in enumerate(s["rows"]):
                r = top + 2 + i
                _ex_cell(ws, r, 3, v)
                _ex_cell(ws, r, 4, cnt, "0")
                _ex_cell(ws, r, 5, ic, _EX_MONEY_FMT)
                _ex_cell(ws, r, 6, moic, _MOIC_FMT)
                _ex_cell(ws, r, 7, loss, "0%")
                ws.cell(row=r, column=9,
                        value=f"=CONCATENATE(C{r},CHAR(10),D{r})")
            gr = top + 2 + len(s["rows"])
            _ex_cell(ws, gr, 3, "Grand Total", bold=True)
            tc, ti, tm, tl = s["total"]
            _ex_cell(ws, gr, 4, tc, "0", True)
            _ex_cell(ws, gr, 5, ti, _EX_MONEY_FMT, True)
            _ex_cell(ws, gr, 6, tm, _MOIC_FMT, True)
            _ex_cell(ws, gr, 7, tl, "0%", True)
        else:
            if s["pages"]:
                _ex_filters(ws, s["pages"], top)
            fmt = "0" if s["value"] == "count" else _MOIC_FMT
            vintages, sectors = vs["_vintages"], vs["_sectors"]
            _ex_cell(ws, top, 3, s["title"].split(" by ")[0])
            _ex_cell(ws, top, 4, "Column Labels")
            _ex_cell(ws, top + 1, 3, "Vintage")
            for k, sec in enumerate(sectors):
                _ex_cell(ws, top + 1, 4 + k, sec)
            _ex_cell(ws, top + 1, 4 + len(sectors), "Grand Total")
            for i, v in enumerate(vintages):
                r = top + 2 + i
                _ex_cell(ws, r, 3, v)
                for k, sec in enumerate(sectors):
                    _ex_cell(ws, r, 4 + k, s["cells"].get((v, sec)), fmt)
                _ex_cell(ws, r, 4 + len(sectors), s["rowt"].get(v), fmt)
            gr = top + 2 + len(vintages)
            _ex_cell(ws, gr, 3, "Grand Total", bold=True)
            for k, sec in enumerate(sectors):
                _ex_cell(ws, gr, 4 + k, s["colt"].get(sec), fmt, True)
            _ex_cell(ws, gr, 4 + len(sectors), s["grand"], fmt, True)


def _write_ua_sheet(ws, ua) -> None:
    _ex_meta(ws, "List of Underperforming Assets")
    ws["B8"] = "Loss Ratio = Invested Capital Below 1.0x / Total Invested Capital"
    ws["B9"] = "Impaired Capital = Amount of Money Currently Lost / Total Invested Capital"
    for ref in ("B8", "B9"):
        ws[ref].font = Font(italic=True, size=9)
    top = ua["top"]
    _ex_cell(ws, top - 3, 2, "Performing (1=Underperform)")
    _ex_cell(ws, top - 3, 3, 1)
    _ex_cell(ws, top, 3, "Values")
    _ex_cell(ws, top + 1, 2, "Row Labels")
    caps = ("Loss Deals", "MOIC", "Current Value", "Impaired Capital", "Hold Period")
    for j, h in enumerate(caps):
        _ex_cell(ws, top + 1, 3 + j, h)
    fmts = ("0", _MOIC_FMT, _EX_MONEY_FMT, "0%", "0.0")
    r = top + 2
    for fund, agg, kids in ua["rows"]:
        _ex_cell(ws, r, 2, fund, bold=True)
        for j, v in enumerate(agg):
            _ex_cell(ws, r, 3 + j, v, fmts[j], True)
        r += 1
        for comp, vals in kids:
            _ex_cell(ws, r, 2, comp)
            for j, v in enumerate(vals):
                _ex_cell(ws, r, 3 + j, v, fmts[j])
            r += 1
    _ex_cell(ws, r, 2, "Grand Total", bold=True)
    for j, v in enumerate(ua["total"]):
        _ex_cell(ws, r, 3 + j, v, fmts[j], True)


def _write_pa_sheet(ws, pa) -> None:
    _ex_meta(ws, "Partner / Deal Lead Attribution")
    top = pa["top"]
    _ex_cell(ws, top, 3, "Values")
    _ex_cell(ws, top + 1, 2, "Row Labels")
    _ex_cell(ws, top + 1, 3, "Count")
    _ex_cell(ws, top + 1, 4, "MOIC")
    for i, (p, cnt, moic) in enumerate(pa["rows"]):
        r = top + 2 + i
        _ex_cell(ws, r, 2, p)
        _ex_cell(ws, r, 3, cnt, "0")
        _ex_cell(ws, r, 4, moic, _MOIC_FMT)
    gr = top + 2 + len(pa["rows"])
    _ex_cell(ws, gr, 2, "Grand Total", bold=True)
    _ex_cell(ws, gr, 3, pa["total"][0], "0", True)
    _ex_cell(ws, gr, 4, pa["total"][1], _MOIC_FMT, True)


_HDR_BLUE = lambda: Font(bold=True, size=10, color="1F4E78")


def _write_op_tables(ws, tables, first_heading, grand_label="Grand Total",
                     pages=None):
    """Pivot-look sections: caption (col C, blue), report-filter rows, Values
    caption row, blue header captions, fund rows, grand row. Stashes
    hrow/top/thead per section for the pivot XML, charts and mini-TOC."""
    h = first_heading
    for t in tables:
        t["hrow"] = h
        c = ws.cell(row=h, column=3, value=t["title"]); c.font = _HDR_BLUE()
        top = h + 6                               # filters h+2..h+4, blank h+5
        t["top"] = top
        if pages:
            _ex_filters(ws, pages, top, label_col=2)
        ws.cell(row=top, column=3, value="Values").font = _HDR_BLUE()
        r0 = top + 1
        t["thead"] = r0
        ws.cell(row=r0, column=2, value="Row Labels").font = _HDR_BLUE()
        ws.cell(row=r0, column=3, value=t["h1"]).font = _HDR_BLUE()
        ws.cell(row=r0, column=4, value=t["h2"]).font = _HDR_BLUE()
        for i, (f, a, b) in enumerate(t["rows"]):
            _ex_cell(ws, r0 + 1 + i, 2, f)
            _ex_cell(ws, r0 + 1 + i, 3, a, t["fmt"])
            _ex_cell(ws, r0 + 1 + i, 4, b, t["fmt"])
        gr = r0 + 1 + len(t["rows"])
        _ex_cell(ws, gr, 2, grand_label, bold=True)
        _ex_cell(ws, gr, 3, t["total"][0], t["fmt"], True)
        _ex_cell(ws, gr, 4, t["total"][1], t["fmt"], True)
        h = max(gr + 1, t["thead"] - 2 + 12) + 3


def _write_op_sheet(ws, op) -> None:
    _ex_meta(ws, "Operating Performance")
    _write_op_tables(ws, op["tables"], first_heading=14, pages=op["pages"])
    _write_mini_toc(ws, [(t["title"], t["hrow"]) for t in op["tables"]],
                    start_row=8)


def _write_opu_sheet(ws, opu) -> None:
    _ex_meta(ws, "Operational Performance")
    _ex_cell(ws, 15, 3, "Realized vs Unrealized Value by Fund").font = _HDR_BLUE()
    top = 21                                      # filters 17-19, blank 20
    opu["rv_top"] = top
    _ex_filters(ws, opu["pages_all"], top, label_col=2)
    acc = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'
    ws.cell(row=top, column=3, value="Values").font = _HDR_BLUE()
    ws.cell(row=top + 1, column=2, value="Row Labels").font = _HDR_BLUE()
    ws.cell(row=top + 1, column=3, value="Realized Value").font = _HDR_BLUE()
    ws.cell(row=top + 1, column=4, value="Current Value").font = _HDR_BLUE()
    ws.cell(row=top + 1, column=5,
            value="% of Value that is Unrealized").font = _HDR_BLUE()
    for i, (f, rv, cv) in enumerate(opu["rv_rows"]):
        r = top + 2 + i
        _ex_cell(ws, r, 2, f)
        _ex_cell(ws, r, 3, rv, acc)
        _ex_cell(ws, r, 4, cv, acc)
        c = ws.cell(row=r, column=5, value=f"=IF(C{r}+D{r}=0,\"\",D{r}/(C{r}+D{r}))")
        c.number_format = "0%"; c.font = _F10()
    gr = top + 2 + len(opu["rv_rows"])
    _ex_cell(ws, gr, 2, "Grand Total", bold=True)
    _ex_cell(ws, gr, 3, opu["rv_total"][0], acc, True)
    _ex_cell(ws, gr, 4, opu["rv_total"][1], acc, True)
    c = ws.cell(row=gr, column=5, value=f"=IF(C{gr}+D{gr}=0,\"\",D{gr}/(C{gr}+D{gr}))")
    c.number_format = "0%"; c.font = _F10B()

    _write_op_tables(ws, opu["tables"], first_heading=gr + 4,
                     grand_label="Consolidated", pages=opu["pages_unreal"])
    last = opu["tables"][-1]
    h = max(last["thead"] + len(last["rows"]) + 2, last["thead"] - 2 + 12) + 3
    opu["hp_hrow"] = h
    _ex_cell(ws, h, 3, "IC Weighted Hold Period").font = _HDR_BLUE()
    hp_top = h + 6
    opu["hp_top"] = hp_top
    _ex_filters(ws, opu["pages_unreal"], hp_top, label_col=2)
    r0 = hp_top
    ws.cell(row=r0, column=2, value="Row Labels").font = _HDR_BLUE()
    ws.cell(row=r0, column=3,
            value="Sum of CalcWghtedHoldPeriod").font = _HDR_BLUE()
    hp_fmt = '0.0\\ "yrs"'
    for i, (f, v) in enumerate(opu["hp_rows"]):
        _ex_cell(ws, r0 + 1 + i, 2, f)
        _ex_cell(ws, r0 + 1 + i, 3, v, hp_fmt)
    _ex_cell(ws, r0 + 1 + len(opu["hp_rows"]), 2, "Grand Total", bold=True)
    _ex_cell(ws, r0 + 1 + len(opu["hp_rows"]), 3, opu["hp_total"], hp_fmt, True)
    _write_mini_toc(
        ws,
        [("Realized vs Unrealized Value by Fund", 15)]
        + [(t["title"], t["hrow"]) for t in opu["tables"]]
        + [("IC Weighted Hold Period", opu["hp_hrow"])],
        start_row=8)


def _write_de_sheet(ws, de) -> None:
    _ex_meta(ws, "Deployment/Pacing")

    def matrix(top, left, caption, row_caption, rows, cols, cells, fmt,
               rowt=None, colt=None, grand=None, grand_col=False,
               grand_row=None):
        ws.cell(row=top, column=left, value=caption).font = _HDR_BLUE()
        ws.cell(row=top, column=left + 1, value="Column Labels").font = _HDR_BLUE()
        ws.cell(row=top + 1, column=left, value=row_caption).font = _HDR_BLUE()
        for k, cv in enumerate(cols):
            ws.cell(row=top + 1, column=left + 1 + k, value=cv).font = _HDR_BLUE()
        if grand_col:
            ws.cell(row=top + 1, column=left + 1 + len(cols),
                    value="Grand Total").font = _HDR_BLUE()
        for i, rv in enumerate(rows):
            r = top + 2 + i
            _ex_cell(ws, r, left, rv)
            for k, cv in enumerate(cols):
                _ex_cell(ws, r, left + 1 + k, cells.get((rv, cv), 0), fmt)
            if grand_col:
                _ex_cell(ws, r, left + 1 + len(cols), rowt.get(rv), fmt)
        gr = top + 2 + len(rows)
        _ex_cell(ws, gr, left, "Grand Total", bold=True)
        for k, cv in enumerate(cols):
            v = grand_row[cv] if grand_row is not None else colt.get(cv)
            _ex_cell(ws, gr, left + 1 + k, v, fmt, True)
        if grand_col:
            _ex_cell(ws, gr, left + 1 + len(cols), grand, fmt, True)

    secs = de["_layout"] = []
    # 1. InvCap % — vintage rows × fund columns, % of column (template: no
    #    grand-total column, row header captioned "Fund")
    secs.append({"title": "InvCap %", "hrow": 14, "top": 16})
    s = secs[-1]
    _ex_cell(ws, s["hrow"], 2, s["title"]).font = _SECTION_FONT
    matrix(s["top"], 3, "Sum of Total Invested Capital (mlns)", "Fund",
           de["vis_vintages"], de["funds"], de["ic_pct"], "0%",
           grand_row={f: 1 for f in de["funds"]})
    anchor = s["top"] + 2 + len(de["vis_vintages"]) + 4
    # 2. Deal Count — vintage rows × fund columns, with grand-total column
    secs.append({"title": "Deal Count", "hrow": anchor, "top": anchor + 2})
    s = secs[-1]
    _ex_cell(ws, s["hrow"], 2, s["title"]).font = _SECTION_FONT
    cnt_cells, cnt_rowt, cnt_colt, cnt_g = de["cnt"]
    matrix(s["top"], 3, "Count of Company", "Fund",
           de["vis_vintages"], de["funds"], cnt_cells, "0",
           rowt=cnt_rowt, colt=cnt_colt, grand=cnt_g, grand_col=True)
    anchor = s["top"] + 2 + len(de["vis_vintages"]) + 4
    # 3. Exits % of IC by Fund — realized IC share by exit year (% of row)
    secs.append({"title": "Exits % of IC by Fund", "hrow": anchor,
                 "top": anchor + 4})
    s = secs[-1]
    _ex_cell(ws, s["hrow"], 2, s["title"]).font = _SECTION_FONT
    _ex_filters(ws, [de["status_page"]], s["top"], label_col=4)
    pct_cells, pct_grand = de["exits_pct"]
    matrix(s["top"], 4, "Sum of Total IC mlns for Buckets", "Row Labels",
           de["ex_funds"], de["vis_years"], pct_cells, "0%",
           grand_row=pct_grand)
    anchor = s["top"] + 2 + len(de["ex_funds"]) + 4
    # 4. Exits by Year — realized deal count by exit year (no grand column)
    secs.append({"title": "Exits by Year", "hrow": anchor, "top": anchor + 4})
    s = secs[-1]
    _ex_cell(ws, s["hrow"], 2, s["title"]).font = _SECTION_FONT
    _ex_filters(ws, [de["status_page"]], s["top"], label_col=4)
    exc_cells, exc_colt, _exc_g = de["exits_cnt"]
    matrix(s["top"], 4, "Count of Company", "Row Labels",
           de["ex_funds"], de["vis_years"], exc_cells, "0", colt=exc_colt)
    _write_mini_toc(ws, [(x["title"], x["hrow"]) for x in secs], start_row=8)


def _extra_jobs(ex: dict) -> list:
    """[(sheet_name, [pivot_xml, …], [(chart_xml, c0, r0, w, h), …]), …]
    consumed by the injector after the sheets are written (writers stash the
    final layout rows)."""
    import pathlib
    here = pathlib.Path(__file__).parent
    vint_tpl = (here / "chart_vintage.xml").read_text()
    op_tpl = (here / "chart_op.xml").read_text()
    jobs = []
    seq = [0]

    def nm():
        seq[0] += 1
        return f"EX{seq[0]}"

    dims = ex["_dims"]
    ic_idx = _dl_field_index("Total Invested Capital (mlns)")
    cur_idx = _dl_field_index("Current\nValue")
    real_idx = _dl_field_index("Realized\nValue")
    hp_idx = _dl_field_index("Hold\nPeriod")
    icb_idx = _dl_field_index("Total IC mlns for Buckets")
    cm, cl = dims["calc_moic"], dims["calc_loss"]
    ci = cl + 1                                   # CalcImpairedLossRatio

    # ── Vintage Perf by Sector ───────────────────────────────────────────
    vs = ex["Vintage Perf by Sector"]
    pivots, charts = [], []
    n_sec = len(dims["sectors"])
    sec_hidden = n_sec if dims["sector_blank"] else None
    for s in vs["sections"]:
        if s["type"] == "vpivot":
            dfs = [_df_xml("Count", 0, 1, subtotal="count"),
                   _df_xml("Invested Capital", ic_idx, 3),
                   _df_xml("MOIC", cm, 217),
                   _df_xml("Loss Ratio", cl, 9)]
            pivots.append(_bx_axis_pivot_xml(
                nm(), s["top"], 3, s["axis_idx"], s["n_axis"], None,
                s["pages"], dfs, {0, ic_idx, cm, cl}))
            cats = [f"{v}\n{cnt}" for v, cnt, _ic, _m, _l in s["rows"]]
            cx = (vint_tpl
                  .replace("{CATS}", _str_lit(cats))
                  .replace("{VBAR}", _num_lit([ic for _v, _c, ic, _m, _l in s["rows"]],
                                              _EX_MONEY_FMT))
                  .replace("{VLINE}", _num_lit([m for _v, _c, _ic, m, _l in s["rows"]],
                                               "0.0\\x;\\(0.0\\x\\)"))
                  .replace("{TITLE}", _esc(s["title"])))
            charts.append((cx, 10, max(s["top"] - 2, 0), 11, 14))
        else:
            df = (_df_xml("Count of Company", 0, 1, subtotal="count")
                  if s["value"] == "count" else _df_xml("MOIC", cm, 217))
            flds = {0} if s["value"] == "count" else {cm}
            pivots.append(_bx_matrix_xml(
                nm(), s["top"], 3, dims["v_idx"], len(vs["_vintages"]), None,
                dims["s_idx"], list(range(n_sec)), sec_hidden,
                df, flds, pages=s["pages"], grand_col=True))
    jobs.append(("Vintage Perf by Sector", pivots, charts))

    # ── Underperforming Assets ───────────────────────────────────────────  [EWL: tabs removed — kept for later]
    # ua = ex["Underperforming Assets"]
    # dfs = [_df_xml("Loss Deals", ua["p_idx"], 1),
    #        _df_xml("MOIC", cm, 217),
    #        _df_xml("Current Value", cur_idx, 3),
    #        _df_xml("Impaired Capital", ci, 9),
    #        _df_xml("Hold Period", hp_idx, 2, subtotal="average")]
    # jobs.append(("Underperforming Assets",
    #              [_bx_two_level_xml(nm(), ua["top"], 2, ua["f_idx"], ua["c_idx"],
    #                                 ua["tuples"], ua["n_funds"], ua["n_comp"],
    #                                 [(ua["p_idx"], ua["n_perf"], set(), ua["sel_perf"])],
    #                                 dfs, {ua["p_idx"], cm, cur_idx, ci, hp_idx})],
    #              []))

    # ── Partner Attribution ──────────────────────────────────────────────
    # pa = ex["Partner Attribution"]
    # dfs = [_df_xml("Count", 0, 1, subtotal="count"), _df_xml("MOIC", cm, 217)]
    # jobs.append(("Partner Attribution",
    #              [_bx_axis_pivot_xml(nm(), pa["top"], 2, pa["axis_idx"],
    #                                  len(pa["rows"]), pa["hidden_blank"],
    #                                  [], dfs, {0, cm}, order_x=pa["order_x"])],
    #              []))

    # ── Op Performance (+ Unrealized): filtered calc-field pivots + charts ─
    # def op_charts(tables, sheet):
    #     out = []
    #     for t in tables:
    #         cats = [f for f, _a, _b in t["rows"]]
    #         cx = (op_tpl
    #               .replace("{S1}", _esc(t["h1"])).replace("{S2}", _esc(t["h2"]))
    #               .replace("{CATS}", _str_lit(cats))
    #               .replace("{V1}", _num_lit([a for _f, a, _b in t["rows"]], t["fmt"]))
    #               .replace("{V2}", _num_lit([b for _f, _a, b in t["rows"]], t["fmt"])))
    #         out.append((cx, 6, max(t["thead"] - 2, 0), 9, 12))
    #     return out

    # _OP_CALC = {"cagr": (cm + 3, cm + 4, 9), "margin": (cm + 5, cm + 6, 9),
    #             "mult": (cm + 7, cm + 8, 217), "lev": (cm + 9, cm + 10, 217)}

    # def op_pivots(tab, pages, consolidated=False):
    #     root = ' grandTotalCaption="Consolidated"' if consolidated else ""
    #     hidden = tab["n_funds"] if tab["fund_blank"] else None
    #     out = []
    #     for t in tab["tables"]:
    #         f1, f2, nf = _OP_CALC[t["kind"]]
    #         dfs = [_df_xml(t["h1"], f1, nf), _df_xml(t["h2"], f2, nf)]
    #         out.append(_bx_axis_pivot_xml(
    #             nm(), t["top"], 2, tab["f_idx"], tab["n_funds"], hidden,
    #             pages, dfs, {f1, f2}, row_x=t["row_x"], extra_root=root))
    #     return out

    # op = ex["Op Performance"]
    # jobs.append(("Op Performance", op_pivots(op, op["pages"]),
    #              op_charts(op["tables"], "Op Performance")))

    # opu = ex["Op Performance - Unrealized"]
    # dfs = [_df_xml("Realized Value", real_idx, 3),
    #        _df_xml("Current Value", cur_idx, 3)]
    # opu_pivots = [_bx_axis_pivot_xml(
    #     nm(), opu["rv_top"], 2, opu["f_idx"], opu["n_funds"],
    #     opu["n_funds"] if opu["fund_blank"] else None,
    #     opu["pages_all"], dfs, {real_idx, cur_idx})]
    # opu_pivots += op_pivots(opu, opu["pages_unreal"], consolidated=True)
    # opu_pivots.append(_bx_axis_pivot_xml(
    #     nm(), opu["hp_top"], 2, opu["f_idx"], opu["n_funds"],
    #     opu["n_funds"] if opu["fund_blank"] else None,
    #     opu["pages_unreal"],
    #     [_df_xml("Sum of CalcWghtedHoldPeriod", cm + 11, 216)],
    #     {cm + 11}, row_x=opu["hp_row_x"],
    #     extra_root=' grandTotalCaption="Consolidated"'))
    # jobs.append(("Op Performance - Unrealized", opu_pivots,
    #              op_charts(opu["tables"], "Op Performance - Unrealized")))

    # ── Deployment & Exits ───────────────────────────────────────────────
    de = ex["Deployment & Exits"]
    L = de["_layout"]
    n_v, n_f = len(de["vintages"]), len(de["funds"])
    f_hidden = n_f if de["fund_blank"] else None
    rhc = ' rowHeaderCaption="Fund"'
    de_pivots = [
        _bx_matrix_xml(nm(), L[0]["top"], 3, de["v_idx"], n_v, None,
                       de["f_idx"], list(range(n_f)), f_hidden,
                       _df_xml("Sum of Total Invested Capital (mlns)", ic_idx, 9,
                               show_as="percentOfCol"),
                       {ic_idx}, grand_col=False, row_x=de["v_row_x"],
                       extra_root=rhc),
        _bx_matrix_xml(nm(), L[1]["top"], 3, de["v_idx"], n_v, None,
                       de["f_idx"], list(range(n_f)), f_hidden,
                       _df_xml("Count of Company", 0, 1, subtotal="count"),
                       {0}, grand_col=True, row_x=de["v_row_x"],
                       extra_root=rhc),
        _bx_matrix_xml(nm(), L[2]["top"], 4, de["f_idx"], n_f, f_hidden,
                       de["e_idx"], de["y_col_x"], de["y_hidden"],
                       _df_xml("Sum of Total IC mlns for Buckets", icb_idx, 9,
                               show_as="percentOfRow"),
                       {icb_idx}, pages=[de["status_page"]], grand_col=False,
                       row_x=de["f_row_x"], n_cols_all=len(de["exit_years"])),
        _bx_matrix_xml(nm(), L[3]["top"], 4, de["f_idx"], n_f, f_hidden,
                       de["e_idx"], de["y_col_x"], de["y_hidden"],
                       _df_xml("Count of Company", 0, 1, subtotal="count"),
                       {0}, pages=[de["status_page"]], grand_col=False,
                       row_x=de["f_row_x"], n_cols_all=len(de["exit_years"])),
    ]
    jobs.append(("Deployment & Exits", de_pivots, []))

    return jobs


def _rel_targets(rels_xml: str) -> dict[str, str]:
    out: dict[str, str] = {}
    for m in re.finditer(r"<Relationship\b[^>]*/>", rels_xml):
        tag = m.group(0)
        i = re.search(r'Id="([^"]+)"', tag)
        g = re.search(r'Target="([^"]+)"', tag)
        if i and g:
            out[i.group(1)] = g.group(1)
    return out


def _sheet_part(wb_xml: str, id2t: dict, sheet_name: str) -> str | None:
    esc_name = sheet_name.replace("&", "&amp;")
    for m in re.finditer(r"<sheet\b[^>]*/>", wb_xml):
        tag = m.group(0)
        if f'name="{esc_name}"' in tag or f'name="{sheet_name}"' in tag:
            rid = re.search(r'r:id="([^"]+)"', tag)
            if rid:
                tgt = id2t.get(rid.group(1), "")
                return tgt.lstrip("/") if tgt.startswith("/") else "xl/" + tgt
    return None


def _inject_pivots(wb_bytes: bytes, records: list[dict], plan: list[dict],
                   pc_plan: list[dict] | None = None,
                   rd_plan: list[dict] | None = None,
                   extra_jobs: list | None = None) -> bytes:
    cache_def, cache_rec, _maps = _build_cache_parts(records)
    z = zipfile.ZipFile(io.BytesIO(wb_bytes))

    # Resolve sheet parts (openpyxl writes Id last + absolute targets —
    # parse robustly).
    wb = z.read("xl/workbook.xml").decode("utf8")
    wr = z.read("xl/_rels/workbook.xml.rels").decode("utf8")
    id2t = _rel_targets(wr)
    rl_target = _sheet_part(wb, id2t, "Return & Loss Ratios")
    pc_target = _sheet_part(wb, id2t, "Portfolio Construction")
    rd_target = _sheet_part(wb, id2t, "Return Dispersion")
    if not rl_target:
        return wb_bytes

    # Pivot parts
    added: dict[str, bytes] = {
        "xl/pivotCache/pivotCacheDefinition1.xml": cache_def,
        "xl/pivotCache/pivotCacheRecords1.xml": cache_rec,
        "xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheRecords" '
            'Target="pivotCacheRecords1.xml"/></Relationships>').encode("utf8"),
    }
    # Sheet rel ids must look like plain "rId<N>" — Excel mishandles mixed
    # custom ids (e.g. rIdpt1) alongside a drawing rel on the same sheet.
    rl_rels_path_probe = rl_target.replace("worksheets/", "worksheets/_rels/") + ".rels"
    existing_sheet_rels = (z.read(rl_rels_path_probe).decode("utf8")
                           if rl_rels_path_probe in z.namelist() else "")
    next_rid = max((int(x) for x in re.findall(r'Id="rId(\d+)"', existing_sheet_rels)),
                   default=0) + 1
    sheet_rel_entries = []
    ct_overrides = [
        '<Override PartName="/xl/pivotCache/pivotCacheDefinition1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheDefinition+xml"/>',
        '<Override PartName="/xl/pivotCache/pivotCacheRecords1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheRecords+xml"/>',
    ]
    live = [p for p in plan if not p.get("empty")]
    for i, p in enumerate(live, start=1):
        part = f"xl/pivotTables/pivotTable{i}.xml"
        added[part] = _build_pivot_table_xml(p)
        added[f"xl/pivotTables/_rels/pivotTable{i}.xml.rels"] = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition" '
            'Target="../pivotCache/pivotCacheDefinition1.xml"/></Relationships>').encode("utf8")
        sheet_rel_entries.append(
            f'<Relationship Id="rId{next_rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable" '
            f'Target="../pivotTables/pivotTable{i}.xml"/>')
        next_rid += 1
        ct_overrides.append(
            f'<Override PartName="/{part}" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml"/>')

    # ── Charts: clone the template's combo chart per pivot ──────────────
    import pathlib
    _here = pathlib.Path(__file__).parent
    chart_tpl = (_here / "chart_template.xml").read_text()
    anchor_tpl = (_here / "drawing_anchor_template.xml").read_text()
    anchors, drawing_rels = [], []
    n_charts = 0
    for p in ([] if _DEBUG_NO_CHARTS else live):
        cats = p.get("chart_cats") or []
        if not cats:                             # no non-blank items → no chart
            continue
        n_charts += 1
        cx = chart_tpl
        cx = cx.replace("{CATS}", _str_lit(cats))
        cx = cx.replace("{VBAR}", _num_lit(p["chart_moic"], "0.0\\x;\\(0.0\\x\\)"))
        cx = cx.replace("{VLINE}", _num_lit(p["chart_loss"], "0%"))
        cx = cx.replace("{TITLE}", _esc(p["title"].split(" by ", 1)[-1]))
        added[f"xl/charts/chart{n_charts}.xml"] = cx.encode("utf8")
        ct_overrides.append(
            f'<Override PartName="/xl/charts/chart{n_charts}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>')
        drawing_rels.append(
            f'<Relationship Id="rId{n_charts}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart" '
            f'Target="../charts/chart{n_charts}.xml"/>')
        r0 = max(p["top"] - 2, 0)                   # 0-based top (≈ template offset)
        a = (anchor_tpl.replace("{C0}", str(_CHART_C0)).replace("{C1}", str(_CHART_C1))
                       .replace("{R0}", str(r0)).replace("{R1}", str(r0 + _CHART_ROWS))
                       .replace("{RID}", f"rId{n_charts}").replace("{FID}", str(n_charts + 1)))
        anchors.append(a)

    if n_charts:
        wsdr = ('<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
                'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
                'xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart">'
                + "".join(anchors) + "</xdr:wsDr>")
        added["xl/drawings/drawing1.xml"] = wsdr.encode("utf8")
        added["xl/drawings/_rels/drawing1.xml.rels"] = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            + "".join(drawing_rels) + "</Relationships>").encode("utf8")
        ct_overrides.append(
            '<Override PartName="/xl/drawings/drawing1.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>')
        drawing_rid = f"rId{next_rid}"
        next_rid += 1
        sheet_rel_entries.append(
            f'<Relationship Id="{drawing_rid}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
            'Target="../drawings/drawing1.xml"/>')
        rl_sheet_xml = z.read(rl_target).decode("utf8")
        if "xmlns:r=" not in rl_sheet_xml.split(">", 1)[0] + ">":
            rl_sheet_xml = rl_sheet_xml.replace(
                "<worksheet ",
                '<worksheet xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" ',
                1)
        rl_sheet_xml = rl_sheet_xml.replace(
            "</worksheet>", f'<drawing r:id="{drawing_rid}"/></worksheet>', 1)
        sheet_replacements = {rl_target: rl_sheet_xml.encode("utf8")}
    else:
        sheet_replacements = {}

    def _merge_sheet_rels(target: str, entries: list[str]) -> None:
        path = target.replace("worksheets/", "worksheets/_rels/") + ".rels"
        if path in z.namelist():
            existing = z.read(path).decode("utf8")
            added[path] = existing.replace(
                "</Relationships>", "".join(entries) + "</Relationships>").encode("utf8")
        else:
            added[path] = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                + "".join(entries) + "</Relationships>").encode("utf8")

    _merge_sheet_rels(rl_target, sheet_rel_entries)

    pc_pivots = ([p for p in pc_plan if p.get("type") in ("matrix", "count", "total")
                  and not (p.get("type") == "matrix"
                           and (not p.get("cols") or not p.get("funds")))]
                 if (pc_plan and pc_target) else [])
    rd_pivots = ([p for p in rd_plan if p.get("type") == "rd"]
                 if (rd_plan and rd_target) else [])

    # ── Portfolio Construction: its own pivots, charts, and drawing ─────
    if pc_pivots:
        pc_rel_entries: list[str] = []
        pc_probe = pc_target.replace("worksheets/", "worksheets/_rels/") + ".rels"
        pc_existing = z.read(pc_probe).decode("utf8") if pc_probe in z.namelist() else ""
        next_rid_pc = max((int(x) for x in re.findall(r'Id="rId(\d+)"', pc_existing)),
                          default=0) + 1
        pv_no = len(live)
        for p in pc_pivots:
            pv_no += 1
            part = f"xl/pivotTables/pivotTable{pv_no}.xml"
            builder = {"matrix": _build_pc_matrix_xml, "count": _build_pc_count_xml,
                       "total": _build_pc_total_xml}[p["type"]]
            added[part] = builder(p)
            added[f"xl/pivotTables/_rels/pivotTable{pv_no}.xml.rels"] = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition" '
                'Target="../pivotCache/pivotCacheDefinition1.xml"/></Relationships>').encode("utf8")
            pc_rel_entries.append(
                f'<Relationship Id="rId{next_rid_pc}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable" '
                f'Target="../pivotTables/pivotTable{pv_no}.xml"/>')
            next_rid_pc += 1
            ct_overrides.append(
                f'<Override PartName="/{part}" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml"/>')

        # Charts: percent-stacked column per matrix, pie per count pivot
        stack_frame = (_here / "chart_pc_stacked.xml").read_text()
        ser_tpl = (_here / "chart_pc_ser.xml").read_text()
        pie_tpl = (_here / "chart_pc_pie.xml").read_text()
        pc_anchors, pc_draw_rels = [], []
        pc_chart_no = 0
        for p in ([] if _DEBUG_NO_CHARTS else pc_pivots):
            top = p["top"]
            if p["type"] == "matrix":
                n_vis = len(p["cols"])
                if n_vis == 0 or not p["matrix_rows"]:
                    continue
                cat_lit = _str_lit([f for f, _pcts in p["matrix_rows"]])
                sers = []
                for k, cv in enumerate(p["cols"]):
                    vals = [pcts[k] for _f, pcts in p["matrix_rows"]]
                    sers.append(ser_tpl
                                .replace("{IDX}", str(k)).replace("{ORDER}", str(k))
                                .replace("{TXF}", f"<c:v>{_esc(str(cv))}</c:v>")
                                .replace("{FILL}", _PC_STACK_FILLS[k % len(_PC_STACK_FILLS)])
                                .replace("{CATF}", cat_lit)
                                .replace("{VALF}", _num_lit(vals, '0%;\\(0%\\);"-"')))
                cxml = stack_frame.replace("{SERIES}", "".join(sers))
                c0 = max(10, 4 + len(p["cols"]) + 1)
                r0 = top - 1
            elif p["type"] == "count":
                if not p["items"]:
                    continue
                cxml = (pie_tpl
                        .replace("{TXF}", "<c:v>Count of Company</c:v>")
                        .replace("{CATS}", _str_lit(p["items"]))
                        .replace("{VALS}", _num_lit([p["counts"][i] for i in p["items"]], "0")))
                c0 = 5
                r0 = max(top - 3, 0)
            else:
                continue                          # total-count pivot: no chart
            pc_chart_no += 1
            n_charts += 1
            added[f"xl/charts/chart{n_charts}.xml"] = cxml.encode("utf8")
            ct_overrides.append(
                f'<Override PartName="/xl/charts/chart{n_charts}.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>')
            pc_draw_rels.append(
                f'<Relationship Id="rId{pc_chart_no}" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart" '
                f'Target="../charts/chart{n_charts}.xml"/>')
            width = 9 if p["type"] == "matrix" else 4
            pc_anchors.append(
                anchor_tpl.replace("{C0}", str(c0)).replace("{C1}", str(c0 + width))
                          .replace("{R0}", str(r0)).replace("{R1}", str(r0 + _PC_CHART_ROWS))
                          .replace("{RID}", f"rId{pc_chart_no}").replace("{FID}", str(pc_chart_no + 1)))

        if pc_chart_no:
            wsdr2 = ('<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
                     'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
                     'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
                     'xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart">'
                     + "".join(pc_anchors) + "</xdr:wsDr>")
            added["xl/drawings/drawing2.xml"] = wsdr2.encode("utf8")
            added["xl/drawings/_rels/drawing2.xml.rels"] = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                + "".join(pc_draw_rels) + "</Relationships>").encode("utf8")
            ct_overrides.append(
                '<Override PartName="/xl/drawings/drawing2.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>')
            pc_drawing_rid = f"rId{next_rid_pc}"
            next_rid_pc += 1
            pc_rel_entries.append(
                f'<Relationship Id="{pc_drawing_rid}" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
                'Target="../drawings/drawing2.xml"/>')
            pc_sheet_xml = z.read(pc_target).decode("utf8")
            if "xmlns:r=" not in pc_sheet_xml.split(">", 1)[0] + ">":
                pc_sheet_xml = pc_sheet_xml.replace(
                    "<worksheet ",
                    '<worksheet xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" ',
                    1)
            pc_sheet_xml = pc_sheet_xml.replace(
                "</worksheet>", f'<drawing r:id="{pc_drawing_rid}"/></worksheet>', 1)
            sheet_replacements[pc_target] = pc_sheet_xml.encode("utf8")

        _merge_sheet_rels(pc_target, pc_rel_entries)

    # ── Return Dispersion: two dispersion pivots + one chart each ────────
    if rd_pivots:
        rd_rel_entries: list[str] = []
        rd_probe = rd_target.replace("worksheets/", "worksheets/_rels/") + ".rels"
        rd_existing = z.read(rd_probe).decode("utf8") if rd_probe in z.namelist() else ""
        next_rid_rd = max((int(x) for x in re.findall(r'Id="rId(\d+)"', rd_existing)),
                          default=0) + 1
        pv_no = len(live) + len(pc_pivots)
        rd_tpl = (_here / "chart_rd.xml").read_text()
        rd_anchors, rd_draw_rels = [], []
        rd_chart_no = 0
        for p in rd_pivots:
            pv_no += 1
            part = f"xl/pivotTables/pivotTable{pv_no}.xml"
            added[part] = _build_rd_pivot_xml(p)
            added[f"xl/pivotTables/_rels/pivotTable{pv_no}.xml.rels"] = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition" '
                'Target="../pivotCache/pivotCacheDefinition1.xml"/></Relationships>').encode("utf8")
            rd_rel_entries.append(
                f'<Relationship Id="rId{next_rid_rd}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable" '
                f'Target="../pivotTables/pivotTable{pv_no}.xml"/>')
            next_rid_rd += 1
            ct_overrides.append(
                f'<Override PartName="/{part}" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml"/>')

            # charts plot the % IC distribution over the real buckets — the
            # "n/a" bucket (no reported MOIC/IRR) stays in the pivot only
            vis = [(it, cnt, pct) for it, cnt, pct, _avg in p["rows"] if it != "n/a"]
            if _DEBUG_NO_CHARTS or not vis:
                continue
            rd_chart_no += 1
            n_charts += 1
            cats = [f"{it}\n{cnt}" for it, cnt, _pct in vis]
            pcts = [pct for _it, _cnt, pct in vis]
            cxml = (rd_tpl.replace("{CATS}", _str_lit(cats))
                          .replace("{VALS}", _num_lit(pcts, "0%")))
            added[f"xl/charts/chart{n_charts}.xml"] = cxml.encode("utf8")
            ct_overrides.append(
                f'<Override PartName="/xl/charts/chart{n_charts}.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>')
            rd_draw_rels.append(
                f'<Relationship Id="rId{rd_chart_no}" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart" '
                f'Target="../charts/chart{n_charts}.xml"/>')
            r0 = max(p["top"] - 2, 0)
            rd_anchors.append(
                anchor_tpl.replace("{C0}", "7").replace("{C1}", "16")
                          .replace("{R0}", str(r0)).replace("{R1}", str(r0 + _CHART_ROWS))
                          .replace("{RID}", f"rId{rd_chart_no}").replace("{FID}", str(rd_chart_no + 1)))

        if rd_chart_no:
            wsdr3 = ('<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
                     'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
                     'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
                     'xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart">'
                     + "".join(rd_anchors) + "</xdr:wsDr>")
            added["xl/drawings/drawing3.xml"] = wsdr3.encode("utf8")
            added["xl/drawings/_rels/drawing3.xml.rels"] = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                + "".join(rd_draw_rels) + "</Relationships>").encode("utf8")
            ct_overrides.append(
                '<Override PartName="/xl/drawings/drawing3.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>')
            rd_drawing_rid = f"rId{next_rid_rd}"
            next_rid_rd += 1
            rd_rel_entries.append(
                f'<Relationship Id="{rd_drawing_rid}" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
                'Target="../drawings/drawing3.xml"/>')
            rd_sheet_xml = z.read(rd_target).decode("utf8")
            if "xmlns:r=" not in rd_sheet_xml.split(">", 1)[0] + ">":
                rd_sheet_xml = rd_sheet_xml.replace(
                    "<worksheet ",
                    '<worksheet xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" ',
                    1)
            rd_sheet_xml = rd_sheet_xml.replace(
                "</worksheet>", f'<drawing r:id="{rd_drawing_rid}"/></worksheet>', 1)
            sheet_replacements[rd_target] = rd_sheet_xml.encode("utf8")

        _merge_sheet_rels(rd_target, rd_rel_entries)

    # ── Extra analysis tabs: generic per-sheet pivots + charts ──────────
    pv_no = len(live) + len(pc_pivots) + len(rd_pivots)
    drawing_no = 3
    for sheet_name, sheet_pivots, sheet_charts in (extra_jobs or []):
        target = _sheet_part(wb, id2t, sheet_name)
        if not target:
            continue
        entries: list[str] = []
        probe = target.replace("worksheets/", "worksheets/_rels/") + ".rels"
        existing_x = z.read(probe).decode("utf8") if probe in z.namelist() else ""
        next_rid_x = max((int(x) for x in re.findall(r'Id="rId(\d+)"', existing_x)),
                         default=0) + 1
        for xml in sheet_pivots:
            if xml is None:                       # degenerate pivot skipped
                continue
            pv_no += 1
            part = f"xl/pivotTables/pivotTable{pv_no}.xml"
            added[part] = xml
            added[f"xl/pivotTables/_rels/pivotTable{pv_no}.xml.rels"] = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition" '
                'Target="../pivotCache/pivotCacheDefinition1.xml"/></Relationships>').encode("utf8")
            entries.append(
                f'<Relationship Id="rId{next_rid_x}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable" '
                f'Target="../pivotTables/pivotTable{pv_no}.xml"/>')
            next_rid_x += 1
            ct_overrides.append(
                f'<Override PartName="/{part}" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml"/>')

        if sheet_charts and not _DEBUG_NO_CHARTS:
            drawing_no += 1
            x_anchors, x_rels = [], []
            for k, (cxml, c0, r0, w, h) in enumerate(sheet_charts, start=1):
                n_charts += 1
                added[f"xl/charts/chart{n_charts}.xml"] = cxml.encode("utf8")
                ct_overrides.append(
                    f'<Override PartName="/xl/charts/chart{n_charts}.xml" '
                    'ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>')
                x_rels.append(
                    f'<Relationship Id="rId{k}" '
                    'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart" '
                    f'Target="../charts/chart{n_charts}.xml"/>')
                x_anchors.append(
                    anchor_tpl.replace("{C0}", str(c0)).replace("{C1}", str(c0 + w))
                              .replace("{R0}", str(r0)).replace("{R1}", str(r0 + h))
                              .replace("{RID}", f"rId{k}").replace("{FID}", str(k + 1)))
            added[f"xl/drawings/drawing{drawing_no}.xml"] = (
                '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
                'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
                'xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart">'
                + "".join(x_anchors) + "</xdr:wsDr>").encode("utf8")
            added[f"xl/drawings/_rels/drawing{drawing_no}.xml.rels"] = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                + "".join(x_rels) + "</Relationships>").encode("utf8")
            ct_overrides.append(
                f'<Override PartName="/xl/drawings/drawing{drawing_no}.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>')
            x_drawing_rid = f"rId{next_rid_x}"
            next_rid_x += 1
            entries.append(
                f'<Relationship Id="{x_drawing_rid}" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
                'Target="../drawings/drawing%d.xml"/>' % drawing_no)
            sx = z.read(target).decode("utf8")
            if "xmlns:r=" not in sx.split(">", 1)[0] + ">":
                sx = sx.replace(
                    "<worksheet ",
                    '<worksheet xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" ',
                    1)
            sx = sx.replace("</worksheet>",
                            f'<drawing r:id="{x_drawing_rid}"/></worksheet>', 1)
            sheet_replacements[target] = sx.encode("utf8")

        if entries:
            _merge_sheet_rels(target, entries)

    # workbook: pivotCaches must come near the END of workbook.xml (child order).
    # xmlns:r is declared INLINE on the element: openpyxl (≥3.1) only declares
    # it locally on <sheet>, not on the <workbook> root, so a bare r: prefix
    # here is an unbound-namespace hard corrupt (Excel refuses the file).
    max_rid = max((int(x) for x in re.findall(r'Id="rId(\d+)"', wr)), default=0)
    cache_rid = f"rId{max_rid + 1}"
    wb = wb.replace(
        "</workbook>",
        f'<pivotCaches><pivotCache cacheId="1" '
        f'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        f'r:id="{cache_rid}"/></pivotCaches></workbook>', 1)
    wr = wr.replace(
        "</Relationships>",
        f'<Relationship Id="{cache_rid}" '
        f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition" '
        f'Target="pivotCache/pivotCacheDefinition1.xml"/></Relationships>')

    ct = z.read("[Content_Types].xml").decode("utf8")
    ct = ct.replace("</Types>", "".join(ct_overrides) + "</Types>")

    # Custom number formats referenced by dataField numFmtId: 217 = MOIC
    # multiples (2.0x / (0.4x)), 192 = matrix row-percentages (0%;(0%);"-").
    st = z.read("xl/styles.xml").decode("utf8")
    for fmt_id, code in (("217", "0.0\\x;\\(0.0\\x\\)"),
                         ("192", "0%;\\(0%\\);&quot;-&quot;"),
                         ("218", "0.0%"),
                         ("216", "0.0\\ &quot;yrs&quot;")):
        if f'numFmtId="{fmt_id}"' not in st:
            fmt = f'<numFmt numFmtId="{fmt_id}" formatCode="{code}"/>'
            m = re.search(r'<numFmts count="(\d+)">', st)
            if m:
                st = st.replace(m.group(0), f'<numFmts count="{int(m.group(1)) + 1}">' + fmt, 1)
            else:
                st = re.sub(r"(<styleSheet[^>]*>)", r"\1" + f'<numFmts count="1">{fmt}</numFmts>', st, count=1)

    replaced = {
        "xl/workbook.xml": wb.encode("utf8"),
        "xl/_rels/workbook.xml.rels": wr.encode("utf8"),
        "[Content_Types].xml": ct.encode("utf8"),
        "xl/styles.xml": st.encode("utf8"),
    }
    replaced.update(sheet_replacements)

    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zn:
        for item in z.infolist():
            data = replaced.get(item.filename)
            zn.writestr(item, data if data is not None else z.read(item.filename))
        for name, data in added.items():
            zn.writestr(name, data)
    return out.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# Public API
# ═══════════════════════════════════════════════════════════════════════════════

def build_inputs_workbook(records: list[dict], gp_name: str,
                          currency: str = "USD",
                          track_record_date: date | None = None) -> bytes:
    """Standalone "Gross Deal Level Input" workbook: just the Deal Level
    Inputs sheet, ready to be imported into the VBA analyzer template."""
    for rec in records:                        # same EWL defaults as build_output
        if not _cell_str(rec.get(90)):
            rec[90] = currency
        if _cell_num(rec.get(17)) is None:
            rec[17] = 0.0
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deal Level Inputs"
    _write_inputs(ws, records, gp_name, currency, track_record_date)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_output(records: list[dict], gp_name: str, currency: str = "USD",
                 phase_errors: list | None = None,
                 track_record_date: date | None = None) -> bytes:
    """Build the 3-tab output workbook (values + linking formulas + real pivots)."""
    errs = phase_errors if phase_errors is not None else []
    for rec in records:                        # EWL defaults
        if not _cell_str(rec.get(90)):
            rec[90] = currency                 # Fund Currency
        if _cell_num(rec.get(17)) is None:
            rec[17] = 0.0                      # Realized Value: 0 if missing
    wb = openpyxl.Workbook()
    ws_in = wb.active; ws_in.title = "Deal Level Inputs"
    wb.create_sheet("Table of Contents", 0)
    ws_dl = wb.create_sheet("Deal List")
    ws_rl = wb.create_sheet("Return & Loss Ratios")
    ws_rd = wb.create_sheet("Return Dispersion")
    ws_pc = wb.create_sheet("Portfolio Construction")
    ws_vs = wb.create_sheet("Vintage Perf by Sector")
    # ws_ua = wb.create_sheet("Underperforming Assets")   # EWL: not needed yet
    # ws_pa = wb.create_sheet("Partner Attribution")      # EWL: not needed yet
    # ws_op = wb.create_sheet("Op Performance")           # EWL: not needed yet
    # ws_ou = wb.create_sheet("Op Performance - Unrealized")  # EWL: not needed yet
    ws_de = wb.create_sheet("Deployment & Exits")
    wb.active = 0

    plan = plan_pivots(records)
    rd_plan = plan_rd(records)
    pc_plan = plan_pc(records)
    ex = plan_extra(records)
    try:
        _write_inputs(ws_in, records, gp_name, currency, track_record_date)
        _write_deal_list(ws_dl, records, gp_name, track_record_date, currency)
        _write_rl_titles(ws_rl, plan)
        _render_pivot_cells(ws_rl, plan, records)
        _write_rl_graphics(ws_rl, plan)
        _write_rd_sheet(ws_rd, rd_plan)
        _write_pc_sheet(ws_pc, pc_plan, records)
        _write_vs_sheet(ws_vs, ex["Vintage Perf by Sector"])
        # _write_ua_sheet(ws_ua, ex["Underperforming Assets"])   # EWL
        # _write_pa_sheet(ws_pa, ex["Partner Attribution"])      # EWL
        # _write_op_sheet(ws_op, ex["Op Performance"])           # EWL
        # _write_opu_sheet(ws_ou, ex["Op Performance - Unrealized"])  # EWL
        _write_de_sheet(ws_de, ex["Deployment & Exits"])
        _write_toc(wb)
    except Exception as e:
        errs.append({"phase": "Build workbook", "detail": str(e), "row": ""})

    wb.calculation.fullCalcOnLoad = True
    buf = io.BytesIO(); wb.save(buf)

    try:
        return _inject_pivots(buf.getvalue(), records, plan, pc_plan, rd_plan,
                              _extra_jobs(ex))
    except Exception as e:
        errs.append({"phase": "Pivot injection", "detail": str(e), "row": ""})
        return buf.getvalue()